"""Profitability Report — project-level Excel export.

Columns per project row:
  Project | Billing (Revenue) | Total Hours | Billable Hours | Utilization %
  | Manpower Cost | Direct Expenses | Indirect / Overhead | Total Cost
  | Net Profit / Loss | Net Margin % | Billing Recovery %

Formulas:
  Manpower Cost      = SUM(work_hours.hours_spent × user.cost_rate) per project
  Direct Expenses    = SUM(project_costs.cost) WHERE category != 'Indirect / Overhead'
  Indirect / Overhead= SUM(project_costs.cost) WHERE category = 'Indirect / Overhead'
  Total Cost         = Manpower Cost + Direct Expenses + Indirect / Overhead
  Net Profit / Loss  = Billing Amount − Total Cost
  Net Margin %       = Net Profit / Billing Amount × 100   (blank if Billing = 0)
  Billing Recovery % = Billing Amount / Total Cost × 100   (blank if Total Cost = 0)
  Utilization %      = Billable Hours / Total Hours × 100  (blank if Total Hours = 0)
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.database import get_db
from app.models.models import Project, WorkHours, ProjectCost, User, ProjectBilling
from app.core.deps import get_current_user

router = APIRouter(prefix="/profitability-report", tags=["Profitability Report"])

INDIRECT_CATEGORY = "Indirect / Overhead"


# ── xlsx builder ─────────────────────────────────────────────────────────────
def _build_xlsx(title: str, headers: list, rows: list, filename: str, subtitle: str = ""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
    def bdr():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    HDR_FILL  = fill("1F3864")
    COL_FILL  = fill("BDD7EE")
    EVEN_FILL = fill("EBF3FB")
    ODD_FILL  = fill("FFFFFF")
    PROFIT_FILL = fill("E2EFDA")   # green tint for profit rows
    LOSS_FILL   = fill("FCE4D6")   # red tint for loss rows

    ncols = len(headers)
    last_col = chr(ord("A") + ncols - 1)

    # Title row
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.fill = HDR_FILL; c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    if subtitle:
        ws.merge_cells(f"A2:{last_col}2")
        c2 = ws["A2"]
        c2.value = subtitle
        c2.font = Font(italic=True, color="555555", size=9, name="Calibri")
        c2.fill = fill("D9E8F5"); c2.alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 16
        header_row = 4
    else:
        header_row = 3

    for col, h in enumerate(headers, 1):
        hc = ws.cell(header_row, col, h)
        hc.font = Font(bold=True, color="1F3864", size=9, name="Calibri")
        hc.fill = COL_FILL; hc.border = bdr()
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 28

    row_num = header_row + 1
    for i, r in enumerate(rows):
        # Detect profit/loss from Net Profit column (index 9)
        net = r[9] if len(r) > 9 else None
        if isinstance(net, (int, float)):
            bg = PROFIT_FILL if net >= 0 else LOSS_FILL
        else:
            bg = EVEN_FILL if i % 2 == 0 else ODD_FILL

        for col, val in enumerate(r, 1):
            cell = ws.cell(row_num, col, val if val not in (None, "") else "—")
            cell.font = Font(size=9, name="Calibri")
            cell.fill = bg; cell.border = bdr()
            # Right-align numeric columns
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

    if not rows:
        ws.merge_cells(f"A{row_num}:{last_col}{row_num}")
        ws.cell(row_num, 1, "No projects found for the selected filters")
        ws.cell(row_num, 1).font = Font(italic=True, size=9, color="999999")

    # Column widths
    col_widths = [28, 16, 12, 14, 13, 16, 16, 18, 13, 16, 13, 16]
    for i, w in enumerate(col_widths[:ncols], 1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w

    output = BytesIO()
    wb.save(output); output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


def _fmt(v, decimals=2):
    """Round float to decimals; return as float (openpyxl handles numeric cells)."""
    if v is None: return 0.0
    return round(float(v), decimals)

def _pct(numerator, denominator, decimals=1):
    """Safe percentage string. Returns float or '' if denominator is 0."""
    if not denominator: return ""
    return round(float(numerator) / float(denominator) * 100, decimals)


# ── Core calculation helper ───────────────────────────────────────────────────
def _project_metrics(db: Session, project: Project):
    """Compute all profitability metrics for one project."""
    pid = project.id

    # ── Hours ────────────────────────────────────────────────────────────────
    total_hours = db.query(func.coalesce(func.sum(WorkHours.hours_spent), 0.0))\
        .filter(WorkHours.project_id == pid).scalar() or 0.0

    billable_hours = db.query(func.coalesce(func.sum(WorkHours.hours_spent), 0.0))\
        .filter(WorkHours.project_id == pid, WorkHours.work_type == 'Billable').scalar() or 0.0

    # ── Manpower Cost = SUM(hours × user.cost_rate) ──────────────────────────
    # Join WorkHours → User to get each entry's rate, then aggregate.
    manpower_cost = 0.0
    wh_rows = db.query(WorkHours.hours_spent, WorkHours.user_id)\
        .filter(WorkHours.project_id == pid).all()
    user_rate_cache = {}
    for hours, uid in wh_rows:
        if uid not in user_rate_cache:
            u = db.query(User.cost_rate).filter_by(id=uid).first()
            user_rate_cache[uid] = float(u.cost_rate or 0) if u else 0.0
        manpower_cost += float(hours or 0) * user_rate_cache[uid]

    # ── Cost entries ─────────────────────────────────────────────────────────
    all_costs = db.query(ProjectCost).filter_by(project_id=pid).all()
    direct_expenses  = sum(c.cost for c in all_costs if c.category != INDIRECT_CATEGORY)
    indirect_cost    = sum(c.cost for c in all_costs if c.category == INDIRECT_CATEGORY)

    # ── Derived KPIs ─────────────────────────────────────────────────────────
    total_cost   = manpower_cost + direct_expenses + indirect_cost
    # Billing = SUM of all ProjectBilling entries for this project.
    # Falls back to 0 if no entries recorded yet.
    billing = float(
        db.query(func.coalesce(func.sum(ProjectBilling.amount), 0.0))
          .filter(ProjectBilling.project_id == pid).scalar() or 0.0
    )
    net_profit   = billing - total_cost
    margin_pct   = _pct(net_profit, billing)
    recovery_pct = _pct(billing, total_cost)
    util_pct     = _pct(billable_hours, total_hours)

    return {
        "billing":         _fmt(billing),
        "total_hours":     _fmt(total_hours),
        "billable_hours":  _fmt(billable_hours),
        "util_pct":        util_pct,
        "manpower_cost":   _fmt(manpower_cost),
        "direct_expenses": _fmt(direct_expenses),
        "indirect_cost":   _fmt(indirect_cost),
        "total_cost":      _fmt(total_cost),
        "net_profit":      _fmt(net_profit),
        "margin_pct":      margin_pct,
        "recovery_pct":    recovery_pct,
    }


# ── Report endpoint ───────────────────────────────────────────────────────────
@router.get("/export")
def profitability_report_export(
    project_id:  Optional[int] = None,
    status:      Optional[str] = None,
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(Project)
    if project_id:
        q = q.filter(Project.id == project_id)
    if status:
        q = q.filter(Project.status == status)
    projects = q.order_by(Project.name).all()

    headers = [
        "Project", "Billing (Revenue ₹)", "Total Hours", "Billable Hours",
        "Utilization %", "Manpower Cost (₹)", "Direct Expenses (₹)",
        "Indirect / Overhead (₹)", "Total Cost (₹)", "Net Profit / Loss (₹)",
        "Net Margin %", "Billing Recovery %",
    ]

    rows = []
    for p in projects:
        m = _project_metrics(db, p)
        rows.append([
            p.name,
            m["billing"],
            m["total_hours"],
            m["billable_hours"],
            m["util_pct"] if m["util_pct"] != "" else "—",
            m["manpower_cost"],
            m["direct_expenses"],
            m["indirect_cost"],
            m["total_cost"],
            m["net_profit"],
            m["margin_pct"] if m["margin_pct"] != "" else "—",
            m["recovery_pct"] if m["recovery_pct"] != "" else "—",
        ])

    subtitle_parts = []
    if start_date or end_date:
        subtitle_parts.append(f"Period: {start_date or '—'} → {end_date or '—'}")
    if status:
        subtitle_parts.append(f"Status: {status}")

    return _build_xlsx(
        "Profitability Report", headers, rows,
        "profitability-report.xlsx",
        " | ".join(subtitle_parts),
    )


# ── Filter options ────────────────────────────────────────────────────────────
@router.get("/filter-options")
def filter_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    projects = db.query(Project).order_by(Project.name).all()
    statuses = (
        db.query(Project.status)
        .filter(Project.status.isnot(None))
        .distinct().order_by(Project.status).all()
    )
    return {
        "projects": [{"id": p.id, "name": p.name} for p in projects],
        "statuses": [s[0] for s in statuses],
    }
