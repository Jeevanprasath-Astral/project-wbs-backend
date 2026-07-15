"""Billing Statement Report — per-project billing history with running totals.

Sheet 1 — Summary (one row per project that has billing entries):
  Project | Client | Status | Total Billed (₹) | # Entries
  | Most Recent Billing | Billing Types Used

Sheet 2 — All Entries (one row per billing entry, grouped by project):
  Project | Client | Date | Billing Type | Amount (₹)
  | Running Total (₹) | Description | Milestone | Remarks
"""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.database import get_db
from app.models.models import ProjectBilling, Project, CustomMilestone, User
from app.core.deps import get_current_user
from app.api.routes.project_billings import BILLING_TYPES

router = APIRouter(prefix="/billing-statement-report", tags=["Billing Statement Report"])


# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL  = _fill("1F3864")
COL_FILL  = _fill("BDD7EE")
EVEN_FILL = _fill("EBF3FB")
ODD_FILL  = _fill("FFFFFF")
TOT_FILL  = _fill("D9E8F5")
PROJ_FILL = _fill("E8F4FD")   # light blue for project header rows in Sheet 2
AMT_FILL  = _fill("E2EFDA")   # green tint for amount / running total cells


def _write_title(ws, title, subtitle, ncols):
    last_col = chr(ord("A") + ncols - 1)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    if subtitle:
        ws.merge_cells(f"A2:{last_col}2")
        c2 = ws["A2"]
        c2.value = subtitle
        c2.font = Font(italic=True, color="555555", size=9, name="Calibri")
        c2.fill = _fill("D9E8F5")
        c2.alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 16
        return 4
    return 3


def _write_header(ws, headers, hrow):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(hrow, col, h)
        cell.font = Font(bold=True, color="1F3864", size=9, name="Calibri")
        cell.fill = COL_FILL
        cell.border = _bdr()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hrow].height = 28


def _write_row(ws, row_num, values, fill, right_cols=None, bold=False):
    right_cols = right_cols or []
    for col, val in enumerate(values, 1):
        display = val if val not in (None, "") else "—"
        cell = ws.cell(row_num, col, display)
        cell.font = Font(size=9, name="Calibri", bold=bold)
        cell.fill = fill
        cell.border = _bdr()
        align = "right" if (isinstance(val, (int, float)) or col in right_cols) else "left"
        cell.alignment = Alignment(horizontal=align)


def _fmt(v, dec=2):
    if v is None:
        return 0.0
    return round(float(v), dec)


# ── Export ────────────────────────────────────────────────────────────────────
@router.get("/export")
def billing_statement_export(
    project_id:   Optional[int] = None,
    start_date:   Optional[str] = None,
    end_date:     Optional[str] = None,
    billing_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import date as date_cls

    # ── Fetch billing entries ─────────────────────────────────────────────────
    q = db.query(ProjectBilling)
    if project_id:
        q = q.filter(ProjectBilling.project_id == project_id)
    if start_date:
        q = q.filter(ProjectBilling.date >= date_cls.fromisoformat(start_date))
    if end_date:
        q = q.filter(ProjectBilling.date <= date_cls.fromisoformat(end_date))
    if billing_type:
        q = q.filter(ProjectBilling.billing_type == billing_type)

    # Order by project_id then date for grouping + running total
    entries = q.order_by(ProjectBilling.project_id, ProjectBilling.date).all()

    # ── Load related data ─────────────────────────────────────────────────────
    proj_ids = list({e.project_id for e in entries})
    proj_map = {p.id: p for p in db.query(Project).filter(Project.id.in_(proj_ids)).all()}

    milestone_cache = {}  # mid -> name
    for e in entries:
        if e.milestone_id and e.milestone_id not in milestone_cache:
            m = db.query(CustomMilestone.name).filter_by(id=e.milestone_id).first()
            milestone_cache[e.milestone_id] = m.name if m else None

    # ── Group entries by project ──────────────────────────────────────────────
    proj_entries = {}   # {pid: [entry, ...]}
    for e in entries:
        proj_entries.setdefault(e.project_id, []).append(e)

    # ── Subtitle ──────────────────────────────────────────────────────────────
    parts = []
    if start_date or end_date:
        parts.append(f"Period: {start_date or '—'} → {end_date or '—'}")
    if billing_type:
        parts.append(f"Type: {billing_type}")
    subtitle = " | ".join(parts)

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    HEADERS_S = [
        "Project", "Client", "Status",
        "Total Billed (₹)", "# Entries",
        "Most Recent Billing", "Billing Types Used",
    ]
    hrow = _write_title(ws1, "Billing Statement — Project Summary", subtitle, len(HEADERS_S))
    _write_header(ws1, HEADERS_S, hrow)

    grand_total = 0.0
    grand_count = 0
    row_num = hrow + 1

    for i, pid in enumerate(sorted(proj_ids, key=lambda x: proj_map.get(x, Project()).name or "")):
        p = proj_map.get(pid)
        p_entries = proj_entries.get(pid, [])
        total = sum(float(e.amount or 0) for e in p_entries)
        count = len(p_entries)
        most_recent = max((str(e.date) for e in p_entries if e.date), default="—")
        types_used = ", ".join(sorted({e.billing_type for e in p_entries if e.billing_type}))

        grand_total += total
        grand_count += count

        bg = EVEN_FILL if i % 2 == 0 else ODD_FILL
        _write_row(ws1, row_num, [
            p.name if p else "—",
            p.client if p else "—",
            p.status if p else "—",
            _fmt(total), count,
            most_recent, types_used or "—",
        ], bg, right_cols=[4, 5])
        row_num += 1

    # Grand total row
    if proj_ids:
        _write_row(ws1, row_num, [
            "TOTAL", "", "",
            _fmt(grand_total), grand_count, "", "",
        ], TOT_FILL, right_cols=[4, 5], bold=True)
        row_num += 1

    if not proj_ids:
        ws1.merge_cells(f"A{row_num}:G{row_num}")
        ws1.cell(row_num, 1, "No billing entries found for the selected filters")
        ws1.cell(row_num, 1).font = Font(italic=True, size=9, color="999999")

    # Column widths Sheet 1
    for i, w in enumerate([30, 20, 14, 18, 10, 18, 40], 1):
        ws1.column_dimensions[chr(ord("A") + i - 1)].width = w

    # ── Sheet 2: All Entries ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Entries")

    HEADERS_E = [
        "Project", "Client", "Date", "Billing Type",
        "Amount (₹)", "Running Total (₹)",
        "Description", "Milestone", "Remarks",
    ]
    hrow2 = _write_title(ws2, "Billing Statement — All Entries", subtitle, len(HEADERS_E))
    _write_header(ws2, HEADERS_E, hrow2)

    row_num2 = hrow2 + 1
    any_rows = False

    sorted_pids = sorted(proj_ids, key=lambda x: proj_map.get(x, Project()).name or "")

    for pid in sorted_pids:
        p = proj_map.get(pid)
        p_entries = proj_entries.get(pid, [])
        running = 0.0

        for j, e in enumerate(p_entries):
            amt = float(e.amount or 0)
            running += amt
            m_name = milestone_cache.get(e.milestone_id) if e.milestone_id else None

            bg = EVEN_FILL if j % 2 == 0 else ODD_FILL

            row_vals = [
                p.name if p else "—",
                p.client if p else "—",
                str(e.date) if e.date else "—",
                e.billing_type or "—",
                _fmt(amt),
                _fmt(running),
                e.description or "—",
                m_name or "—",
                e.remarks or "—",
            ]
            _write_row(ws2, row_num2, row_vals, bg, right_cols=[5, 6])

            # Highlight amount and running total columns green
            ws2.cell(row_num2, 5).fill = AMT_FILL
            ws2.cell(row_num2, 6).fill = AMT_FILL
            ws2.cell(row_num2, 5).font = Font(size=9, name="Calibri", bold=True, color="1A7A3E")
            ws2.cell(row_num2, 6).font = Font(size=9, name="Calibri", bold=True, color="1A7A3E")

            row_num2 += 1
            any_rows = True

        # Project subtotal row
        if p_entries:
            proj_total_val = sum(float(e.amount or 0) for e in p_entries)
            _write_row(ws2, row_num2, [
                f"  {p.name if p else '—'} — Total", "", "", "",
                _fmt(proj_total_val), "", "", "", "",
            ], PROJ_FILL, right_cols=[5], bold=True)
            row_num2 += 1

    if not any_rows:
        ws2.merge_cells(f"A{row_num2}:I{row_num2}")
        ws2.cell(row_num2, 1, "No billing entries found for the selected filters")
        ws2.cell(row_num2, 1).font = Font(italic=True, size=9, color="999999")

    # Column widths Sheet 2
    for i, w in enumerate([28, 18, 12, 20, 16, 18, 30, 25, 25], 1):
        ws2.column_dimensions[chr(ord("A") + i - 1)].width = w

    # ── Sheet 3: Month-wise Summary ───────────────────────────────────────────
    ws3 = wb.create_sheet("Month-wise Summary")

    HEADERS_M = ["Month", "Total Billed (₹)", "# Entries", "Projects"]
    hrow3 = _write_title(ws3, "Billing Statement — Month-wise Summary", subtitle, len(HEADERS_M))
    _write_header(ws3, HEADERS_M, hrow3)

    # Group all billing entries by YYYY-MM
    from collections import defaultdict
    month_data = defaultdict(lambda: {"total": 0.0, "count": 0, "projects": set()})
    for e in entries:
        if e.date:
            key = e.date.strftime("%Y-%m")
            month_data[key]["total"]   += float(e.amount or 0)
            month_data[key]["count"]   += 1
            p = proj_map.get(e.project_id)
            if p:
                month_data[key]["projects"].add(p.name)

    row_num3 = hrow3 + 1
    grand_m_total = 0.0
    grand_m_count = 0

    for i, month_key in enumerate(sorted(month_data.keys())):
        md = month_data[month_key]
        grand_m_total += md["total"]
        grand_m_count += md["count"]
        bg = EVEN_FILL if i % 2 == 0 else ODD_FILL
        _write_row(ws3, row_num3, [
            month_key,
            _fmt(md["total"]),
            md["count"],
            ", ".join(sorted(md["projects"])) or "—",
        ], bg, right_cols=[2, 3])
        row_num3 += 1

    if month_data:
        _write_row(ws3, row_num3, [
            "TOTAL", _fmt(grand_m_total), grand_m_count, "",
        ], TOT_FILL, right_cols=[2, 3], bold=True)
        row_num3 += 1

    if not month_data:
        ws3.merge_cells(f"A{row_num3}:D{row_num3}")
        ws3.cell(row_num3, 1, "No billing entries found for the selected filters")
        ws3.cell(row_num3, 1).font = Font(italic=True, size=9, color="999999")

    for i, w in enumerate([18, 20, 12, 50], 1):
        ws3.column_dimensions[chr(ord("A") + i - 1)].width = w

    # ── Stream ────────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=billing-statement-report.xlsx"},
    )


# ── Filter options ────────────────────────────────────────────────────────────
@router.get("/filter-options")
def filter_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    projects = db.query(Project).order_by(Project.name).all()
    return {
        "projects":      [{"id": p.id, "name": p.name} for p in projects],
        "billing_types": BILLING_TYPES,
    }
