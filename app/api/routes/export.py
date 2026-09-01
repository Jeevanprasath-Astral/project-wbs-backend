from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from io import BytesIO
from datetime import datetime
from app.db.database import get_db
from app.models.models import (Project, ProjectMilestone, Milestone, Task,
                                Subtask, Question, Response, SubtaskStatus, User,
                                CustomMilestone, CustomTask, CustomSubtask,
                                Activity, WorkHours)
from app.core.deps import get_current_user
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(tags=["Export"])

# ── Shared helpers ────────────────────────────────────────────────────────────
def _load_project_data(db, project_id):
    """Load all milestone data in bulk to avoid N+1 queries.

    "All Milestones" must only ever mean the milestones actually selected
    for THIS project via Milestone Configuration (CustomMilestone), not the
    full standard 10-milestone catalog. Different projects can select
    different subsets, so this filter is computed fresh per project.
    """
    from sqlalchemy.orm import joinedload
    project = db.query(Project).filter_by(id=project_id).first()
    selected_nums = {
        cm.num for cm in db.query(CustomMilestone).filter_by(
            project_id=project_id, is_active=True
        ).all()
    }
    pm_query = db.query(ProjectMilestone).filter_by(project_id=project_id)
    if selected_nums:
        pm_query = pm_query.filter(ProjectMilestone.num.in_(selected_nums))
    else:
        # No milestones selected yet for this project — export nothing
        # rather than silently falling back to all 10 standard milestones.
        pm_query = pm_query.filter(ProjectMilestone.num.in_([-1]))
    pms = pm_query.order_by(ProjectMilestone.num).all()
    milestones = {
        ms.num: ms for ms in db.query(Milestone).options(
            joinedload(Milestone.tasks)
            .joinedload(Task.subtasks)
            .joinedload(Subtask.questions)
        ).all()
    }
    # Bulk load all responses and statuses
    from sqlalchemy import or_
    all_responses = db.query(Response).filter_by(project_id=project_id).all()
    resp_by_question = {r.question_id: r.value for r in all_responses if r.question_id}
    resp_by_subtask  = {r.subtask_id: r.value  for r in all_responses if r.subtask_id and not r.question_id}
    all_ss = db.query(SubtaskStatus).filter_by(project_id=project_id).all()
    ss_by_subtask = {ss.subtask_id: ss for ss in all_ss}
    return project, pms, milestones, resp_by_question, resp_by_subtask, ss_by_subtask


# ── helpers ───────────────────────────────────────────────────────────────────
def _days_between(start, end) -> str:
    """Return integer day count between two DateTime values, or '' if either is None."""
    if not start or not end:
        return ""
    try:
        s = start if isinstance(start, datetime) else datetime.fromisoformat(str(start))
        e = end   if isinstance(end,   datetime) else datetime.fromisoformat(str(end))
        return max(0, (e - s).days)
    except Exception:
        return ""


def _fmt_date(dt) -> str:
    """Return YYYY-MM-DD string or '' if None."""
    if not dt:
        return ""
    try:
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(dt)[:10]


# ── Excel Export ──────────────────────────────────────────────────────────────
@router.get("/projects/{project_id}/export/xlsx")
def export_excel(project_id: int, milestone: int = None, db: Session = Depends(get_db),
                 current_user: User = Depends(get_current_user)):
    """Export using the Custom Milestone system (CustomMilestone → CustomTask →
    CustomSubtask → Activity).  Status is read from CustomMilestone.status so
    it always reflects the actual configured value rather than the legacy
    ProjectMilestone default."""

    project = db.query(Project).filter_by(id=project_id).first()

    # Load active custom milestones with full hierarchy (eager load to avoid N+1)
    cm_q = (db.query(CustomMilestone)
              .filter_by(project_id=project_id, is_active=True)
              .options(
                  joinedload(CustomMilestone.tasks)
                  .joinedload(CustomTask.subtasks)
                  .joinedload(CustomSubtask.activities)
              )
              .order_by(CustomMilestone.num))
    if milestone:
        cm_q = cm_q.filter(CustomMilestone.num == milestone)
    custom_milestones = cm_q.all()

    # Bulk-fetch actual consumed hours per subtask and per activity
    all_sub_ids = [s.id for cm in custom_milestones
                   for t  in cm.tasks
                   for s  in t.subtasks]
    all_act_ids = [a.id for cm in custom_milestones
                   for t  in cm.tasks
                   for s  in t.subtasks
                   for a  in s.activities]

    sub_actual: dict[int, float] = {}
    if all_sub_ids:
        rows = (db.query(WorkHours.custom_subtask_id, func.sum(WorkHours.hours_spent))
                  .filter(WorkHours.custom_subtask_id.in_(all_sub_ids))
                  .group_by(WorkHours.custom_subtask_id).all())
        sub_actual = {r[0]: round(float(r[1] or 0), 2) for r in rows}

    act_actual: dict[int, float] = {}
    if all_act_ids:
        rows = (db.query(WorkHours.activity_id, func.sum(WorkHours.hours_spent))
                  .filter(WorkHours.activity_id.in_(all_act_ids))
                  .group_by(WorkHours.activity_id).all())
        act_actual = {r[0]: round(float(r[1] or 0), 2) for r in rows}

    # ── Workbook setup ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
    def bdr():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    HDR_FILL  = fill("1F3864")
    TASK_FILL = fill("5B9BD5")
    COL_FILL  = fill("BDD7EE")
    EVEN_FILL = fill("EBF3FB")
    ODD_FILL  = fill("FFFFFF")
    DONE_FILL = fill("E2EFDA")
    PROG_FILL = fill("FFF2CC")
    OVER_FILL = fill("FCE4EC")
    TODO_FILL = fill("F0F0F0")
    ACT_FILL  = fill("F5F0FF")   # light purple tint for Activity rows

    STATUS_FILLS = {
        "Completed":   DONE_FILL, "In Progress": PROG_FILL,
        "Overdue":     OVER_FILL, "Not Started": TODO_FILL,
    }
    STATUS_COLORS = {
        "Completed":   "375623", "In Progress": "7F6000",
        "Overdue":     "A32D2D", "Not Started": "666666",
    }

    # New 9-column layout
    COL_HEADERS = [
        "Subtask / Question",   # A
        "Response / Input",     # B
        "Assignee (Multi)",     # C
        "Status",               # D
        "Planned Start",        # E
        "Planned End",          # F
        "Planned Total Days",   # G
        "Estimated Hours",      # H
        "Actual Consumed Hours",# I
    ]
    COL_WIDTHS = [42, 30, 28, 16, 14, 14, 18, 18, 22]
    LAST_COL   = "I"
    N_COLS     = 9

    for cm in custom_milestones:
        ws = wb.create_sheet(f"M{cm.num:02d}-{cm.name[:18]}")
        for col_letter, w in zip("ABCDEFGHI", COL_WIDTHS):
            ws.column_dimensions[col_letter].width = w

        # ── Title row (Status from CustomMilestone — actual configured value) ─
        ws.merge_cells(f"A1:{LAST_COL}1")
        c = ws["A1"]
        c.value = (f"M{cm.num:02d} — {cm.name}"
                   f"  |  Status: {cm.status or 'Not Started'}"
                   f"  |  Assignee: {cm.assignee or '—'}")
        c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

        # ── Project info row ──────────────────────────────────────────────────
        ws.merge_cells(f"A2:{LAST_COL}2")
        c2 = ws["A2"]
        c2.value = (f"Project: {project.name if project else ''}  "
                    f"|  Client: {project.client if project else ''}  "
                    f"|  Exported by: {current_user.name}")
        c2.font = Font(italic=True, color="555555", size=9, name="Calibri")
        c2.fill = fill("D9E8F5")
        c2.alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 16

        row = 4
        tasks_sorted = sorted(cm.tasks, key=lambda x: x.num or 0)

        for task in tasks_sorted:
            # ── Task header ───────────────────────────────────────────────────
            ws.merge_cells(f"A{row}:{LAST_COL}{row}")
            tc = ws.cell(row, 1, f"  Task {(task.num or 0):02d} — {task.name.upper()}")
            tc.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            tc.fill = TASK_FILL
            ws.row_dimensions[row].height = 20
            row += 1

            # ── Column headers ────────────────────────────────────────────────
            for col_idx, h in enumerate(COL_HEADERS, 1):
                hc = ws.cell(row, col_idx, h)
                hc.font = Font(bold=True, color="1F3864", size=9, name="Calibri")
                hc.fill = COL_FILL
                hc.border = bdr()
                hc.alignment = Alignment(horizontal="center")
            ws.row_dimensions[row].height = 17
            row += 1

            # ── Subtask rows ──────────────────────────────────────────────────
            subs_sorted = sorted(task.subtasks, key=lambda x: x.num or 0)
            for idx, sub in enumerate(subs_sorted):
                sub_status = sub.status or "Not Started"
                s_fill  = STATUS_FILLS.get(sub_status, TODO_FILL)
                s_color = STATUS_COLORS.get(sub_status, "666666")
                bg = EVEN_FILL if idx % 2 == 0 else ODD_FILL
                actual_hrs = sub_actual.get(sub.id, 0.0)
                total_days = _days_between(sub.planned_start, sub.planned_end)

                def _cell(c, v, bold=False, italic=False, color="333333", align="left", bg_=None):
                    cell = ws.cell(row, c, v if v not in (None, "") else "")
                    cell.font = Font(size=9, bold=bold, italic=italic, color=color, name="Calibri")
                    cell.fill = bg_ or bg
                    cell.border = bdr()
                    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=(c in (1, 2)))
                    return cell

                _cell(1, sub.name, bold=True)
                resp_val = sub.response or ""
                _cell(2, resp_val if resp_val else "— not filled —",
                      italic=not bool(resp_val),
                      color="0D47A1" if resp_val else "BBBBBB",
                      bg_=bg if resp_val else fill("FFFDE7"))
                _cell(3, sub.assignee or "—")
                sc = _cell(4, sub_status, bold=True, color=s_color, align="center", bg_=s_fill)
                _cell(5, _fmt_date(sub.planned_start), align="center")
                _cell(6, _fmt_date(sub.planned_end),   align="center")
                _cell(7, total_days, align="center")
                _cell(8, sub.estimated_hours or 0, align="center")
                _cell(9, actual_hrs,               align="center")
                ws.row_dimensions[row].height = 17
                row += 1

                # ── Activity rows (indented under their subtask) ──────────────
                for act in sorted(sub.activities, key=lambda x: x.id):
                    act_status = act.status or "Not Started"
                    a_fill  = STATUS_FILLS.get(act_status, TODO_FILL)
                    a_color = STATUS_COLORS.get(act_status, "666666")
                    act_hrs    = act_actual.get(act.id, 0.0)
                    act_days   = _days_between(act.planned_start, act.planned_end)

                    def _acell(c, v, bold=False, italic=False, color="555555", align="left"):
                        cell = ws.cell(row, c, v if v not in (None, "") else "")
                        cell.font = Font(size=8, bold=bold, italic=italic, color=color, name="Calibri")
                        cell.fill = ACT_FILL
                        cell.border = bdr()
                        cell.alignment = Alignment(horizontal=align, vertical="top")
                        return cell

                    _acell(1, f"  ↳ {act.name}")
                    _acell(2, "")   # activities carry no free-text response
                    _acell(3, act.assignee or "—")
                    _acell(4, act_status, bold=True, color=a_color, align="center")
                    _acell(5, _fmt_date(act.planned_start), align="center")
                    _acell(6, _fmt_date(act.planned_end),   align="center")
                    _acell(7, act_days,            align="center")
                    _acell(8, act.estimated_hours or 0, align="center")
                    _acell(9, act_hrs,             align="center")
                    ws.row_dimensions[row].height = 15
                    row += 1

            row += 1  # blank gap between tasks

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=project-wbs-{project_id}.xlsx"})


# ── PDF Export ────────────────────────────────────────────────────────────────
@router.get("/projects/{project_id}/export/pdf")
def export_pdf(project_id: int, milestone: int = None, db: Session = Depends(get_db),
               current_user: User = Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, PageBreak, HRFlowable)
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from datetime import date

    project, pms, milestones, resp_by_q, resp_by_sub, ss_by_sub = _load_project_data(db, project_id)
    if milestone:
        pms = [pm for pm in pms if pm.num == milestone]

    # Build a num → CustomMilestone lookup so PDF headers show the actual status
    # rather than the ProjectMilestone default ("Not Started").
    cm_by_num = {
        cm.num: cm for cm in db.query(CustomMilestone).filter_by(
            project_id=project_id, is_active=True
        ).all()
    }

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)

    # Styles
    styles = getSampleStyleSheet()
    style_title   = ParagraphStyle("title",   fontSize=18, fontName="Helvetica-Bold",
                                   textColor=colors.HexColor("#1F3864"), spaceAfter=4, alignment=TA_LEFT)
    style_sub     = ParagraphStyle("sub",     fontSize=9,  fontName="Helvetica",
                                   textColor=colors.HexColor("#555555"), spaceAfter=12)
    style_ms_hdr  = ParagraphStyle("ms_hdr",  fontSize=12, fontName="Helvetica-Bold",
                                   textColor=colors.white, spaceAfter=0)
    style_task    = ParagraphStyle("task",    fontSize=10, fontName="Helvetica-Bold",
                                   textColor=colors.white, spaceAfter=0)
    style_subtask = ParagraphStyle("subtask", fontSize=9,  fontName="Helvetica-Bold",
                                   textColor=colors.HexColor("#1F3864"), spaceAfter=0)
    style_cell    = ParagraphStyle("cell",    fontSize=8,  fontName="Helvetica",
                                   textColor=colors.HexColor("#333333"), spaceAfter=0, leading=10)
    style_resp    = ParagraphStyle("resp",    fontSize=8,  fontName="Helvetica",
                                   textColor=colors.HexColor("#0D47A1"), spaceAfter=0, leading=10)
    style_empty   = ParagraphStyle("empty",   fontSize=8,  fontName="Helvetica-Oblique",
                                   textColor=colors.HexColor("#AAAAAA"), spaceAfter=0)

    # Color constants
    C_NAVY   = colors.HexColor("#1F3864")
    C_BLUE   = colors.HexColor("#2E75B6")
    C_LBLUE  = colors.HexColor("#5B9BD5")
    C_LIGHT  = colors.HexColor("#EBF3FB")
    C_WHITE  = colors.white
    C_HDRB   = colors.HexColor("#BDD7EE")
    C_DONE   = colors.HexColor("#E2EFDA")
    C_PROG   = colors.HexColor("#FFF2CC")
    C_OVER   = colors.HexColor("#FCE4EC")
    C_TODO   = colors.HexColor("#F0F0F0")
    C_UNFILL = colors.HexColor("#FFFDE7")

    STATUS_BG = {
        "Completed": C_DONE, "In Progress": C_PROG,
        "Overdue": C_OVER, "Not Started": C_TODO,
    }
    STATUS_FG = {
        "Completed": colors.HexColor("#375623"),
        "In Progress": colors.HexColor("#7F6000"),
        "Overdue": colors.HexColor("#A32D2D"),
        "Not Started": colors.HexColor("#666666"),
    }

    story = []
    page_w = landscape(A4)[0] - 3*cm
    col_widths = [1.2*cm, 8*cm, 9*cm, 4*cm, 3*cm, 3*cm]

    # ── Cover page ────────────────────────────────────────────────────────────
    story.append(Spacer(1, 3*cm))
    story.append(Paragraph("PROJECT WBS", ParagraphStyle("cover1", fontSize=32,
        fontName="Helvetica-Bold", textColor=C_NAVY, alignment=TA_CENTER)))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("Requirement Gathering & Tracking Report",
        ParagraphStyle("cover2", fontSize=16, fontName="Helvetica",
                       textColor=C_BLUE, alignment=TA_CENTER)))
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="80%", thickness=2, color=C_NAVY, spaceAfter=0.5*cm))

    cover_data = [
        ["Project:", project.name if project else "—"],
        ["Client:", project.client if project else "—"],
        ["Owner:", project.owner if project else "—"],
        ["Exported by:", current_user.name],
        ["Export date:", date.today().strftime("%d %B %Y")],
        ["Total milestones:", str(len(pms))],
    ]
    cover_table = Table(cover_data, colWidths=[5*cm, 12*cm])
    cover_table.setStyle(TableStyle([
        ("FONTNAME",    (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",    (0,0), (-1,-1), 11),
        ("FONTNAME",    (0,0), (0,-1),  "Helvetica-Bold"),
        ("TEXTCOLOR",   (0,0), (0,-1),  C_NAVY),
        ("TEXTCOLOR",   (1,0), (1,-1),  colors.HexColor("#333333")),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.HexColor("#F8F8FF"), C_WHITE]),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("ALIGN",       (0,0), (-1,-1), "LEFT"),
    ]))
    story.append(cover_table)
    story.append(PageBreak())

    # ── Milestone pages ───────────────────────────────────────────────────────
    for pm in pms:
        ms = milestones.get(pm.num)
        if not ms: continue

        # Milestone header — use CustomMilestone.status (actual configured value)
        _cm = cm_by_num.get(pm.num)
        _cm_status = _cm.status if _cm and _cm.status else pm.status
        ms_hdr = Table([[Paragraph(
            f"Milestone {pm.num:02d} — {ms.name.upper()}   |   Status: {_cm_status}   |   Progress: {pm.progress:.1f}%",
            style_ms_hdr)]], colWidths=[page_w])
        ms_hdr.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), C_NAVY),
            ("TOPPADDING",    (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ]))
        story.append(ms_hdr)
        story.append(Spacer(1, 0.2*cm))

        for task in sorted(ms.tasks, key=lambda x: x.num or 0):
            # Task header
            task_hdr = Table([[Paragraph(
                f"Task {task.num:02d} — {task.name}", style_task)]], colWidths=[page_w])
            task_hdr.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,-1), C_LBLUE),
                ("TOPPADDING",    (0,0), (-1,-1), 6),
                ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                ("LEFTPADDING",   (0,0), (-1,-1), 10),
            ]))
            story.append(task_hdr)

            # Column headers
            col_hdr = Table([[
                Paragraph("#", style_subtask),
                Paragraph("Subtask / Question", style_subtask),
                Paragraph("Response / Input", style_subtask),
                Paragraph("Owner", style_subtask),
                Paragraph("Status", style_subtask),
                Paragraph("Sign-off", style_subtask),
            ]], colWidths=col_widths)
            col_hdr.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,-1), C_HDRB),
                ("TOPPADDING",    (0,0), (-1,-1), 5),
                ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ("LEFTPADDING",   (0,0), (-1,-1), 6),
                ("GRID",          (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ]))
            story.append(col_hdr)

            rows = []
            row_styles = []
            idx = 0

            for sub in sorted(task.subtasks, key=lambda x: x.num or 0):
                ss = ss_by_sub.get(sub.id)
                sub_status = ss.status if ss else "Not Started"
                signed = ss.reviewer if ss and ss.signed_off_at else "—"
                bg = STATUS_BG.get(sub_status, C_TODO)
                fg = STATUS_FG.get(sub_status, colors.HexColor("#666666"))

                if sub.is_format and sub.questions:
                    # Subtask group label
                    rows.append([
                        Paragraph("", style_cell),
                        Paragraph(f"▸  {sub.name}", style_subtask),
                        Paragraph("", style_cell),
                        Paragraph(task.responsibility, style_cell),
                        Paragraph(sub_status, ParagraphStyle("st", fontSize=8,
                            fontName="Helvetica-Bold", textColor=fg, spaceAfter=0)),
                        Paragraph(signed, style_cell),
                    ])
                    row_styles.append(("BACKGROUND", (0,idx), (-1,idx), colors.HexColor("#EEF4FB")))
                    idx += 1

                    for q in sorted(sub.questions, key=lambda x: x.num or 0):
                        val = resp_by_q.get(q.id, "")
                        alt_bg = C_LIGHT if idx % 2 == 0 else C_WHITE
                        rows.append([
                            Paragraph(str(q.num), style_cell),
                            Paragraph(q.question_text or "", style_cell),
                            Paragraph(val, style_resp) if val else Paragraph("— not filled —", style_empty),
                            Paragraph(task.responsibility, style_cell),
                            Paragraph(sub_status, ParagraphStyle("st2", fontSize=8,
                                fontName="Helvetica-Bold", textColor=fg, spaceAfter=0)),
                            Paragraph(signed, style_cell),
                        ])
                        row_styles.append(("BACKGROUND", (0,idx), (-1,idx), alt_bg if val else C_UNFILL))
                        idx += 1
                else:
                    val = resp_by_sub.get(sub.id, "")
                    alt_bg = C_LIGHT if idx % 2 == 0 else C_WHITE
                    rows.append([
                        Paragraph(str(sub.num or ""), style_cell),
                        Paragraph(sub.name, ParagraphStyle("sn", fontSize=9,
                            fontName="Helvetica-Bold", textColor=colors.HexColor("#1F3864"), spaceAfter=0)),
                        Paragraph(val, style_resp) if val else Paragraph("— not filled —", style_empty),
                        Paragraph(task.responsibility, style_cell),
                        Paragraph(sub_status, ParagraphStyle("st3", fontSize=8,
                            fontName="Helvetica-Bold", textColor=fg, spaceAfter=0)),
                        Paragraph(signed, style_cell),
                    ])
                    row_styles.append(("BACKGROUND", (0,idx), (-1,idx), alt_bg if val else C_UNFILL))
                    idx += 1

            if rows:
                t = Table(rows, colWidths=col_widths, repeatRows=0)
                ts = TableStyle([
                    ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#DDDDDD")),
                    ("TOPPADDING",    (0,0), (-1,-1), 4),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                    ("LEFTPADDING",   (0,0), (-1,-1), 6),
                    ("RIGHTPADDING",  (0,0), (-1,-1), 4),
                    ("VALIGN",        (0,0), (-1,-1), "TOP"),
                ] + row_styles)
                t.setStyle(ts)
                story.append(t)

            story.append(Spacer(1, 0.3*cm))

        story.append(PageBreak())

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=project-wbs-{project_id}.pdf"})
