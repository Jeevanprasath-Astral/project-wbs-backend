"""Billing Reports — Monthly Billing Tracker & Billing Status Report.

Endpoints:
  GET /billing-reports/monthly-tracker         — plan vs actual grouped by month × billing_type
  GET /billing-reports/monthly-tracker/export  — same data as xlsx download
  GET /billing-reports/billing-status          — per-entry with computed billing status
  GET /billing-reports/billing-status/export   — same data as xlsx download
"""
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sa_func
from typing import Optional
from datetime import date as date_cls
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db.database import get_db
from app.models.models import ProjectBilling, CustomMilestone, Project, User
from app.core.deps import get_current_user

router = APIRouter(prefix="/billing-reports", tags=["Billing Reports"])


def _compute_status(planned_date, actual_date, today):
    """Return (status_label, days_variance)."""
    if actual_date is None:
        # Billing hasn't happened yet
        if planned_date is None:
            return "Upcoming", None
        delta = (planned_date - today).days
        if delta >= 0:
            return "Upcoming", delta
        else:
            return "Overdue", abs(delta)
    # Billing has happened
    if planned_date is None:
        return "On Time", 0
    delta = (actual_date - planned_date).days
    if delta < 0:
        return "Before Schedule", abs(delta)
    if delta == 0:
        return "On Time", 0
    return "Delayed", delta


@router.get("/monthly-tracker")
def monthly_billing_tracker(
    project_id:  Optional[int] = None,
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    billing_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Plan vs actual amounts grouped by calendar month and billing type.

    Uses milestone.planned_end as the Planned Billing Date for grouping.
    """
    q = db.query(ProjectBilling)
    if project_id:
        q = q.filter(ProjectBilling.project_id == project_id)
    if billing_type:
        q = q.filter(ProjectBilling.billing_type == billing_type)
    entries = q.all()

    # Resolve planned_billing_date from milestone.planned_end
    ms_cache = {}
    proj_cache = {}
    for e in entries:
        if e.milestone_id and e.milestone_id not in ms_cache:
            m = db.query(CustomMilestone).filter_by(id=e.milestone_id).first()
            ms_cache[e.milestone_id] = m
        if e.project_id and e.project_id not in proj_cache:
            p = db.query(Project).filter_by(id=e.project_id).first()
            proj_cache[e.project_id] = p

    # Filter by start/end on planned_billing_date.
    # planned_end is a DateTime column — extract .date() so comparisons with date objects work.
    def get_pbd(e):
        m = ms_cache.get(e.milestone_id)
        return m.planned_end.date() if m and m.planned_end else None

    filtered = []
    for e in entries:
        pbd = get_pbd(e)
        if start_date and (pbd is None or pbd < date_cls.fromisoformat(start_date)):
            continue
        if end_date and (pbd is None or pbd > date_cls.fromisoformat(end_date)):
            continue
        filtered.append(e)

    # Group by (month_key, billing_type)
    #   month_key derived from planned_billing_date (YYYY-MM)
    groups = defaultdict(lambda: {"planned": 0.0, "actual": 0.0, "count": 0})
    for e in filtered:
        pbd = get_pbd(e)
        month = pbd.strftime("%Y-%m") if pbd else "Unscheduled"
        bt    = e.billing_type or "Unspecified"
        key   = (month, bt)
        groups[key]["planned"] += float(e.planned_billing_amount or 0)
        groups[key]["actual"]  += float(e.actual_billing_amount  or 0)
        groups[key]["count"]   += 1

    rows = []
    for (month, bt), data in sorted(groups.items()):
        rows.append({
            "month":           month,
            "billing_type":    bt,
            "planned_amount":  round(data["planned"], 2),
            "actual_amount":   round(data["actual"],  2),
            "variance":        round(data["actual"] - data["planned"], 2),
            "entry_count":     data["count"],
        })

    # Month-level subtotals
    month_totals = defaultdict(lambda: {"planned": 0.0, "actual": 0.0, "count": 0})
    for r in rows:
        month_totals[r["month"]]["planned"] += r["planned_amount"]
        month_totals[r["month"]]["actual"]  += r["actual_amount"]
        month_totals[r["month"]]["count"]   += r["entry_count"]

    return {
        "rows":         rows,
        "month_totals": [
            {
                "month":          m,
                "planned_amount": round(d["planned"], 2),
                "actual_amount":  round(d["actual"],  2),
                "variance":       round(d["actual"] - d["planned"], 2),
                "entry_count":    d["count"],
            }
            for m, d in sorted(month_totals.items())
        ],
        "grand_total": {
            "planned_amount": round(sum(r["planned_amount"] for r in rows), 2),
            "actual_amount":  round(sum(r["actual_amount"]  for r in rows), 2),
            "variance":       round(sum(r["variance"]       for r in rows), 2),
            "entry_count":    sum(r["entry_count"] for r in rows),
        },
        "applied_filters": {
            "project_id":   project_id,
            "start_date":   start_date,
            "end_date":     end_date,
            "billing_type": billing_type,
        },
    }


@router.get("/billing-status")
def billing_status_report(
    project_id:   Optional[int] = None,
    status_filter: Optional[str] = None,   # Upcoming | Overdue | On Time | Delayed | Before Schedule
    billing_type:  Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Per-entry billing status report with computed status.

    Status logic:
      Upcoming       — no actual date, planned date is today or future
      Overdue        — no actual date, planned date is in the past
      Before Schedule — actual date < planned date  (billed early)
      On Time         — actual date == planned date
      Delayed         — actual date > planned date  (billed late)
    """
    today = date_cls.today()

    q = db.query(ProjectBilling)
    if project_id:
        q = q.filter(ProjectBilling.project_id == project_id)
    if billing_type:
        q = q.filter(ProjectBilling.billing_type == billing_type)
    entries = q.all()

    ms_cache   = {}
    proj_cache = {}
    for e in entries:
        if e.milestone_id and e.milestone_id not in ms_cache:
            m = db.query(CustomMilestone).filter_by(id=e.milestone_id).first()
            ms_cache[e.milestone_id] = m
        if e.project_id and e.project_id not in proj_cache:
            p = db.query(Project).filter_by(id=e.project_id).first()
            proj_cache[e.project_id] = p

    rows = []
    for e in entries:
        ms   = ms_cache.get(e.milestone_id)
        proj = proj_cache.get(e.project_id)
        # planned_end is a DateTime column — extract .date() so date comparisons and str() work correctly
        pbd  = ms.planned_end.date() if ms and ms.planned_end else None   # planned_billing_date

        status, days_var = _compute_status(pbd, e.actual_billing_date, today)

        if status_filter and status != status_filter:
            continue

        rows.append({
            "id":                     e.id,
            "project":                proj.name   if proj else "—",
            "project_id":             e.project_id,
            "milestone":              ms.name     if ms   else "—",
            "milestone_id":           e.milestone_id,
            "billing_type":           e.billing_type,
            "planned_billing_date":   str(pbd)                       if pbd                   else None,
            "planned_billing_amount": float(e.planned_billing_amount or 0),
            "actual_billing_date":    str(e.actual_billing_date)     if e.actual_billing_date else None,
            "actual_billing_amount":  float(e.actual_billing_amount) if e.actual_billing_amount is not None else None,
            "description":            e.description,
            "remarks":                e.remarks,
            "status":                 status,
            "days_variance":          days_var,
        })

    # Sort: Overdue first, then Upcoming, then Delayed, On Time, Before Schedule
    STATUS_ORDER = {"Overdue": 0, "Upcoming": 1, "Delayed": 2, "On Time": 3, "Before Schedule": 4}
    rows.sort(key=lambda r: (STATUS_ORDER.get(r["status"], 9), r["planned_billing_date"] or "9999"))

    # Summary counts
    summary = {}
    for r in rows:
        summary[r["status"]] = summary.get(r["status"], 0) + 1

    return {
        "rows":         rows,
        "summary":      summary,
        "total_entries": len(rows),
        "applied_filters": {
            "project_id":    project_id,
            "status_filter": status_filter,
            "billing_type":  billing_type,
        },
    }


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _xl_header_style():
    """Returns (Font, PatternFill, Alignment, Border) for header cells."""
    fill   = PatternFill("solid", fgColor="4F46E5")
    font   = Font(bold=True, color="FFFFFF", size=10)
    align  = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return font, fill, align, border


def _xl_subtotal_style():
    fill   = PatternFill("solid", fgColor="EDE9FE")
    font   = Font(bold=True, size=9)
    align  = Alignment(horizontal="center", vertical="center")
    return font, fill, align


def _xl_grand_style():
    fill   = PatternFill("solid", fgColor="7C3AED")
    font   = Font(bold=True, color="FFFFFF", size=10)
    align  = Alignment(horizontal="center", vertical="center")
    return font, fill, align


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt_inr(v):
    return f"₹{v:,.2f}" if v is not None else "—"


# ── Monthly Tracker Export ────────────────────────────────────────────────────

@router.get("/monthly-tracker/export")
def monthly_billing_tracker_export(
    project_id:   Optional[int] = None,
    start_date:   Optional[str] = None,
    end_date:     Optional[str] = None,
    billing_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download Monthly Billing Tracker as .xlsx"""
    # Reuse the same data logic
    data = monthly_billing_tracker(
        project_id=project_id, start_date=start_date,
        end_date=end_date, billing_type=billing_type,
        db=db, current_user=current_user,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Billing Tracker"

    # ── Header row
    headers = ["Month", "Billing Type", "Planned Amount (₹)", "Actual Amount (₹)", "Variance (₹)", "Entry Count"]
    hfont, hfill, halign, hborder = _xl_header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = hborder

    ws.row_dimensions[1].height = 22

    # ── Data rows + month subtotals
    row_num = 2
    rows_by_month = defaultdict(list)
    for r in data["rows"]:
        rows_by_month[r["month"]].append(r)

    for month_key in sorted(rows_by_month.keys()):
        month_rows = rows_by_month[month_key]
        for r in month_rows:
            ws.cell(row=row_num, column=1, value=r["month"])
            ws.cell(row=row_num, column=2, value=r["billing_type"])
            ws.cell(row=row_num, column=3, value=r["planned_amount"])
            ws.cell(row=row_num, column=4, value=r["actual_amount"])
            ws.cell(row=row_num, column=5, value=r["variance"])
            ws.cell(row=row_num, column=6, value=r["entry_count"])
            # Right-align currency columns
            for col in (3, 4, 5):
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="right")
            row_num += 1

        # Month subtotal row
        mt = next((m for m in data["month_totals"] if m["month"] == month_key), None)
        if mt:
            sfont, sfill, salign = _xl_subtotal_style()
            vals = [month_key, "SUBTOTAL", mt["planned_amount"], mt["actual_amount"], mt["variance"], mt["entry_count"]]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=v)
                cell.font = sfont; cell.fill = sfill
                cell.alignment = Alignment(horizontal="right") if col >= 3 else salign
            row_num += 1

    # Grand total row
    gt = data["grand_total"]
    gfont, gfill, galign = _xl_grand_style()
    vals = ["GRAND TOTAL", "", gt["planned_amount"], gt["actual_amount"], gt["variance"], gt["entry_count"]]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=row_num, column=col, value=v)
        cell.font = gfont; cell.fill = gfill
        cell.alignment = Alignment(horizontal="right") if col >= 3 else galign

    _set_col_widths(ws, [14, 22, 20, 20, 18, 12])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=monthly_billing_tracker.xlsx"},
    )


# ── Billing Status Export ─────────────────────────────────────────────────────

@router.get("/billing-status/export")
def billing_status_report_export(
    project_id:    Optional[int] = None,
    status_filter: Optional[str] = None,
    billing_type:  Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download Billing Status Report as .xlsx"""
    data = billing_status_report(
        project_id=project_id, status_filter=status_filter,
        billing_type=billing_type, db=db, current_user=current_user,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Billing Status"

    headers = [
        "Project", "Milestone", "Billing Type",
        "Planned Date", "Planned Amount (₹)",
        "Actual Date", "Actual Amount (₹)",
        "Status", "Days Variance", "Description", "Remarks",
    ]
    hfont, hfill, halign, hborder = _xl_header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = hborder
    ws.row_dimensions[1].height = 22

    STATUS_COLORS_XL = {
        "Overdue":         "FEE2E2",
        "Upcoming":        "DBEAFE",
        "Delayed":         "FEF3C7",
        "On Time":         "D1FAE5",
        "Before Schedule": "EDE9FE",
    }

    for row_num, r in enumerate(data["rows"], 2):
        vals = [
            r["project"], r["milestone"], r["billing_type"],
            r["planned_billing_date"],  r["planned_billing_amount"],
            r["actual_billing_date"],   r["actual_billing_amount"],
            r["status"], r["days_variance"], r["description"], r["remarks"],
        ]
        status_color = STATUS_COLORS_XL.get(r["status"], "FFFFFF")
        fill = PatternFill("solid", fgColor=status_color)
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.fill = fill
            if col in (5, 7):
                cell.alignment = Alignment(horizontal="right")

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Status").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Count").font = Font(bold=True)
    for i, (status, count) in enumerate(data["summary"].items(), 2):
        ws2.cell(row=i, column=1, value=status)
        ws2.cell(row=i, column=2, value=count)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10

    _set_col_widths(ws, [20, 22, 18, 14, 20, 14, 20, 16, 13, 25, 25])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=billing_status_report.xlsx"},
    )
