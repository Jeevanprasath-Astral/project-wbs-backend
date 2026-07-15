"""Timesheet Calendar — Report exports.

Five Excel report exports, one per report type. Two configurable pieces,
both finalized with the user before implementation:

  1. Timeline filter — each report has its OWN start_date/end_date range,
     independent of the other reports and of the Month filter elsewhere on
     the Timesheet Calendar page. No shared/common period control. Defaults
     to the current calendar month when omitted.
  2. Configurable columns — each report has its own fixed, small set of
     fields (REPORT_FIELDS below). Every field is independently
     included/excluded — fields never swap into another field's slot. The
     frontend sends `columns` as an ordered, comma-separated list of the
     field keys the user kept (fields toggled to "None" are simply omitted
     from the list — the backend never sees "None" itself). Output column
     order follows the list order. At least one field must be selected.

Mirrors the styling/streaming convention established in export.py
(openpyxl Workbook -> BytesIO -> StreamingResponse).

Other design notes (finalized earlier):
  - Individual Time Report: one row per WorkHours entry (not per calendar
    day), since a single day can have several entries across projects.
  - Overtime Report ("> 7h/day"): based on WORKED hours only (sum of
    hours_spent per user/day) — permission hours are intentionally excluded.
  - Leave / Permission Reports: Approved requests only.
  - Holiday Report: Date + Holiday Name only — Holiday rows aren't tied to
    an individual employee.
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date as date_cls
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.database import get_db
from app.models.models import WorkHours, LeaveRequest, Permission, Holiday, User, Project
from app.core.deps import get_current_user

router = APIRouter(prefix="/timesheet/reports", tags=["Timesheet Reports"])

MIN_DAILY_HOURS = 7.0


# ── Per-report timeline filter ────────────────────────────────────────────────
# No shared/common period control — each report endpoint takes its own
# start_date/end_date directly. Defaults to the current calendar month when
# neither is given, so an export with no filters still returns something
# sensible.
def _resolve_range(start_date: Optional[str] = None, end_date: Optional[str] = None):
    today = date_cls.today()
    if not start_date and not end_date:
        return today.replace(day=1), today
    if not start_date or not end_date:
        raise HTTPException(400, "Both start_date and end_date are required")
    try:
        s = date_cls.fromisoformat(start_date)
        e = date_cls.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(400, "start_date/end_date must be in YYYY-MM-DD format")
    if s > e:
        raise HTTPException(400, "start_date must be on or before end_date")
    return s, e


# ── Configurable columns ──────────────────────────────────────────────────────
# Per-report field keys -> column headers. Order here is just the default;
# actual output order/selection is driven by the `columns` query param.
REPORT_FIELDS = {
    "individual_time": [("start_date", "Start Date"), ("end_date", "End Date"), ("project", "Project"),
                         ("billable", "Billable / Non-Billable"), ("hours", "Number of Hours")],
    "leave": [("employee", "Employee Name"), ("leave_reason", "Leave Reason"),
              ("leave_start_date", "Leave Start Date"), ("leave_end_date", "Leave End Date"),
              ("leave_days", "Leave Days")],
    "overtime": [("employee", "Employee Name"), ("start_date", "Start Date"), ("end_date", "End Date"),
                 ("hours_worked", "Hours Worked"), ("overtime_hours", "Overtime Hours")],
    "holiday": [("start_date", "Start Date"), ("end_date", "End Date"), ("holiday_name", "Holiday Name")],
    "permission": [("employee", "Employee Name"), ("start_date", "Start Date"), ("end_date", "End Date"),
                    ("hours", "Hours"), ("permission_reason", "Permission Reason")],
}


def _resolve_columns(report_key: str, columns: Optional[str]):
    field_defs = REPORT_FIELDS[report_key]
    valid_keys = [k for k, _ in field_defs]
    label_of = dict(field_defs)

    if not columns:
        keys = valid_keys
    else:
        requested = [c.strip() for c in columns.split(",") if c.strip()]
        seen = set()
        keys = [k for k in requested if k in valid_keys and not (k in seen or seen.add(k))]

    if not keys:
        raise HTTPException(400, "Select at least one report field")
    return keys, [label_of[k] for k in keys]


# ── Shared xlsx builder ───────────────────────────────────────────────────────
def _build_xlsx(title: str, headers: list, rows: list, filename: str, subtitle: str = ""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
    def bdr():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    HDR_FILL = fill("1F3864"); COL_FILL = fill("BDD7EE")
    EVEN_FILL = fill("EBF3FB"); ODD_FILL = fill("FFFFFF")

    ncols = len(headers)
    last_col_letter = chr(ord("A") + ncols - 1)

    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.fill = HDR_FILL; c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    if subtitle:
        ws.merge_cells(f"A2:{last_col_letter}2")
        c2 = ws["A2"]
        c2.value = subtitle
        c2.font = Font(italic=True, color="555555", size=9, name="Calibri")
        c2.fill = fill("D9E8F5"); c2.alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 16
        header_row = 4
    else:
        header_row = 3

    for col, h in enumerate(headers, 1):
        hc = ws.cell(header_row, col, h)
        hc.font = Font(bold=True, color="1F3864", size=9, name="Calibri")
        hc.fill = COL_FILL; hc.border = bdr()
        hc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[header_row].height = 18

    row = header_row + 1
    for i, r in enumerate(rows):
        bg = EVEN_FILL if i % 2 == 0 else ODD_FILL
        for col, val in enumerate(r, 1):
            cell = ws.cell(row, col, val if val not in (None, "") else "—")
            cell.font = Font(size=9, name="Calibri")
            cell.fill = bg; cell.border = bdr()
        row += 1

    if not rows:
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        ws.cell(row, 1, "No records for the selected filters").font = Font(italic=True, size=9, color="999999")

    for col_idx in range(1, ncols + 1):
        col_letter = chr(ord("A") + col_idx - 1)
        ws.column_dimensions[col_letter].width = 22

    output = BytesIO()
    wb.save(output); output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


def _period_label(start, end):
    return f"Period: {start}" if start == end else f"Period: {start} → {end}"


# ── 1. Individual Time Report ─────────────────────────────────────────────────
@router.get("/individual-time")
def individual_time_report(start_date: Optional[str] = None, end_date: Optional[str] = None,
                            user_id: Optional[int] = None, project_id: Optional[int] = None,
                            work_type: Optional[str] = None, is_billable: Optional[bool] = None,
                            columns: Optional[str] = None,
                            db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    start, end = _resolve_range(start_date, end_date)
    keys, headers = _resolve_columns("individual_time", columns)

    q = db.query(WorkHours).filter(WorkHours.date >= start, WorkHours.date <= end)
    if user_id: q = q.filter(WorkHours.user_id == user_id)
    if project_id: q = q.filter(WorkHours.project_id == project_id)
    if work_type: q = q.filter(WorkHours.work_type == work_type)
    elif is_billable is not None: q = q.filter(WorkHours.is_billable == is_billable)
    entries = q.order_by(WorkHours.date).all()

    rows = []
    for w in entries:
        project = db.query(Project).filter_by(id=w.project_id).first() if w.project_id else None
        values = {
            "start_date": str(w.date), "end_date": str(w.date),
            "project": project.name if project else "📋 General",
            "billable": w.work_type if w.work_type else ("Billable" if (w.is_billable if w.is_billable is not None else True) else "Non-Billable"),
            "hours": w.hours_spent or 0,
        }
        rows.append([values[k] for k in keys])

    employee = db.query(User).filter_by(id=user_id).first() if user_id else None
    subtitle = _period_label(start, end) + (f" | Employee: {employee.name}" if employee else "")
    return _build_xlsx("Individual Time Report", headers, rows,
                        f"individual-time-report-{start}-to-{end}.xlsx", subtitle)


# ── 2. Leave Report ────────────────────────────────────────────────────────────
@router.get("/leave")
def leave_report(start_date: Optional[str] = None, end_date: Optional[str] = None,
                  user_id: Optional[int] = None, columns: Optional[str] = None,
                  db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    start, end = _resolve_range(start_date, end_date)
    keys, headers = _resolve_columns("leave", columns)

    q = db.query(LeaveRequest).filter(LeaveRequest.date_from <= end, LeaveRequest.date_to >= start,
                                       LeaveRequest.status == "Approved")
    if user_id: q = q.filter(LeaveRequest.user_id == user_id)
    leaves = q.order_by(LeaveRequest.date_from).all()

    rows = []
    for l in leaves:
        u = db.query(User).filter_by(id=l.user_id).first()
        days = (l.date_to - l.date_from).days + 1 if l.date_from and l.date_to else 0
        values = {
            "employee": u.name if u else "—",
            "leave_reason": l.reason,
            "leave_start_date": str(l.date_from) if l.date_from else "—",
            "leave_end_date": str(l.date_to) if l.date_to else "—",
            "leave_days": days,
        }
        rows.append([values[k] for k in keys])

    return _build_xlsx("Leave Report", headers, rows,
                        f"leave-report-{start}-to-{end}.xlsx", _period_label(start, end) + " | Approved leave only")


# ── 3. Overtime Report ─────────────────────────────────────────────────────────
@router.get("/overtime")
def overtime_report(start_date: Optional[str] = None, end_date: Optional[str] = None,
                     user_id: Optional[int] = None, project_id: Optional[int] = None, columns: Optional[str] = None,
                     db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    start, end = _resolve_range(start_date, end_date)
    keys, headers = _resolve_columns("overtime", columns)

    q = db.query(WorkHours).filter(WorkHours.date >= start, WorkHours.date <= end)
    if user_id: q = q.filter(WorkHours.user_id == user_id)
    if project_id: q = q.filter(WorkHours.project_id == project_id)
    entries = q.all()

    # Worked-hours-only basis: sum hours_spent per (user, date), ignoring
    # permission hours, then flag any day over the 7h/day minimum.
    totals = {}
    for w in entries:
        key = (w.user_id, w.date)
        totals[key] = totals.get(key, 0) + (w.hours_spent or 0)

    rows = []
    for (uid, d), hours in sorted(totals.items(), key=lambda kv: kv[0][1]):
        if hours > MIN_DAILY_HOURS:
            u = db.query(User).filter_by(id=uid).first()
            values = {
                "employee": u.name if u else "—",
                "start_date": str(d), "end_date": str(d),
                "hours_worked": round(hours, 2),
                "overtime_hours": round(hours - MIN_DAILY_HOURS, 2),
            }
            rows.append([values[k] for k in keys])

    subtitle = _period_label(start, end) + f" | Threshold: >{MIN_DAILY_HOURS}h/day worked"
    return _build_xlsx("Overtime Report", headers, rows, f"overtime-report-{start}-to-{end}.xlsx", subtitle)


# ── 4. Holiday Report ──────────────────────────────────────────────────────────
@router.get("/holiday")
def holiday_report(start_date: Optional[str] = None, end_date: Optional[str] = None,
                    project_id: Optional[int] = None, holiday_name: Optional[str] = None,
                    columns: Optional[str] = None,
                    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    start, end = _resolve_range(start_date, end_date)
    keys, headers = _resolve_columns("holiday", columns)

    q = db.query(Holiday).filter(Holiday.date >= start, Holiday.date <= end)
    if project_id: q = q.filter((Holiday.project_id == project_id) | (Holiday.project_id.is_(None)))
    if holiday_name: q = q.filter(Holiday.name == holiday_name)
    holidays = q.order_by(Holiday.date).all()

    rows = []
    for h in holidays:
        values = {"start_date": str(h.date), "end_date": str(h.date), "holiday_name": h.name}
        rows.append([values[k] for k in keys])

    return _build_xlsx("Holiday Report", headers, rows, f"holiday-report-{start}-to-{end}.xlsx", _period_label(start, end))


# ── 5. Permission Report ───────────────────────────────────────────────────────
@router.get("/permission")
def permission_report(start_date: Optional[str] = None, end_date: Optional[str] = None,
                       user_id: Optional[int] = None, columns: Optional[str] = None,
                       db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    start, end = _resolve_range(start_date, end_date)
    keys, headers = _resolve_columns("permission", columns)

    q = db.query(Permission).filter(Permission.date >= start, Permission.date <= end,
                                     Permission.status == "Approved")
    if user_id: q = q.filter(Permission.user_id == user_id)
    perms = q.order_by(Permission.date).all()

    rows = []
    for p in perms:
        u = db.query(User).filter_by(id=p.user_id).first()
        values = {"employee": u.name if u else "—", "start_date": str(p.date), "end_date": str(p.date),
                  "hours": p.hours or 0, "permission_reason": p.reason}
        rows.append([values[k] for k in keys])

    return _build_xlsx("Permission Report", headers, rows,
                        f"permission-report-{start}-to-{end}.xlsx", _period_label(start, end) + " | Approved permissions only")
