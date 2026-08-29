"""
Proposal Estimate API — Phase 1
Endpoints:
  GET    /proposal-estimates               list (with filters)
  POST   /proposal-estimates               create
  GET    /proposal-estimates/{id}          detail
  PATCH  /proposal-estimates/{id}          update header
  DELETE /proposal-estimates/{id}          delete
  POST   /proposal-estimates/{id}/submit   Draft → Submitted
  POST   /proposal-estimates/{id}/approve  Submitted → Approved
  POST   /proposal-estimates/{id}/reject   Submitted → Rejected
  POST   /proposal-estimates/{id}/archive  → Archived

  GET    /proposal-estimates/{id}/sections             list sections
  PUT    /proposal-estimates/{id}/sections/{stype}     upsert section

  GET    /proposal-estimates/{id}/reports              list report rows
  POST   /proposal-estimates/{id}/reports              add report row
  PATCH  /proposal-estimates/{id}/reports/{rid}        update row
  DELETE /proposal-estimates/{id}/reports/{rid}        delete row

  GET    /proposal-estimates/{id}/features             get feature toggles
  PUT    /proposal-estimates/{id}/features             upsert feature toggles

Access Control:
  Admin, PM          — create, edit, delete, approve
  FC Lead, TC Lead   — create, edit, delete  (no approve)
  All others         — view only
"""

from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session, selectinload

from app.db.database import get_db
from app.models.models import (
    User, ProposalEstimate, ProposalSection, ProposalReport, ProposalFeatures,
    ProposalEstimationRow, ProposalAuditLog,
)
from app.api.routes.auth import get_current_user
from app.core.permissions import is_admin

router = APIRouter(tags=["proposal-estimates"])

# ─── Permission helpers ────────────────────────────────────────────────────────

_CAN_EDIT_ROLES  = {"Admin", "Project Manager", "FC Lead", "TC Lead"}
_CAN_APPROVE_ROLES = {"Admin", "Project Manager"}


def _can_edit(user: User) -> bool:
    return user.role in _CAN_EDIT_ROLES


def _can_approve(user: User) -> bool:
    return user.role in _CAN_APPROVE_ROLES


def _require_edit(user: User):
    if not _can_edit(user):
        raise HTTPException(403, "You do not have permission to modify proposals")


def _require_approve(user: User):
    if not _can_approve(user):
        raise HTTPException(403, "Only Admin or Project Manager can approve proposals")


def _require_delete(proposal: ProposalEstimate, user: User):
    """Admin/PM can delete any; FC Lead/TC Lead can delete proposals they created."""
    if is_admin(user) or user.role == "Project Manager":
        return
    if user.role in {"FC Lead", "TC Lead"} and proposal.created_by == user.id:
        return
    raise HTTPException(403, "You do not have permission to delete this proposal")


# ─── Proposal Number helpers ──────────────────────────────────────────────────

def _get_financial_year() -> tuple:
    """Return (fy_start, fy_end) for the current Indian financial year (Apr–Mar)."""
    from datetime import date
    today = date.today()
    fy_start = today.year if today.month >= 4 else today.year - 1
    return fy_start, fy_start + 1


def _generate_proposal_number(db) -> str:
    """Return the next sequential proposal number for the current financial year.
    Format: YYYY-YYYY+1/NNN  e.g. 2026-2027/001
    The UNIQUE constraint on the column prevents duplicates on concurrent requests."""
    fy_start, fy_end = _get_financial_year()
    fy_prefix = f"{fy_start}-{fy_end}"
    existing_count = db.query(ProposalEstimate).filter(
        ProposalEstimate.proposal_number.like(f"{fy_prefix}/%")
    ).count()
    return f"{fy_prefix}/{existing_count + 1:03d}"


# ─── Serializers ──────────────────────────────────────────────────────────────

def _ser_proposal(p: ProposalEstimate) -> dict:
    return {
        "id":               p.id,
        "client_name":      p.client_name,
        "project_name":     p.project_name,
        "project_category": p.project_category,
        "status":           p.status,
        "version":          p.version,
        "estimation_mode":  p.estimation_mode or "hours",
        "created_by":       p.created_by,
        "creator_name":     p.creator.name if p.creator else None,
        "approved_by":      p.approved_by,
        "approver_name":    p.approver.name if p.approver else None,
        "created_at":       p.created_at.isoformat() if p.created_at else None,
        "updated_at":       p.updated_at.isoformat() if p.updated_at else None,
        "submitted_at":     p.submitted_at.isoformat() if p.submitted_at else None,
        "approved_at":      p.approved_at.isoformat() if p.approved_at else None,
        "bd_status":        p.bd_status,
        "bd_status_date":   p.bd_status_date.isoformat() if p.bd_status_date else None,
        "proposal_number":  p.proposal_number,
        "estimation_total_cost": sum(r.total_cost or 0 for r in p.estimation_rows) or None,
    }


def _ser_section(s: ProposalSection) -> dict:
    return {
        "id":           s.id,
        "proposal_id":  s.proposal_id,
        "section_type": s.section_type,
        "content":      s.content,
    }


def _ser_report(r: ProposalReport) -> dict:
    import json
    def _parse(v):
        if not v:
            return []
        try:
            return json.loads(v)
        except Exception:
            return []
    return {
        "id":                    r.id,
        "proposal_id":           r.proposal_id,
        "sl_no":                 r.sl_no,
        "report_name":           r.report_name,
        "frequency":             r.frequency,
        "output_automated":      r.output_automated,
        "output_methods":        _parse(r.output_methods),
        "rough_input":           r.rough_input,
        "input_form":            _parse(r.input_form),
        "input_gen_automated":   r.input_gen_automated,
        "data_validated":        r.data_validated,
        "validation_complexity": r.validation_complexity,
    }


def _ser_features(f: ProposalFeatures) -> dict:
    import json
    def _parse(v):
        if not v:
            return []
        try:
            return json.loads(v)
        except Exception:
            return []
    # answers is stored as JSON (dict); return {} if not set
    answers = f.answers if isinstance(f.answers, dict) else {}
    return {
        "id":                 f.id,
        "proposal_id":        f.proposal_id,
        "feat_viz":           f.feat_viz,
        "feat_viz_types":     _parse(f.feat_viz_types),
        "feat_alerts":        f.feat_alerts,
        "feat_alerts_detail": f.feat_alerts_detail,
        "feat_access":        f.feat_access,
        "feat_access_detail": f.feat_access_detail,
        "feat_rules":         f.feat_rules,
        "feat_rules_detail":  f.feat_rules_detail,
        "feat_mobile":        f.feat_mobile,
        "feat_mobile_detail": f.feat_mobile_detail,
        "feat_master":        f.feat_master,
        "feat_master_detail": f.feat_master_detail,
        "feat_audit":         f.feat_audit,
        "feat_audit_detail":  f.feat_audit_detail,
        "answers":            answers,
    }


# ─── Pydantic schemas ──────────────────────────────────────────────────────────

class ProposalCreate(BaseModel):
    client_name:      str
    project_name:     Optional[str] = None
    project_category: Optional[str] = None


class ProposalUpdate(BaseModel):
    client_name:      Optional[str] = None
    project_name:     Optional[str] = None
    project_category: Optional[str] = None
    bd_status:        Optional[str] = None
    bd_status_date:   Optional[str] = None  # ISO date string "YYYY-MM-DD" or ""


class SectionUpsert(BaseModel):
    content: Optional[str] = None


class ReportCreate(BaseModel):
    sl_no:                 Optional[int]  = None
    report_name:           str
    frequency:             Optional[str]  = None
    output_automated:      Optional[bool] = False
    output_methods:        Optional[list] = []
    rough_input:           Optional[str]  = None
    input_form:            Optional[list] = []
    input_gen_automated:   Optional[bool] = False
    data_validated:        Optional[bool] = False
    validation_complexity: Optional[str]  = None


class ReportUpdate(BaseModel):
    sl_no:                 Optional[int]  = None
    report_name:           Optional[str]  = None
    frequency:             Optional[str]  = None
    output_automated:      Optional[bool] = None
    output_methods:        Optional[list] = None
    rough_input:           Optional[str]  = None
    input_form:            Optional[list] = None
    input_gen_automated:   Optional[bool] = None
    data_validated:        Optional[bool] = None
    validation_complexity: Optional[str]  = None


class FeaturesUpsert(BaseModel):
    feat_viz:            Optional[bool] = None
    feat_viz_types:      Optional[list] = None
    feat_alerts:         Optional[bool] = None
    feat_alerts_detail:  Optional[str]  = None
    feat_access:         Optional[bool] = None
    feat_access_detail:  Optional[str]  = None
    feat_rules:          Optional[bool] = None
    feat_rules_detail:   Optional[str]  = None
    feat_mobile:         Optional[bool] = None
    feat_mobile_detail:  Optional[str]  = None
    feat_master:         Optional[bool] = None
    feat_master_detail:  Optional[str]  = None
    feat_audit:          Optional[bool] = None
    feat_audit_detail:   Optional[str]  = None
    answers:             Optional[dict] = None  # structured Q&A: question_id → answer string


# ─── Internal query helper ────────────────────────────────────────────────────

def _get_proposal(db: Session, proposal_id: int) -> "ProposalEstimate | None":
    """Load a proposal with creator+approver eagerly to avoid N+1 on serialisation."""
    return (
        db.query(ProposalEstimate)
        .options(
            selectinload(ProposalEstimate.creator),
            selectinload(ProposalEstimate.approver),
        )
        .filter_by(id=proposal_id)
        .first()
    )


# ─── Proposal CRUD ────────────────────────────────────────────────────────────

@router.get("/proposal-estimates/team-members")
def list_team_members(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return all active users for the Role/Team dropdown in Estimation."""
    users = (
        db.query(User)
        .filter(User.is_active == True)
        .order_by(User.name)
        .all()
    )
    return [
        {"id": u.id, "name": u.name, "role": u.role, "cost_rate": u.cost_rate or 0}
        for u in users
    ]


@router.get("/proposal-estimates")
def list_proposals(
    status:    Optional[str] = None,
    category:  Optional[str] = None,
    bd_status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = (
        db.query(ProposalEstimate)
        .options(
            selectinload(ProposalEstimate.creator),
            selectinload(ProposalEstimate.approver),
            selectinload(ProposalEstimate.estimation_rows),
        )
    )
    if status:
        q = q.filter(ProposalEstimate.status == status)
    if category:
        q = q.filter(ProposalEstimate.project_category == category)
    if bd_status:
        q = q.filter(ProposalEstimate.bd_status == bd_status)
    proposals = q.order_by(ProposalEstimate.created_at.desc()).all()
    return [_ser_proposal(p) for p in proposals]


@router.post("/proposal-estimates", status_code=201)
def create_proposal(
    payload: ProposalCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_edit(current_user)
    p = ProposalEstimate(
        client_name=payload.client_name,
        project_name=payload.project_name,
        project_category=payload.project_category,
        status="Draft",
        version=1,
        created_by=current_user.id,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    # seed the 3 default sections
    for stype in ("companies", "units", "workflow"):
        db.add(ProposalSection(proposal_id=p.id, section_type=stype, content=""))
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


@router.get("/proposal-estimates/{proposal_id}")
def get_proposal(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    p = _get_proposal(db, proposal_id)
    if not p:
        raise HTTPException(404, "Proposal not found")
    return _ser_proposal(p)


@router.patch("/proposal-estimates/{proposal_id}")
def update_proposal(
    proposal_id: int,
    payload: ProposalUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_edit(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    # Editing is allowed regardless of status so CAN_EDIT_ROLES can revise
    # Approved / Archived proposals (e.g. scope corrections, V2 updates).
    if payload.client_name is not None:
        p.client_name = payload.client_name
    if payload.project_name is not None:
        p.project_name = payload.project_name
    if payload.project_category is not None:
        p.project_category = payload.project_category
    if payload.bd_status is not None:
        p.bd_status = payload.bd_status or None
        # Auto-assign a Proposal Number the first time BD Stage is set to "Proposal"
        if p.bd_status == "Proposal" and not p.proposal_number:
            p.proposal_number = _generate_proposal_number(db)
    if payload.bd_status_date is not None:
        if payload.bd_status_date == "":
            p.bd_status_date = None
        else:
            from datetime import date as _date
            try:
                p.bd_status_date = _date.fromisoformat(payload.bd_status_date)
            except ValueError:
                pass
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


@router.delete("/proposal-estimates/{proposal_id}")
def delete_proposal(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    _require_delete(p, current_user)
    db.delete(p)
    db.commit()
    return {"status": "deleted"}


# ─── Status workflow ──────────────────────────────────────────────────────────

@router.post("/proposal-estimates/{proposal_id}/submit")
def submit_proposal(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_edit(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    if p.status != "Draft":
        raise HTTPException(400, f"Cannot submit a proposal with status '{p.status}'")
    p.status = "Submitted"
    p.submitted_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


@router.post("/proposal-estimates/{proposal_id}/approve")
def approve_proposal(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_approve(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    if p.status != "Submitted":
        raise HTTPException(400, f"Cannot approve a proposal with status '{p.status}'")
    p.status = "Approved"
    p.approved_by = current_user.id
    p.approved_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


@router.post("/proposal-estimates/{proposal_id}/reject")
def reject_proposal(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_approve(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    if p.status != "Submitted":
        raise HTTPException(400, f"Cannot reject a proposal with status '{p.status}'")
    p.status = "Rejected"
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


@router.post("/proposal-estimates/{proposal_id}/archive")
def archive_proposal(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_approve(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    p.status = "Archived"
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


# ─── Sections ─────────────────────────────────────────────────────────────────

@router.get("/proposal-estimates/{proposal_id}/sections")
def list_sections(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sections = db.query(ProposalSection).filter_by(proposal_id=proposal_id).all()
    return [_ser_section(s) for s in sections]


@router.put("/proposal-estimates/{proposal_id}/sections/{section_type}")
def upsert_section(
    proposal_id:  int,
    section_type: str,
    payload: SectionUpsert,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_edit(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    s = db.query(ProposalSection).filter_by(
        proposal_id=proposal_id, section_type=section_type
    ).first()
    if s:
        s.content = payload.content
    else:
        s = ProposalSection(
            proposal_id=proposal_id,
            section_type=section_type,
            content=payload.content,
        )
        db.add(s)
    db.commit()
    db.refresh(s)
    return _ser_section(s)


# ─── Report rows ──────────────────────────────────────────────────────────────

@router.get("/proposal-estimates/{proposal_id}/reports")
def list_reports(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rows = (
        db.query(ProposalReport)
        .filter_by(proposal_id=proposal_id)
        .order_by(ProposalReport.sl_no)
        .all()
    )
    return [_ser_report(r) for r in rows]


@router.post("/proposal-estimates/{proposal_id}/reports", status_code=201)
def add_report(
    proposal_id: int,
    payload: ReportCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import json
    _require_edit(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    # auto sl_no if not provided
    sl_no = payload.sl_no
    if sl_no is None:
        max_row = (
            db.query(ProposalReport)
            .filter_by(proposal_id=proposal_id)
            .order_by(ProposalReport.sl_no.desc())
            .first()
        )
        sl_no = (max_row.sl_no + 1) if max_row else 1
    r = ProposalReport(
        proposal_id=proposal_id,
        sl_no=sl_no,
        report_name=payload.report_name,
        frequency=payload.frequency,
        output_automated=payload.output_automated or False,
        output_methods=json.dumps(payload.output_methods or []),
        rough_input=payload.rough_input,
        input_form=json.dumps(payload.input_form or []),
        input_gen_automated=payload.input_gen_automated or False,
        data_validated=payload.data_validated or False,
        validation_complexity=payload.validation_complexity,
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return _ser_report(r)


@router.patch("/proposal-estimates/{proposal_id}/reports/{report_id}")
def update_report(
    proposal_id: int,
    report_id:   int,
    payload: ReportUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import json
    _require_edit(current_user)
    r = db.query(ProposalReport).filter_by(
        id=report_id, proposal_id=proposal_id
    ).first()
    if not r:
        raise HTTPException(404, "Report row not found")
    if payload.sl_no is not None:
        r.sl_no = payload.sl_no
    if payload.report_name is not None:
        r.report_name = payload.report_name
    if payload.frequency is not None:
        r.frequency = payload.frequency
    if payload.output_automated is not None:
        r.output_automated = payload.output_automated
    if payload.output_methods is not None:
        r.output_methods = json.dumps(payload.output_methods)
    if payload.rough_input is not None:
        r.rough_input = payload.rough_input
    if payload.input_form is not None:
        r.input_form = json.dumps(payload.input_form)
    if payload.input_gen_automated is not None:
        r.input_gen_automated = payload.input_gen_automated
    if payload.data_validated is not None:
        r.data_validated = payload.data_validated
    if payload.validation_complexity is not None:
        r.validation_complexity = payload.validation_complexity
    db.commit()
    db.refresh(r)
    return _ser_report(r)


@router.delete("/proposal-estimates/{proposal_id}/reports/{report_id}")
def delete_report(
    proposal_id: int,
    report_id:   int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_edit(current_user)
    r = db.query(ProposalReport).filter_by(
        id=report_id, proposal_id=proposal_id
    ).first()
    if not r:
        raise HTTPException(404, "Report row not found")
    db.delete(r)
    db.commit()
    return {"status": "deleted"}


# ─── Features ─────────────────────────────────────────────────────────────────

@router.get("/proposal-estimates/{proposal_id}/features")
def get_features(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    f = db.query(ProposalFeatures).filter_by(proposal_id=proposal_id).first()
    if not f:
        return {}
    return _ser_features(f)


@router.put("/proposal-estimates/{proposal_id}/features")
def upsert_features(
    proposal_id: int,
    payload: FeaturesUpsert,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import json
    _require_edit(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    f = db.query(ProposalFeatures).filter_by(proposal_id=proposal_id).first()
    if not f:
        f = ProposalFeatures(proposal_id=proposal_id)
        db.add(f)

    for field, val in payload.dict(exclude_none=True).items():
        if field.endswith("_types") or field in ("feat_viz_types",):
            setattr(f, field, json.dumps(val) if isinstance(val, list) else val)
        elif field == "answers":
            # answers is a dict; SQLAlchemy JSON column handles serialization natively
            setattr(f, field, val if isinstance(val, dict) else {})
        else:
            setattr(f, field, val)

    db.commit()
    db.refresh(f)
    return _ser_features(f)


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 2 — Estimation Engine
# ═══════════════════════════════════════════════════════════════════════════════

def _to_hours(quantity: float, mode: str) -> float:
    """Convert display quantity to canonical hours. Days mode: qty × 7."""
    return quantity * 7 if mode == "days" else quantity


def _compute_total(quantity_hours: float, cost_rate: float) -> float:
    """Total cost = quantity_hours × cost_rate (always hours-based)."""
    return round((quantity_hours or 0) * (cost_rate or 0), 2)


def _ser_estimation_row(r: ProposalEstimationRow, mode: str) -> dict:
    # quantity_hours is canonical; derive display quantity from it
    qty_hours = r.quantity_hours if (r.quantity_hours is not None and r.quantity_hours > 0) else (r.quantity or 0)
    display_qty = round(qty_hours / 7, 4) if mode == "days" else qty_hours
    return {
        "id":               r.id,
        "proposal_id":      r.proposal_id,
        "sl_no":            r.sl_no,
        "description":      r.description,
        "role_description": r.role_description,
        "quantity":         display_qty,          # mode-adjusted display value
        "quantity_hours":   qty_hours,             # always in hours
        "cost_rate":        r.cost_rate,
        "total_cost":       r.total_cost,
        "unit":             "day(s)" if mode == "days" else "hr(s)",
    }


class EstimationModeUpdate(BaseModel):
    mode: str  # "hours" | "days"


class EstimationRowCreate(BaseModel):
    sl_no:            Optional[int]   = None
    description:      str
    role_description: Optional[str]   = None
    quantity:         Optional[float] = 0
    cost_rate:        Optional[float] = 0


class EstimationRowUpdate(BaseModel):
    sl_no:            Optional[int]   = None
    description:      Optional[str]   = None
    role_description: Optional[str]   = None
    quantity:         Optional[float] = None
    cost_rate:        Optional[float] = None


@router.get("/proposal-estimates/{proposal_id}/estimation")
def get_estimation(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    mode = p.estimation_mode or "hours"
    rows = (
        db.query(ProposalEstimationRow)
        .filter_by(proposal_id=proposal_id)
        .order_by(ProposalEstimationRow.sl_no)
        .all()
    )
    serialized = [_ser_estimation_row(r, mode) for r in rows]

    # ── Summary totals — always derived from quantity_hours (canonical) ──────────
    def _row_hours(r: ProposalEstimationRow) -> float:
        return r.quantity_hours if (r.quantity_hours is not None and r.quantity_hours > 0) else (r.quantity or 0)

    total_hours = sum(_row_hours(r) for r in rows)
    total_qty   = round(total_hours / 7, 4) if mode == "days" else total_hours
    total_cost  = sum(r.total_cost or 0 for r in rows)

    # ── Role-wise summary ────────────────────────────────────────────────────────
    role_map: dict = {}
    for r in rows:
        role = (r.role_description or "").strip() or "Unspecified"
        if role not in role_map:
            role_map[role] = {"hours": 0.0, "cost": 0.0}
        role_map[role]["hours"] += _row_hours(r)
        role_map[role]["cost"]  += r.total_cost or 0

    role_totals = [
        {
            "role":  role,
            "hours": round(vals["hours"], 2),
            "qty":   round(vals["hours"] / 7, 2) if mode == "days" else round(vals["hours"], 2),
            "cost":  round(vals["cost"], 2),
        }
        for role, vals in role_map.items()
    ]

    return {
        "mode":        mode,
        "rows":        serialized,
        "total_qty":   round(total_qty, 2),
        "total_hours": round(total_hours, 2),
        "total_cost":  round(total_cost, 2),
        "role_totals": role_totals,
    }


@router.patch("/proposal-estimates/{proposal_id}/estimation-mode")
def set_estimation_mode(
    proposal_id: int,
    payload: EstimationModeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_edit(current_user)
    if payload.mode not in ("hours", "days"):
        raise HTTPException(400, "mode must be 'hours' or 'days'")
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    p.estimation_mode = payload.mode
    # Mode switch is purely cosmetic — quantity_hours (canonical) is unchanged.
    # total_cost = quantity_hours × cost_rate which doesn't depend on mode.
    # No row recompute needed.
    db.commit()
    db.refresh(p)
    return {"mode": p.estimation_mode}


@router.post("/proposal-estimates/{proposal_id}/estimation", status_code=201)
def add_estimation_row(
    proposal_id: int,
    payload: EstimationRowCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_edit(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    mode = p.estimation_mode or "hours"
    sl_no = payload.sl_no
    if sl_no is None:
        last = (
            db.query(ProposalEstimationRow)
            .filter_by(proposal_id=proposal_id)
            .order_by(ProposalEstimationRow.sl_no.desc())
            .first()
        )
        sl_no = (last.sl_no + 1) if last else 1
    qty       = payload.quantity or 0
    rate      = payload.cost_rate or 0
    qty_hours = _to_hours(qty, mode)
    row = ProposalEstimationRow(
        proposal_id=proposal_id,
        sl_no=sl_no,
        description=payload.description,
        role_description=payload.role_description,
        quantity=qty,
        quantity_hours=qty_hours,
        cost_rate=rate,
        total_cost=_compute_total(qty_hours, rate),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _ser_estimation_row(row, mode)


@router.patch("/proposal-estimates/{proposal_id}/estimation/{row_id}")
def update_estimation_row(
    proposal_id: int,
    row_id:      int,
    payload: EstimationRowUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_edit(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    mode = p.estimation_mode or "hours"
    row = db.query(ProposalEstimationRow).filter_by(id=row_id, proposal_id=proposal_id).first()
    if not row:
        raise HTTPException(404, "Row not found")
    if payload.sl_no is not None:
        row.sl_no = payload.sl_no
    if payload.description is not None:
        row.description = payload.description
    if payload.role_description is not None:
        row.role_description = payload.role_description
    if payload.quantity is not None:
        row.quantity = payload.quantity
        row.quantity_hours = _to_hours(payload.quantity, mode)
    if payload.cost_rate is not None:
        row.cost_rate = payload.cost_rate
    # Derive quantity_hours for total computation (use existing canonical if quantity unchanged)
    qty_hours = row.quantity_hours if row.quantity_hours else _to_hours(row.quantity or 0, mode)
    row.total_cost = _compute_total(qty_hours, row.cost_rate or 0)
    db.commit()
    db.refresh(row)
    return _ser_estimation_row(row, mode)


@router.delete("/proposal-estimates/{proposal_id}/estimation/{row_id}")
def delete_estimation_row(
    proposal_id: int,
    row_id:      int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_edit(current_user)
    row = db.query(ProposalEstimationRow).filter_by(id=row_id, proposal_id=proposal_id).first()
    if not row:
        raise HTTPException(404, "Row not found")
    db.delete(row)
    db.commit()
    return {"status": "deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 3 — Audit Log + Version Control + Excel Export
# ═══════════════════════════════════════════════════════════════════════════════

def _log_audit(db: Session, proposal_id: int, action: str, user_id: int,
               from_status: str = None, to_status: str = None, note: str = None):
    db.add(ProposalAuditLog(
        proposal_id=proposal_id,
        action=action,
        from_status=from_status,
        to_status=to_status,
        note=note,
        changed_by=user_id,
    ))


def _ser_audit(log: ProposalAuditLog) -> dict:
    return {
        "id":          log.id,
        "action":      log.action,
        "from_status": log.from_status,
        "to_status":   log.to_status,
        "note":        log.note,
        "actor_name":  log.actor.name if log.actor else None,
        "changed_at":  log.changed_at.isoformat() if log.changed_at else None,
    }


@router.get("/proposal-estimates/{proposal_id}/audit")
def get_audit_log(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    logs = (
        db.query(ProposalAuditLog)
        .options(selectinload(ProposalAuditLog.actor))
        .filter_by(proposal_id=proposal_id)
        .order_by(ProposalAuditLog.changed_at.desc())
        .all()
    )
    return [_ser_audit(l) for l in logs]


# ── Patch status endpoints to write audit log + version-increment on re-submit ─
# We override the submit/approve/reject/archive endpoints with new versions
# that also write the audit log and handle version increment.

@router.post("/proposal-estimates/{proposal_id}/submit-v2")
def submit_proposal_v2(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Submit with audit log + version increment on re-submit (after Rejected)."""
    _require_edit(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    if p.status not in ("Draft", "Rejected"):
        raise HTTPException(400, f"Cannot submit a proposal with status '{p.status}'")
    old_status = p.status
    p.status = "Submitted"
    p.submitted_at = datetime.now(timezone.utc)
    # Increment version on re-submission after rejection
    if old_status == "Rejected":
        p.version = (p.version or 1) + 1
    _log_audit(db, proposal_id, "submitted", current_user.id,
               from_status=old_status, to_status="Submitted",
               note=f"Version {p.version}")
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


@router.post("/proposal-estimates/{proposal_id}/approve-v2")
def approve_proposal_v2(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_approve(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    if p.status != "Submitted":
        raise HTTPException(400, f"Cannot approve a proposal with status '{p.status}'")
    p.status = "Approved"
    p.approved_by = current_user.id
    p.approved_at = datetime.now(timezone.utc)
    _log_audit(db, proposal_id, "approved", current_user.id,
               from_status="Submitted", to_status="Approved")
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


@router.post("/proposal-estimates/{proposal_id}/reject-v2")
def reject_proposal_v2(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_approve(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    if p.status != "Submitted":
        raise HTTPException(400, f"Cannot reject a proposal with status '{p.status}'")
    p.status = "Rejected"
    _log_audit(db, proposal_id, "rejected", current_user.id,
               from_status="Submitted", to_status="Rejected")
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


@router.post("/proposal-estimates/{proposal_id}/archive-v2")
def archive_proposal_v2(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_approve(current_user)
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    old_status = p.status
    p.status = "Archived"
    _log_audit(db, proposal_id, "archived", current_user.id,
               from_status=old_status, to_status="Archived")
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


# ── Admin: force-status override ──────────────────────────────────────────────
_VALID_STATUSES = {"Draft", "Submitted", "Approved", "Rejected", "Archived"}

@router.patch("/proposal-estimates/{proposal_id}/force-status")
def force_proposal_status(
    proposal_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Admin-only: directly set any proposal status (e.g. Archived → Approved)."""
    if not is_admin(current_user):
        raise HTTPException(403, "Only Admin can override proposal status")
    new_status = (payload.get("status") or "").strip()
    if new_status not in _VALID_STATUSES:
        raise HTTPException(400, f"status must be one of: {', '.join(sorted(_VALID_STATUSES))}")
    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")
    old_status = p.status
    p.status = new_status
    _log_audit(db, proposal_id, "force_status", current_user.id,
               from_status=old_status, to_status=new_status,
               note=f"Admin override by {current_user.name}")
    db.commit()
    db.refresh(p)
    return _ser_proposal(p)


# ── Excel Export ───────────────────────────────────────────────────────────────
from fastapi.responses import StreamingResponse
import io


@router.get("/proposal-estimates/{proposal_id}/export/excel")
def export_proposal_excel(
    proposal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export full proposal as a multi-sheet Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import json

    p = db.query(ProposalEstimate).filter_by(id=proposal_id).first()
    if not p:
        raise HTTPException(404, "Proposal not found")

    mode = p.estimation_mode or "hours"

    wb = openpyxl.Workbook()

    # ── helpers ─────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    SUBHDR_FILL = PatternFill("solid", fgColor="E8F0FE")
    SUBHDR_FONT = Font(bold=True, color="1E3A5F", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _h(ws, row, col, val, fill=None, font=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        if fill: c.fill = fill
        if font: c.font = font
        c.border = border
        if wrap: c.alignment = Alignment(wrap_text=True, vertical="top")
        return c

    def _w(ws, col, width):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Sheet 1: Overview ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overview"

    ws1.merge_cells("A1:D1")
    c = ws1.cell(row=1, column=1, value="PROPOSAL ESTIMATE — OVERVIEW")
    c.font = Font(bold=True, size=14, color="1E3A5F")
    c.fill = PatternFill("solid", fgColor="DBEAFE")
    c.alignment = Alignment(horizontal="center")

    fields = [
        ("Client Name",      p.client_name),
        ("Project Name",     p.project_name or "—"),
        ("Project Category", p.project_category or "—"),
        ("Status",           p.status),
        ("Version",          f"v{p.version}"),
        ("Created By",       p.creator.name if p.creator else "—"),
        ("Created Date",     p.created_at.strftime("%d %b %Y") if p.created_at else "—"),
        ("Submitted Date",   p.submitted_at.strftime("%d %b %Y") if p.submitted_at else "—"),
        ("Approved By",      p.approver.name if p.approver else "—"),
        ("Approved Date",    p.approved_at.strftime("%d %b %Y") if p.approved_at else "—"),
    ]
    for i, (label, val) in enumerate(fields, start=3):
        _h(ws1, i, 1, label, fill=SUBHDR_FILL, font=SUBHDR_FONT)
        _h(ws1, i, 2, val)

    _w(ws1, 1, 22); _w(ws1, 2, 40)

    # ── Sheet 2: Scope ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Scope")
    ws2.merge_cells("A1:B1")
    c = ws2.cell(row=1, column=1, value="SCOPE OF WORK")
    c.font = Font(bold=True, size=14, color="1E3A5F")
    c.fill = PatternFill("solid", fgColor="DBEAFE")
    c.alignment = Alignment(horizontal="center")

    sections = db.query(ProposalSection).filter_by(proposal_id=proposal_id).all()
    sec_map = {s.section_type: s.content or "" for s in sections}
    for r, (stype, label) in enumerate([
        ("companies", "Companies / Clients Overview"),
        ("units",     "Business Units / Departments"),
        ("workflow",  "Business Workflow Description"),
    ], start=3):
        _h(ws2, r, 1, label, fill=SUBHDR_FILL, font=SUBHDR_FONT)
        c2 = ws2.cell(row=r, column=2, value=sec_map.get(stype, ""))
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        c2.border = border
        ws2.row_dimensions[r].height = 80

    _w(ws2, 1, 30); _w(ws2, 2, 80)

    # ── Sheet 3: Reports ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Reports")
    rpt_headers = [
        "Sl.", "Report Name", "Frequency", "Output Methods",
        "Output Auto?", "Input Format", "Input Auto?", "Data Validated?", "Validation Complexity"
    ]
    for ci, h in enumerate(rpt_headers, 1):
        _h(ws3, 1, ci, h, fill=HDR_FILL, font=HDR_FONT)

    report_rows = (
        db.query(ProposalReport)
        .filter_by(proposal_id=proposal_id)
        .order_by(ProposalReport.sl_no)
        .all()
    )
    def _parse_json(v):
        if not v: return ""
        try: return ", ".join(json.loads(v))
        except: return v

    for ri, rw in enumerate(report_rows, 2):
        vals = [
            rw.sl_no, rw.report_name, rw.frequency or "",
            _parse_json(rw.output_methods), "Yes" if rw.output_automated else "No",
            _parse_json(rw.input_form), "Yes" if rw.input_gen_automated else "No",
            "Yes" if rw.data_validated else "No", rw.validation_complexity or "",
        ]
        for ci, v in enumerate(vals, 1):
            _h(ws3, ri, ci, v)

    for ci, w in enumerate([6, 35, 14, 22, 12, 22, 12, 16, 30], 1):
        _w(ws3, ci, w)

    # ── Sheet 4: Estimation ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Estimation")
    unit_label = "Man-Days" if mode == "days" else "Hours"
    est_headers = ["Sl.", "Description / Work Item", "Role / Team", unit_label, "Rate (₹/hr)", "Total Cost (₹)"]
    for ci, h in enumerate(est_headers, 1):
        _h(ws4, 1, ci, h, fill=HDR_FILL, font=HDR_FONT)

    est_rows = (
        db.query(ProposalEstimationRow)
        .filter_by(proposal_id=proposal_id)
        .order_by(ProposalEstimationRow.sl_no)
        .all()
    )
    for ri, er in enumerate(est_rows, 2):
        vals = [er.sl_no, er.description, er.role_description or "", er.quantity, er.cost_rate, er.total_cost]
        for ci, v in enumerate(vals, 1):
            _h(ws4, ri, ci, v)

    # Totals row
    tr = len(est_rows) + 2
    ws4.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws4.cell(row=tr, column=5, value="").font = Font(bold=True)
    total_cost = sum(r.total_cost or 0 for r in est_rows)
    tc = ws4.cell(row=tr, column=6, value=round(total_cost, 2))
    tc.font = Font(bold=True, color="1E3A5F")
    tc.fill = PatternFill("solid", fgColor="DBEAFE")

    # Mode note
    ws4.cell(row=tr + 2, column=1,
             value=f"Estimation Mode: {mode.title()} ({'1 day = 7 hours' if mode == 'days' else 'direct hours'})")

    for ci, w in enumerate([6, 40, 25, 14, 16, 18], 1):
        _w(ws4, ci, w)

    # ── Sheet 5: Features ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("Features")
    feats = db.query(ProposalFeatures).filter_by(proposal_id=proposal_id).first()

    ws5.merge_cells("A1:C1")
    c = ws5.cell(row=1, column=1, value="REQUIRED FEATURES")
    c.font = Font(bold=True, size=14, color="1E3A5F")
    c.fill = PatternFill("solid", fgColor="DBEAFE")
    c.alignment = Alignment(horizontal="center")

    # Mirrors frontend FEATURE_QUESTIONS config: feature_key → [(question_id, question_text)]
    FEATURE_QUESTION_DEFS = {
        "feat_viz": [
            ("viz_dashboards_required", "Are dashboards required?"),
            ("viz_dashboard_count",     "How many dashboards are required?"),
            ("viz_drill_down",          "Is multi-level drill-down required? (Group → Division → Region → Entity)"),
            ("viz_chart_types",         "What type of charts are required? (Trend lines, comparisons, heatmaps, funnel views, etc.)"),
            ("viz_date_ranges",         "Are configurable date ranges and period-over-period comparisons required? (MoM, YoY, etc.)"),
        ],
        "feat_alerts": [
            ("alerts_threshold",         "Threshold-based alerts required? (e.g., price deviation, KPI breach)"),
            ("alerts_anomaly",           "Anomaly flagging required?"),
            ("alerts_exception_reports", "Auto-generated exception reports required?"),
        ],
        "feat_access": [
            ("access_rbac",          "Role-based access control (RBAC) — admin, manager, user, view-only tiers?"),
            ("access_org_hierarchy", "Multi-level org hierarchy support (branch/region/division/entity)?"),
            ("access_sso",           "SSO / Google-Microsoft login, or email+OTP?"),
            ("access_provisioning",  "User provisioning/deprovisioning and password reset flows?"),
        ],
        "feat_rules": [
            ("rules_formulas",   "Admin-configurable formulas/weightages (e.g., KPI scoring)?"),
            ("rules_versioning", "Rule versioning (historical calculations don't break on rule change)?"),
            ("rules_scenario",   "Scenario/what-if testing before rule rollout?"),
        ],
        "feat_mobile": [
            ("mobile_required", "Is a mobile app required? (Yes / No)"),
            ("mobile_scope",    "If yes — dashboard view only, or full app logic?"),
        ],
        "feat_master": [
            ("master_configurable", "Admin-configurable master data (departments, categories, products, locations)?"),
            ("master_dropdowns",    "Configurable dropdowns/lookups instead of hardcoded values?"),
            ("master_bulk_import",  "Bulk import/export (usually Excel-based)?"),
        ],
        "feat_audit": [
            ("audit_logging",         "Who-changed-what-when logging required?"),
            ("audit_version_history", "Version history on key records?"),
            ("audit_data_retention",  "Data retention/archival rules?"),
        ],
    }

    FEATURE_DEFS = [
        ("feat_viz",    "Data Visualization",            "feat_viz_types"),
        ("feat_alerts", "Alerts & Notifications",        "feat_alerts_detail"),
        ("feat_access", "Role-Based Access Control",     "feat_access_detail"),
        ("feat_rules",  "Business Rules / Conditional",  "feat_rules_detail"),
        ("feat_mobile", "Mobile Access",                 "feat_mobile_detail"),
        ("feat_master", "Master Data Management",        "feat_master_detail"),
        ("feat_audit",  "Audit Trail / History Log",     "feat_audit_detail"),
    ]
    _h(ws5, 2, 1, "Feature", fill=HDR_FILL, font=HDR_FONT)
    _h(ws5, 2, 2, "Required?", fill=HDR_FILL, font=HDR_FONT)
    _h(ws5, 2, 3, "Details / Notes", fill=HDR_FILL, font=HDR_FONT)

    # Pull saved answers dict (question_id → answer text)
    saved_answers = {}
    if feats and isinstance(feats.answers, dict):
        saved_answers = feats.answers

    for ri, (toggle, label, detail_field) in enumerate(FEATURE_DEFS, 3):
        enabled = getattr(feats, toggle, False) if feats else False
        detail  = getattr(feats, detail_field, "") or "" if feats else ""
        if detail_field == "feat_viz_types" and isinstance(detail, str) and detail.startswith("["):
            try: detail = ", ".join(json.loads(detail))
            except: pass

        # Build combined notes: optional legacy detail + structured Q&A
        parts = []
        if detail:
            parts.append(detail)

        qa_lines = []
        for qid, qtext in FEATURE_QUESTION_DEFS.get(toggle, []):
            ans = saved_answers.get(qid, "") or ""
            qa_lines.append(f"Q: {qtext}\nA: {ans if ans else '—'}")
        if qa_lines:
            parts.append("\n\n".join(qa_lines))

        combined_notes = "\n\n".join(parts)

        _h(ws5, ri, 1, label)
        c_req = ws5.cell(row=ri, column=2, value="Yes" if enabled else "No")
        c_req.font = Font(color="1E7A4A" if enabled else "888888", bold=enabled)
        c_req.border = border
        c_notes = ws5.cell(row=ri, column=3, value=combined_notes)
        c_notes.alignment = Alignment(wrap_text=True, vertical="top")
        c_notes.border = border
        # Auto-height: ~15pt per line of text
        line_count = combined_notes.count("\n") + 1
        ws5.row_dimensions[ri].height = max(40, 15 * line_count)

    for ci, w in enumerate([35, 12, 70], 1):
        _w(ws5, ci, w)

    # ── Stream out ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = p.client_name.replace(" ", "_")[:30]
    filename = f"Proposal_{safe}_v{p.version}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
