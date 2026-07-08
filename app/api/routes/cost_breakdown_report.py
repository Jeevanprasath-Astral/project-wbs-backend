"""Cost Category Breakdown Report — per-project cost distribution by category.

Sheet 1 — Project Summary (one row per project):
  Project | Client | Status | Budget (₹) | Total Cost (₹)
  | Budget Used % | Remaining (₹)

  Row colour:
    Green (E2EFDA) — total cost < budget  (under budget)
    Red   (FCE4D6) — total cost ≥ budget  (at or over budget)
    Grey  (F2F2F2) — no budget set

Sheet 2 — Category Detail (one row per project × category):
  Project | Category | Cost (₹) | Share of Project %
"""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.database import get_db
from app.models.models import ProjectCost, Project, User
from app.core.deps import get_current_user

router = APIRouter(prefix="/cost-breakdown-report", tags=["Cost Breakdown Report"])


# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL   = _fill("1F3864")
COL_FILL   = _fill("BDD7EE")
EVEN_FILL  = _fill("EBF3FB")
ODD_FILL   = _fill("FFFFFF")
TOT_FILL   = _fill("D9E8F5")
GREEN_FILL = _fill("E2EFDA")
RED_FILL   = _fill("FCE4D6")
GREY_FILL  = _fill("F2F2F2")
CAT_FILL   = _fill("EDE7F6")   # light purple tint for category rows


def _write_title(ws, title, subtitle, ncols):
    last_col = chr(ord("A") + ncols - 1)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    if subtitle:
        ws.merge_cells(f"A2:{last_col}2")
        c2 = ws["A2"]
        c2.value = subtitle
        c2.font = Font(italic=True, color="555555", size=9, name="Calibri")
        c2.fill = _fill("D9E8F5")
        c2.alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 16
        return 4
    return 3


def _write_header(ws, headers, hrow):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(hrow, col, h)
        cell.font = Font(bold=True, color="1F3864", size=9, name="Calibri")
        cell.fill = COL_FILL
        cell.border = _bdr()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hrow].height = 28


def _write_row(ws, row_num, values, fill, right_cols=None, bold=False):
    right_cols = right_cols or []
    for col, val in enumerate(values, 1):
        cell = ws.cell(row_num, col, val if val not in (None, "") else "—")
        cell.font = Font(size=9, name="Calibri", bold=bold)
        cell.fill = fill
        cell.border = _bdr()
        align = "right" if (isinstance(val, (int, float)) or col in right_cols) else "left"
        cell.alignment = Alignment(horizontal=align)


def _fmt(v, dec=2):
    if v is None:
        return 0.0
    return round(float(v), dec)


def _pct(num, den, dec=1):
    if not den:
        return ""
    return round(float(num) / float(den) * 100, dec)


# ── Export ────────────────────────────────────────────────────────────────────
@router.get("/export")
def cost_breakdown_export(
    project_id: Optional[int] = None,
    status:     Optional[str] = None,
    start_date: Optional[str] = None,
    end_date:   Optional[str] = None,
    category:   Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import date as date_cls

    # ── Fetch projects ────────────────────────────────────────────────────────
    pq = db.query(Project)
    if project_id:
        pq = pq.filter(Project.id == project_id)
    if status:
        pq = pq.filter(Project.status == status)
    projects = pq.order_by(Project.name).all()
    proj_map = {p.id: p for p in projects}

    if not projects:
        project_ids = []
    else:
        project_ids = [p.id for p in projects]

    # ── Fetch costs ───────────────────────────────────────────────────────────
    cq = db.query(ProjectCost).filter(ProjectCost.project_id.in_(project_ids))
    if start_date:
        cq = cq.filter(ProjectCost.date >= date_cls.fromisoformat(start_date))
    if end_date:
        cq = cq.filter(ProjectCost.date <= date_cls.fromisoformat(end_date))
    if category:
        cq = cq.filter(ProjectCost.category == category)
    costs = cq.all()

    # ── Aggregate: {pid: {cat: total}} ───────────────────────────────────────
    proj_cat = {}   # {pid: {category: cost}}
    proj_total = {} # {pid: total_cost}
    for c in costs:
        pid = c.project_id
        cat = c.category or "Uncategorised"
        proj_cat.setdefault(pid, {})
        proj_cat[pid][cat] = proj_cat[pid].get(cat, 0.0) + float(c.cost or 0)
        proj_total[pid] = proj_total.get(pid, 0.0) + float(c.cost or 0)

    # ── Subtitle ──────────────────────────────────────────────────────────────
    parts = []
    if start_date or end_date:
        parts.append(f"Period: {start_date or '—'} → {end_date or '—'}")
    if status:
        parts.append(f"Status: {status}")
    if category:
        parts.append(f"Category: {category}")
    subtitle = " | ".join(parts)

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Sheet 1: Project Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Project Summary"

    HEADERS_S = [
        "Project", "Client", "Status",
        "Budget (₹)", "Total Cost (₹)", "Budget Used %", "Remaining (₹)",
    ]
    hrow = _write_title(ws1, "Cost Breakdown Report — Project Summary", subtitle, len(HEADERS_S))
    _write_header(ws1, HEADERS_S, hrow)

    grand_budget = grand_cost = 0.0
    row_num = hrow + 1
    for i, p in enumerate(projects):
        total = _fmt(proj_total.get(p.id, 0.0))
        budget = _fmt(p.budget or 0)
        used_pct = _pct(total, budget)
        remaining = _fmt(budget - total)

        grand_budget += budget
        grand_cost   += total

        if budget <= 0:
            bg = GREY_FILL
        elif total >= budget:
            bg = RED_FILL
        else:
            bg = GREEN_FILL

        STATUS_COLORS = {}  # using fill only, no badge in Excel
        _write_row(ws1, row_num, [
            p.name, p.client or "—", p.status or "—",
            budget if budget > 0 else "—",
            total,
            f"{used_pct}%" if used_pct != "" else "—",
            remaining if budget > 0 else "—",
        ], bg, right_cols=[4, 5, 6, 7])
        row_num += 1

    # Grand total
    if projects:
        grand_util = _pct(grand_cost, grand_budget)
        _write_row(ws1, row_num, [
            "TOTAL", "", "",
            _fmt(grand_budget) if grand_budget else "—",
            _fmt(grand_cost),
            f"{grand_util}%" if grand_util != "" else "—",
            _fmt(grand_budget - grand_cost) if grand_budget else "—",
        ], TOT_FILL, right_cols=[4, 5, 6, 7], bold=True)
        row_num += 1

    if not projects:
        ws1.merge_cells(f"A{row_num}:G{row_num}")
        ws1.cell(row_num, 1, "No projects found for the selected filters")
        ws1.cell(row_num, 1).font = Font(italic=True, size=9, color="999999")

    # Legend
    note_row = row_num + 2
    ws1.cell(note_row, 1, "Colour key:")
    ws1.cell(note_row, 1).font = Font(bold=True, size=8, name="Calibri")
    for col, (label, fg) in enumerate([
        ("Under budget", "E2EFDA"),
        ("Over / at budget", "FCE4D6"),
        ("No budget set", "F2F2F2"),
    ], 2):
        c = ws1.cell(note_row, col, label)
        c.font = Font(size=8, name="Calibri")
        c.fill = _fill(fg)
        c.border = _bdr()
        c.alignment = Alignment(horizontal="center")

    # Column widths Sheet 1
    for i, w in enumerate([30, 20, 14, 16, 16, 13, 16], 1):
        ws1.column_dimensions[chr(ord("A") + i - 1)].width = w

    # ── Sheet 2: Category Detail ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Category Detail")

    HEADERS_C = ["Project", "Client", "Category", "Cost (₹)", "Share of Project %"]
    hrow2 = _write_title(ws2, "Cost Breakdown Report — By Category", subtitle, len(HEADERS_C))
    _write_header(ws2, HEADERS_C, hrow2)

    row_num2 = hrow2 + 1
    any_rows = False

    for p in projects:
        pid = p.id
        cats = proj_cat.get(pid)
        if not cats:
            continue

        proj_t = proj_total.get(pid, 0.0)
        sorted_cats = sorted(cats.items(), key=lambda x: -x[1])  # desc by cost

        for j, (cat, cost) in enumerate(sorted_cats):
            share = _pct(cost, proj_t)
            bg = EVEN_FILL if j % 2 == 0 else ODD_FILL
            _write_row(ws2, row_num2, [
                p.name, p.client or "—", cat,
                _fmt(cost),
                f"{share}%" if share != "" else "—",
            ], bg, right_cols=[4, 5])
            row_num2 += 1
            any_rows = True

        # Project subtotal
        _write_row(ws2, row_num2, [
            f"  {p.name} — Subtotal", "", "All Categories",
            _fmt(proj_t), "100%",
        ], CAT_FILL, right_cols=[4, 5], bold=True)
        row_num2 += 1

    if not any_rows:
        ws2.merge_cells(f"A{row_num2}:E{row_num2}")
        ws2.cell(row_num2, 1, "No cost entries found for the selected filters")
        ws2.cell(row_num2, 1).font = Font(italic=True, size=9, color="999999")

    # Column widths Sheet 2
    for i, w in enumerate([32, 20, 26, 16, 18], 1):
        ws2.column_dimensions[chr(ord("A") + i - 1)].width = w

    # ── Stream ────────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cost-breakdown-report.xlsx"},
    )


# ── Filter options ────────────────────────────────────────────────────────────
@router.get("/filter-options")
def filter_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    projects = db.query(Project).order_by(Project.name).all()
    statuses = (
        db.query(Project.status)
        .filter(Project.status.isnot(None))
        .distinct()
        .order_by(Project.status)
        .all()
    )
    categories = (
        db.query(ProjectCost.category)
        .filter(ProjectCost.category.isnot(None))
        .distinct()
        .order_by(ProjectCost.category)
        .all()
    )
    return {
        "projects":   [{"id": p.id, "name": p.name} for p in projects],
        "statuses":   [s[0] for s in statuses],
        "categories": [c[0] for c in categories],
    }
