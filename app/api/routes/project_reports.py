"""Project-level Report exports.

Two Excel report exports spanning all projects (global scope):

  1. Budgeted vs Actual Hours Report — one row per CustomMilestone, with
     Budgeted Hours computed as SUM(subtask.estimated_hours) under the
     milestone and Actual Hours as SUM(work_hours.hours_spent) that reference
     this milestone's tree (milestone-level + task-level + subtask-level +
     activity-level work hours).

  2. Timeline Report — one row per CustomMilestone showing planned vs actual
     end dates and the free-text Schedule Variance Reason field, useful for
     identifying and explaining schedule deviations.

Both endpoints accept data-filter query params that narrow which milestones
appear in the output (project_id, assignee, team_id, status, start_date,
end_date). start_date/end_date filter on the milestone's actual_start/actual_end
for the Budgeted report and on planned_end for the Timeline report.

Mirrors the openpyxl → BytesIO → StreamingResponse convention from
timesheet_reports.py.
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import Optional
from datetime import date as date_cls
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.database import get_db
from app.models.models import (CustomMilestone, CustomTask, CustomSubtask,
                                Activity, WorkHours, Project, User)
from app.core.deps import get_current_user

router = APIRouter(prefix="/project-reports", tags=["Project Reports"])


# ── Shared xlsx builder (same style as timesheet_reports.py) ─────────────────
def _build_xlsx(title: str, headers: list, rows: list, filename: str, subtitle: str = ""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
    def bdr():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    HDR_FILL = fill("1F3864"); COL_FILL = fill("BDD7EE")
    EVEN_FILL = fill("EBF3FB"); ODD_FILL = fill("FFFFFF")

    ncols = len(headers)
    last_col_letter = chr(ord("A") + ncols - 1)

    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.fill = HDR_FILL; c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    if subtitle:
        ws.merge_cells(f"A2:{last_col_letter}2")
        c2 = ws["A2"]
        c2.value = subtitle
        c2.font = Font(italic=True, color="555555", size=9, name="Calibri")
        c2.fill = fill("D9E8F5"); c2.alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 16
        header_row = 4
    else:
        header_row = 3

    for col, h in enumerate(headers, 1):
        hc = ws.cell(header_row, col, h)
        hc.font = Font(bold=True, color="1F3864", size=9, name="Calibri")
        hc.fill = COL_FILL; hc.border = bdr()
        hc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[header_row].height = 18

    row = header_row + 1
    for i, r in enumerate(rows):
        bg = EVEN_FILL if i % 2 == 0 else ODD_FILL
        for col, val in enumerate(r, 1):
            cell = ws.cell(row, col, val if val not in (None, "") else "—")
            cell.font = Font(size=9, name="Calibri")
            cell.fill = bg; cell.border = bdr()
        row += 1

    if not rows:
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        ws.cell(row, 1, "No records for the selected filters").font = Font(italic=True, size=9, color="999999")

    for col_idx in range(1, ncols + 1):
        col_letter = chr(ord("A") + col_idx - 1)
        ws.column_dimensions[col_letter].width = 22

    output = BytesIO()
    wb.save(output); output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


def _fmt_dt(dt):
    """Return YYYY-MM-DD string or empty string."""
    if not dt:
        return ""
    try:
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(dt)[:10]


def _milestone_actual_hours(db: Session, ms_id: int) -> float:
    """Sum all work_hours that are part of this milestone's tree:
    direct milestone hours + task hours (tasks under this milestone) +
    subtask hours + activity hours. Uses subquery IDs to stay in one
    pass per level rather than per-row Python loops."""
    task_ids = [t.id for t in db.query(CustomTask.id).filter_by(milestone_id=ms_id).all()]
    subtask_ids = [s.id for s in db.query(CustomSubtask.id).filter(
        CustomSubtask.task_id.in_(task_ids)).all()] if task_ids else []
    activity_ids = [a.id for a in db.query(Activity.id).filter(
        Activity.subtask_id.in_(subtask_ids)).all()] if subtask_ids else []

    conditions = [WorkHours.custom_milestone_id == ms_id]
    if task_ids:     conditions.append(WorkHours.custom_task_id.in_(task_ids))
    if subtask_ids:  conditions.append(WorkHours.custom_subtask_id.in_(subtask_ids))
    if activity_ids: conditions.append(WorkHours.activity_id.in_(activity_ids))
    result = db.query(func.coalesce(func.sum(WorkHours.hours_spent), 0.0)).filter(
        or_(*conditions)).scalar()
    return round(float(result or 0), 2)


def _milestone_budgeted_hours(db: Session, ms_id: int) -> float:
    """Sum estimated_hours from subtasks + activities under this milestone."""
    task_ids = [t.id for t in db.query(CustomTask.id).filter_by(milestone_id=ms_id).all()]
    subtask_ids = [s.id for s in db.query(CustomSubtask.id).filter(
        CustomSubtask.task_id.in_(task_ids)).all()] if task_ids else []

    sub_hrs = db.query(func.coalesce(func.sum(CustomSubtask.estimated_hours), 0.0)).filter(
        CustomSubtask.id.in_(subtask_ids)).scalar() if subtask_ids else 0.0
    act_hrs = db.query(func.coalesce(func.sum(Activity.estimated_hours), 0.0)).filter(
        Activity.subtask_id.in_(subtask_ids)).scalar() if subtask_ids else 0.0
    return round(float(sub_hrs or 0) + float(act_hrs or 0), 2)


def _parse_date(s: Optional[str]):
    if not s:
        return None
    try:
        return date_cls.fromisoformat(s)
    except ValueError:
        raise HTTPException(400, f"Invalid date format '{s}' — use YYYY-MM-DD")


# ── 1. Budgeted vs Actual Hours Report ───────────────────────────────────────
@router.get("/budgeted-vs-actual")
def budgeted_vs_actual_report(
    project_id:  Optional[int] = None,
    assignee:    Optional[str] = None,
    team:        Optional[str] = None,
    status:      Optional[str] = None,
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    start = _parse_date(start_date)
    end   = _parse_date(end_date)

    q = db.query(CustomMilestone)
    if project_id:
        q = q.filter(CustomMilestone.project_id == project_id)
    if assignee:
        q = q.filter(CustomMilestone.assignee.ilike(f"%{assignee}%"))
    if status:
        q = q.filter(CustomMilestone.status == status)
    if start:
        q = q.filter(CustomMilestone.actual_start >= start)
    if end:
        q = q.filter(CustomMilestone.actual_end <= end)

    milestones = q.order_by(CustomMilestone.project_id, CustomMilestone.num).all()

    # Team filter — resolve assignee → user → team name, then compare
    if team:
        milestones = [
            ms for ms in milestones
            if _assignee_team(db, ms.assignee) == team
        ]

    headers = ["Individual Name", "Project", "Team", "Start Date", "End Date",
               "Budgeted Hours", "Actual Hours", "Status"]
    rows = []
    for ms in milestones:
        project = db.query(Project).filter_by(id=ms.project_id).first()
        team_name = _assignee_team(db, ms.assignee)
        budgeted = _milestone_budgeted_hours(db, ms.id)
        actual   = _milestone_actual_hours(db, ms.id)
        rows.append([
            ms.assignee or "—",
            project.name if project else "—",
            team_name or "—",
            _fmt_dt(ms.actual_start),
            _fmt_dt(ms.actual_end),
            budgeted,
            actual,
            ms.status or "—",
        ])

    subtitle_parts = []
    if start_date or end_date:
        subtitle_parts.append(f"Date range: {start_date or '—'} → {end_date or '—'}")
    if project_id:
        p = db.query(Project).filter_by(id=project_id).first()
        if p: subtitle_parts.append(f"Project: {p.name}")
    subtitle = " | ".join(subtitle_parts)

    return _build_xlsx("Budgeted vs Actual Hours Report", headers, rows,
                        "budgeted-vs-actual-report.xlsx", subtitle)


def _assignee_team(db: Session, assignee: Optional[str]) -> str:
    """Look up the team name for a free-text assignee. Returns '' if not found."""
    if not assignee:
        return ""
    user = db.query(User).filter(User.name.ilike(assignee.strip())).first()
    if not user:
        return ""
    # User.role stores the role label directly ("Functional Consultant",
    # "Technical Team", etc.) — use that as the Team value, not team.name.
    return user.role or ""


# ── 2. Timeline Report ────────────────────────────────────────────────────────
@router.get("/timeline")
def timeline_report(
    project_id:  Optional[int] = None,
    status:      Optional[str] = None,
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    start = _parse_date(start_date)
    end   = _parse_date(end_date)

    q = db.query(CustomMilestone)
    if project_id:
        q = q.filter(CustomMilestone.project_id == project_id)
    if status:
        q = q.filter(CustomMilestone.status == status)
    # Date filter on planned_end for the timeline report
    if start:
        q = q.filter(CustomMilestone.planned_end >= start)
    if end:
        q = q.filter(CustomMilestone.planned_end <= end)

    milestones = q.order_by(CustomMilestone.project_id, CustomMilestone.num).all()

    headers = ["Milestone", "Project", "Planned End Date", "Actual End Date",
               "Schedule Variance Reason", "Status"]
    rows = []
    for ms in milestones:
        project = db.query(Project).filter_by(id=ms.project_id).first()
        rows.append([
            ms.name,
            project.name if project else "—",
            _fmt_dt(ms.planned_end),
            _fmt_dt(ms.actual_end),
            ms.schedule_variance_reason or "",
            ms.status or "—",
        ])

    subtitle_parts = []
    if start_date or end_date:
        subtitle_parts.append(f"Planned end range: {start_date or '—'} → {end_date or '—'}")
    if project_id:
        p = db.query(Project).filter_by(id=project_id).first()
        if p: subtitle_parts.append(f"Project: {p.name}")
    subtitle = " | ".join(subtitle_parts)

    return _build_xlsx("Timeline Report", headers, rows,
                        "timeline-report.xlsx", subtitle)


# ── Filter options (feed the frontend dropdowns) ─────────────────────────────
@router.get("/filter-options")
def filter_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return all distinct values used to populate the frontend filter dropdowns."""
    projects = db.query(Project).order_by(Project.name).all()

    # All active users — used to populate the Individual Name dropdown regardless
    # of whether they have been set as a milestone assignee yet.
    all_users = db.query(User).filter(User.is_active == True).order_by(User.name).all()

    statuses = (db.query(CustomMilestone.status)
                  .filter(CustomMilestone.status.isnot(None))
                  .distinct().order_by(CustomMilestone.status).all())

    # Build assignee list from all active users; include their role as "team"
    # so the frontend can auto-populate the Team field when one is selected.
    assignee_list = [
        {"name": u.name, "team": u.role or ""}
        for u in all_users
        if u.name
    ]

    return {
        "projects":  [{"id": p.id, "name": p.name} for p in projects],
        "assignees": assignee_list,
        "statuses":  [s[0] for s in statuses],
    }
