from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func as sa_func
from typing import Optional, List, Dict
from datetime import datetime
from pydantic import BaseModel
from app.db.database import get_db
from app.models.models import (CustomMilestone, CustomTask, CustomSubtask, Activity,
                                SubtaskQuestion, SubtaskReport, MilestoneReport,
                                TaskFormField,
                                Milestone, Task, Subtask, User, WorkHours, Project,
                                TaskAssignment)
from sqlalchemy.exc import IntegrityError
from app.core.deps import get_current_user
from app.services.audit_service import log_action
from app.services.notification_service import create_notification
from app.services.email_service import send_mailbox_email
import io as _io
import base64 as _b64
import openpyxl as _xl
from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align

router = APIRouter(prefix="/projects/{project_id}/custom-milestones", tags=["Custom Milestones"])


def _parse_dt(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(v)
    except Exception:
        return None


def _parse_date(v):
    if not v:
        return None
    if hasattr(v, "year") and not isinstance(v, str):
        return v
    try:
        return datetime.fromisoformat(v).date()
    except Exception:
        return None


def _total_days(start, end):
    """Total Number of Days between planned/actual start and end (inclusive)."""
    if not start or not end:
        return None
    return (end.date() - start.date()).days + 1


def _wh_sum(db: Session, **filters) -> float:
    """Sum of *actual working* hours (hours_spent minus any logged buffer
    time) for the WorkHours rows matching the given exact-FK filters.

    Buffer hours (breaks/interruptions logged alongside an entry) must be
    excluded here so that "Actual Hours" on Milestone Config matches the
    same "Actual Hours" figure used by the Working Hours module
    (`work_hours.py::_actual_hours`) — previously this summed raw
    `hours_spent` with no buffer subtraction, so the two screens disagreed
    whenever any buffer time had been logged.
    """
    q = db.query(WorkHours.hours_spent, WorkHours.buffer_hours)
    for k, v in filters.items():
        q = q.filter(getattr(WorkHours, k) == v)
    total = 0.0
    for hours_spent, buffer_hours in q.all():
        total += max((hours_spent or 0.0) - (buffer_hours or 0.0), 0.0)
    return round(total, 2)


def _resolve_user_by_name(db: Session, name: Optional[str]) -> Optional[User]:
    if not name:
        return None
    return db.query(User).filter(sa_func.lower(User.name) == name.strip().lower()).first()


def _wh_sum_for_assignee(db: Session, assignee: Optional[str], **filters) -> float:
    """Actual Hours for one Milestone/Task/Subtask/Activity, attributed to
    its own assigned person (testing feedback: "calculate the Actual Total
    Hours at the Milestone, Task, and Subtask levels based on the respective
    person assigned"). When the entity has an assignee that resolves to a
    known user, only that user's logged WorkHours rows count — this stops a
    different contributor's hours (logged against the same entity id) from
    inflating the assigned person's Actual Hours figure. Falls back to
    summing every logger's hours when the entity has no assignee, or the
    assignee name doesn't match a known user (no one specific to attribute
    to, so don't silently drop real logged time)."""
    user = _resolve_user_by_name(db, assignee)
    if user is not None:
        filters = {**filters, "user_id": user.id}
    return _wh_sum(db, **filters)


def _activity_hours(db: Session, a: Activity):
    estimated = a.estimated_hours or 0.0
    actual = _wh_sum_for_assignee(db, a.assignee, activity_id=a.id)
    return estimated, actual


def _subtask_hours(db: Session, s: CustomSubtask):
    estimated = s.estimated_hours or 0.0
    actual = _wh_sum_for_assignee(db, s.assignee, custom_subtask_id=s.id)
    for a in s.activities:
        a_est, a_act = _activity_hours(db, a)
        estimated += a_est
        actual += a_act
    return estimated, actual


def _linked_assignee_id(db: Session, **assignment_filters) -> Optional[int]:
    """The assigned_to user id of the most recent TaskAssignment that links
    directly to this Milestone/Task (via custom_task_id, or milestone_num
    for a milestone-level/General assignment), if one exists."""
    a = db.query(TaskAssignment).filter_by(**assignment_filters) \
          .order_by(TaskAssignment.created_at.desc()).first()
    return a.assigned_to if a else None


def _task_hours(db: Session, t: CustomTask):
    estimated = actual = 0.0
    # A Task's own free-text `assignee` field (set in Milestone Configuration)
    # and the Assignments module's `assigned_to` (set independently, via
    # TaskAssignment.custom_task_id) can drift out of sync — e.g. a task
    # assigned to someone through Assignments while its `assignee` field
    # still says someone else or is blank. When an Assignment links directly
    # to this task, that's the authoritative "who's doing this" — otherwise
    # hours logged by the actually-assigned person were silently excluded
    # from this Task's (and its Milestone's/the project's) Actual Hours,
    # even though the same hours show up correctly in Working Hours' own
    # by-employee breakdown. Prefer it; fall back to the assignee name.
    linked_user_id = _linked_assignee_id(db, custom_task_id=t.id)
    if linked_user_id is not None:
        actual += _wh_sum(db, custom_task_id=t.id, user_id=linked_user_id)
    else:
        actual += _wh_sum_for_assignee(db, t.assignee, custom_task_id=t.id)
    for s in t.subtasks:
        s_est, s_act = _subtask_hours(db, s)
        estimated += s_est
        actual += s_act
    return estimated, actual


def _milestone_hours(db: Session, ms: CustomMilestone):
    estimated = actual = 0.0
    # Same fix as _task_hours, for a milestone-level Assignment (milestone_num
    # set, custom_task_id null — a General/whole-milestone assignment).
    linked_user_id = _linked_assignee_id(db, milestone_num=ms.num, custom_task_id=None)
    if linked_user_id is not None:
        actual += _wh_sum(db, custom_milestone_id=ms.id, user_id=linked_user_id)
    else:
        actual += _wh_sum_for_assignee(db, ms.assignee, custom_milestone_id=ms.id)
    for t in ms.tasks:
        t_est, t_act = _task_hours(db, t)
        estimated += t_est
        actual += t_act
    return estimated, actual


# ── Pre-fetch context for list endpoint (eliminates N+1 queries) ─────────────
def _make_list_ctx(db: Session, project_id: int) -> dict:
    """Pre-fetch all WorkHours and TaskAssignments for a project in 3 queries
    total, then build in-memory lookup dicts that replace the per-entity DB
    calls (_wh_sum / _linked_assignee_id) inside _build and friends.

    Without this, listing N milestones with T tasks each and S subtasks each
    triggered O(N*T*S) separate DB round-trips — 150-670+ queries for a
    typical project. With this, the list endpoint costs exactly 3 queries
    regardless of how many milestones/tasks/subtasks exist.
    """
    from collections import defaultdict

    # Query 1: all work-hour rows for this project
    all_wh = db.query(WorkHours).filter(WorkHours.project_id == project_id).all()

    # wh[(entity_type_char, entity_id)][user_id_or_None] = cumulative hours
    # None key = any-user total (used when no specific user is identified)
    wh: dict = defaultdict(lambda: defaultdict(float))
    for row in all_wh:
        h = max((row.hours_spent or 0.0) - (row.buffer_hours or 0.0), 0.0)
        uid = row.user_id
        for key in [
            ('ms', row.custom_milestone_id),
            ('t',  row.custom_task_id),
            ('s',  row.custom_subtask_id),
            ('a',  row.activity_id),
        ]:
            if key[1] is not None:
                wh[key][uid]  += h
                wh[key][None] += h  # None = sum over all users

    # Query 2: all task assignments for this project (ordered newest-first so
    # first-seen wins when building the "most recent" map)
    all_ta = db.query(TaskAssignment).filter(
        TaskAssignment.project_id == project_id
    ).order_by(TaskAssignment.created_at.desc()).all()

    ta_by_task: dict  = {}   # custom_task_id  -> assigned_to user_id
    ta_by_msnum: dict = {}   # milestone_num   -> assigned_to user_id (task_id=None)
    for ta in all_ta:
        if ta.custom_task_id is not None:
            ta_by_task.setdefault(ta.custom_task_id, ta.assigned_to)
        elif ta.milestone_num is not None:
            ta_by_msnum.setdefault(ta.milestone_num, ta.assigned_to)

    # Query 3: name→id map for assignee resolution
    all_users = db.query(User.id, User.name).all()
    user_name_to_id: dict = {
        u.name.strip().lower(): u.id for u in all_users if u.name
    }

    return {
        "wh": wh,
        "ta_by_task": ta_by_task,
        "ta_by_msnum": ta_by_msnum,
        "user_name_to_id": user_name_to_id,
    }


def _ctx_wh_sum(ctx: dict, entity_key: tuple, user_id=None) -> float:
    return round(ctx["wh"][entity_key].get(user_id, 0.0), 2)


def _ctx_wh_for_assignee(ctx: dict, assignee, entity_key: tuple) -> float:
    uid = None
    if assignee:
        uid = ctx["user_name_to_id"].get(assignee.strip().lower())
    return _ctx_wh_sum(ctx, entity_key, user_id=uid)


def _activity_hours_ctx(ctx: dict, a: "Activity"):
    estimated = a.estimated_hours or 0.0
    actual    = _ctx_wh_for_assignee(ctx, a.assignee, ("a", a.id))
    return estimated, actual


def _subtask_hours_ctx(ctx: dict, s: "CustomSubtask"):
    estimated = s.estimated_hours or 0.0
    actual    = _ctx_wh_for_assignee(ctx, s.assignee, ("s", s.id))
    for a in s.activities:
        a_est, a_act = _activity_hours_ctx(ctx, a)
        estimated += a_est
        actual    += a_act
    return estimated, actual


def _task_hours_ctx(ctx: dict, t: "CustomTask"):
    estimated = 0.0
    linked_uid = ctx["ta_by_task"].get(t.id)
    if linked_uid is not None:
        actual = _ctx_wh_sum(ctx, ("t", t.id), user_id=linked_uid)
    else:
        actual = _ctx_wh_for_assignee(ctx, t.assignee, ("t", t.id))
    for s in t.subtasks:
        s_est, s_act = _subtask_hours_ctx(ctx, s)
        estimated += s_est
        actual    += s_act
    return estimated, actual


def _milestone_hours_ctx(ctx: dict, ms: "CustomMilestone"):
    estimated = 0.0
    linked_uid = ctx["ta_by_msnum"].get(ms.num)
    if linked_uid is not None:
        actual = _ctx_wh_sum(ctx, ("ms", ms.id), user_id=linked_uid)
    else:
        actual = _ctx_wh_for_assignee(ctx, ms.assignee, ("ms", ms.id))
    for t in ms.tasks:
        t_est, t_act = _task_hours_ctx(ctx, t)
        estimated += t_est
        actual    += t_act
    return estimated, actual


def _build_activity_ctx(a: "Activity", ctx: dict):
    est, act = _activity_hours_ctx(ctx, a)
    return {
        "id": a.id, "name": a.name, "status": a.status, "assignee": a.assignee,
        "planned_start": a.planned_start, "planned_end": a.planned_end,
        "actual_start": a.actual_start, "actual_end": a.actual_end,
        "start_time": a.start_time, "end_time": a.end_time,
        "estimated_hours": a.estimated_hours or 0.0,
        "own_estimated_hours": a.estimated_hours or 0.0,
        "actual_hours": act,
        "total_days": _total_days(a.planned_start, a.planned_end) or _total_days(a.actual_start, a.actual_end),
    }


def _build_subtask_ctx(s: "CustomSubtask", ctx: dict):
    est, act = _subtask_hours_ctx(ctx, s)
    return {
        "id": s.id, "num": s.num, "name": s.name,
        "input_type": s.input_type, "response": s.response, "status": s.status,
        "assignee": s.assignee,
        "planned_start": s.planned_start, "planned_end": s.planned_end,
        "actual_start": s.actual_start, "actual_end": s.actual_end,
        "start_time": s.start_time, "end_time": s.end_time,
        "estimated_hours": est, "own_estimated_hours": s.estimated_hours or 0.0,
        "actual_hours": act,
        "total_days": _total_days(s.planned_start, s.planned_end) or _total_days(s.actual_start, s.actual_end),
        "activities": [_build_activity_ctx(a, ctx) for a in sorted(s.activities, key=lambda x: x.id)],
        "questions": [_build_question(q) for q in sorted(s.questions, key=lambda x: (x.num or 0, x.id))],
        "reports": [_build_report(r) for r in sorted(s.reports, key=lambda x: x.id)],
    }


def _build_task_ctx(t: "CustomTask", ctx: dict):
    est, act = _task_hours_ctx(ctx, t)
    return {
        "id": t.id, "num": t.num, "name": t.name,
        "responsibility": t.responsibility,
        "status": t.status, "assignee": t.assignee,
        "planned_start": t.planned_start, "planned_end": t.planned_end,
        "actual_start": t.actual_start, "actual_end": t.actual_end,
        "start_time": t.start_time, "end_time": t.end_time,
        "estimated_hours": est, "actual_hours": act,
        "total_days": _total_days(t.planned_start, t.planned_end) or _total_days(t.actual_start, t.actual_end),
        "form_fields": [_build_form_field(f) for f in sorted(t.form_fields, key=lambda x: x.num or 0)],
        "subtasks": [_build_subtask_ctx(s, ctx) for s in sorted(t.subtasks, key=lambda x: x.num or 0)],
    }


def _build_ctx(ms: "CustomMilestone", ctx: dict):
    est, act = _milestone_hours_ctx(ctx, ms)
    return {
        "id": ms.id, "num": ms.num, "name": ms.name,
        "description": ms.description, "responsible": ms.responsible,
        "is_active": ms.is_active,
        "status": ms.status, "assignee": ms.assignee,
        "planned_start": ms.planned_start, "planned_end": ms.planned_end,
        "actual_start": ms.actual_start, "actual_end": ms.actual_end,
        "start_time": ms.start_time, "end_time": ms.end_time,
        "schedule_variance_reason": ms.schedule_variance_reason,
        "iteration": ms.iteration or 1,
        "revision_reason": ms.revision_reason,
        "revision_description": ms.revision_description,
        "estimated_hours": est, "actual_hours": act,
        "total_days": _total_days(ms.planned_start, ms.planned_end) or _total_days(ms.actual_start, ms.actual_end),
        "reports": [_build_milestone_report(r) for r in sorted(ms.reports, key=lambda x: x.id)],
        "tasks": [_build_task_ctx(t, ctx) for t in sorted(ms.tasks, key=lambda x: x.num or 0)],
    }


def _notify_task_assignment(db: Session, project_id: int, task: CustomTask, milestone_name: str):
    """Requirement: once a Task is assigned, auto-generate an email AND a
    Notifications-tab entry for the assignee."""
    user = _resolve_user_by_name(db, task.assignee)
    project = db.query(Project).filter_by(id=project_id).first()
    project_name = project.name if project else "—"
    message = f"Task '{task.name}' (Milestone: {milestone_name}) assigned to {task.assignee}."
    email_to = user.email if user else None
    email_subject = f"[{project_name}] Task assigned to you — {task.name}"
    due = task.planned_end.strftime("%Y-%m-%d") if task.planned_end else None
    email_body = f"""
    <p>Hi {user.name if user else task.assignee},</p>
    <p>A new task has been assigned to you in <strong>{project_name}</strong> (Milestone: {milestone_name}):</p>
    <p><strong>Task:</strong> {task.name}</p>
    {f'<p><strong>Planned End Date:</strong> {due}</p>' if due else ''}
    <p>Please log in to review and start working on this task.</p>
    <p>Regards,<br>Project WBS System</p>
    """
    create_notification(
        db, project_id, "assignment", message,
        user_id=user.id if user else None,
        email_to=email_to,
        send_now=bool(email_to),
        email_subject=email_subject,
        email_body=email_body,
    )


def _notify_assignee(db: Session, project_id: int, entity_type: str, entity_name: str, assignee_name: str, due_date=None):
    """Generic in-app + email notification when any entity's assignee is set/changed.
    Works for Milestone, Subtask, Activity, Report."""
    user = _resolve_user_by_name(db, assignee_name)
    project = db.query(Project).filter_by(id=project_id).first()
    project_name = project.name if project else "—"
    message = f"{entity_type} '{entity_name}' assigned to {assignee_name} in project '{project_name}'."
    email_to = user.email if user else None
    email_subject = f"[{project_name}] {entity_type} assigned to you — {entity_name}"
    due = due_date.strftime("%Y-%m-%d") if due_date else None
    email_body = f"""
    <p>Hi {user.name if user else assignee_name},</p>
    <p>A {entity_type.lower()} has been assigned to you in <strong>{project_name}</strong>:</p>
    <p><strong>{entity_type}:</strong> {entity_name}</p>
    {f'<p><strong>Planned End Date:</strong> {due}</p>' if due else ''}
    <p>Please log in to review and start working on this item.</p>
    <p>Regards,<br>Project WBS System</p>
    """
    create_notification(
        db, project_id, "assignment", message,
        user_id=user.id if user else None,
        email_to=email_to,
        send_now=bool(email_to),
        email_subject=email_subject,
        email_body=email_body,
    )


# ── Completion auto-bubbling: Activity -> Subtask -> Task -> Milestone ───────
def _bubble_subtask(db: Session, s: CustomSubtask):
    if s.activities and all(a.status == "Completed" for a in s.activities):
        if s.status != "Completed":
            s.status = "Completed"
            if not s.actual_end:
                s.actual_end = datetime.utcnow()
    if s.task_id:
        task = db.query(CustomTask).filter_by(id=s.task_id).first()
        if task:
            _bubble_task(db, task)


def _bubble_task(db: Session, t: CustomTask):
    if t.subtasks and all(s.status == "Completed" for s in t.subtasks):
        if t.status != "Completed":
            t.status = "Completed"
            if not t.actual_end:
                t.actual_end = datetime.utcnow()
    if t.milestone_id:
        ms = db.query(CustomMilestone).filter_by(id=t.milestone_id).first()
        if ms:
            _bubble_milestone(db, ms)


def _bubble_milestone(db: Session, ms: CustomMilestone):
    if ms.tasks and all(t.status == "Completed" for t in ms.tasks):
        if ms.status != "Completed":
            ms.status = "Completed"
            if not ms.actual_end:
                ms.actual_end = datetime.utcnow()


REVISION_REASONS = [
    "New Requirements",
    "Major Changes",
    "Bug Fixes",
    "Enhancements",
    "Other",
]

class RevisitPayload(BaseModel):
    reason: str               # one of REVISION_REASONS
    description: Optional[str] = None

class MilestoneCreate(BaseModel):
    num: int
    name: str
    description: Optional[str] = None
    responsible: Optional[str] = None
    status: Optional[str] = None
    assignee: Optional[str] = None
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    schedule_variance_reason: Optional[str] = None

class MilestoneUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    responsible: Optional[str] = None
    is_active: Optional[bool] = None
    status: Optional[str] = None
    assignee: Optional[str] = None
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    schedule_variance_reason: Optional[str] = None

class TaskCreate(BaseModel):
    name: str
    num: Optional[int] = None
    responsibility: Optional[str] = None
    status: Optional[str] = None
    assignee: Optional[str] = None
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None

class TaskUpdate(BaseModel):
    name: Optional[str] = None
    responsibility: Optional[str] = None
    status: Optional[str] = None
    assignee: Optional[str] = None
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None

class SubtaskCreate(BaseModel):
    name: str
    num: Optional[int] = None
    input_type: str = "text"
    status: Optional[str] = None
    assignee: Optional[str] = None
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    estimated_hours: Optional[float] = None

class SubtaskUpdate(BaseModel):
    name: Optional[str] = None
    input_type: Optional[str] = None
    response: Optional[str] = None
    status: Optional[str] = None
    assignee: Optional[str] = None
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    estimated_hours: Optional[float] = None

class ActivityCreate(BaseModel):
    name: str
    status: Optional[str] = None
    assignee: Optional[str] = None
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    estimated_hours: Optional[float] = None

class ActivityUpdate(BaseModel):
    name: Optional[str] = None
    status: Optional[str] = None
    assignee: Optional[str] = None
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    estimated_hours: Optional[float] = None

class SubtaskQuestionCreate(BaseModel):
    question_text: str
    input_type: str = "text"
    num: Optional[int] = None

class SubtaskQuestionUpdate(BaseModel):
    question_text: Optional[str] = None
    input_type: Optional[str] = None
    response: Optional[str] = None

class FormFieldCreate(BaseModel):
    question_text: str
    input_type: str = "text"
    section_name: Optional[str] = None
    num: Optional[int] = None

class FormFieldUpdate(BaseModel):
    question_text: Optional[str] = None
    input_type: Optional[str] = None
    section_name: Optional[str] = None
    response: Optional[str] = None

class ReportCreate(BaseModel):
    report_number: str
    report_name: str
    department: Optional[str] = None
    status: Optional[str] = None
    assigned_to: Optional[str] = None
    due_date: Optional[str] = None

class ReportUpdate(BaseModel):
    report_number: Optional[str] = None
    report_name: Optional[str] = None
    department: Optional[str] = None
    status: Optional[str] = None
    assigned_to: Optional[str] = None
    due_date: Optional[str] = None

class TemplateSelection(BaseModel):
    """Selective copy from the standard template. Omit entirely to copy the
    whole milestone (all tasks + all subtasks) — the original "select all"
    behaviour. If task_nums is given, only those tasks are copied. If a key
    appears in subtask_nums for a task, only those subtasks are copied for
    that task (otherwise all of that task's subtasks are copied). Selecting a
    subtask whose task isn't listed in task_nums auto-includes that task
    (parent auto-selection)."""
    task_nums: Optional[List[int]] = None
    subtask_nums: Optional[Dict[int, List[int]]] = None


def _set_fields(obj, payload, fields):
    data = payload.model_dump(exclude_none=True)
    for k in fields:
        if k in data:
            v = data[k]
            if k in ("planned_start", "planned_end", "actual_start", "actual_end"):
                v = _parse_dt(v)
            setattr(obj, k, v)


def _build_activity(a: Activity, db: Session = None):
    est, act = _activity_hours(db, a) if db is not None else (a.estimated_hours or 0.0, 0.0)
    return {
        "id": a.id, "name": a.name, "status": a.status, "assignee": a.assignee,
        "planned_start": a.planned_start, "planned_end": a.planned_end,
        "actual_start": a.actual_start, "actual_end": a.actual_end,
        "start_time": a.start_time, "end_time": a.end_time,
        # "estimated_hours" is the rollup total shown for display (== own
        # value here since Activity has no children); "own_estimated_hours"
        # is always the raw, editable column — see _build_subtask for why
        # this distinction matters.
        "estimated_hours": a.estimated_hours or 0.0,
        "own_estimated_hours": a.estimated_hours or 0.0,
        "actual_hours": act,
        "total_days": _total_days(a.planned_start, a.planned_end) or _total_days(a.actual_start, a.actual_end),
    }


def _build_subtask(s: CustomSubtask, db: Session = None):
    est, act = _subtask_hours(db, s) if db is not None else (s.estimated_hours or 0.0, 0.0)
    return {
        "id": s.id, "num": s.num, "name": s.name,
        "input_type": s.input_type, "response": s.response, "status": s.status,
        "assignee": s.assignee,
        "planned_start": s.planned_start, "planned_end": s.planned_end,
        "actual_start": s.actual_start, "actual_end": s.actual_end,
        "start_time": s.start_time, "end_time": s.end_time,
        # "estimated_hours" = rollup total (own + all child Activities) for
        # display at this and parent levels. "own_estimated_hours" = the raw
        # column actually stored on this Subtask, which is what the edit
        # form must read/write — using the rollup there was the root cause
        # of Total Estimated Hours inflating on every save (the displayed
        # rollup got written back into the raw column, then re-added to the
        # activities' hours on the next computation, compounding forever).
        "estimated_hours": est, "own_estimated_hours": s.estimated_hours or 0.0,
        "actual_hours": act,
        "total_days": _total_days(s.planned_start, s.planned_end) or _total_days(s.actual_start, s.actual_end),
        "activities": [_build_activity(a, db) for a in sorted(s.activities, key=lambda x: x.id)],
        "questions": [_build_question(q) for q in sorted(s.questions, key=lambda x: (x.num or 0, x.id))],
        "reports": [_build_report(r) for r in sorted(s.reports, key=lambda x: x.id)],
    }


def _build_question(q: SubtaskQuestion):
    return {
        "id": q.id, "num": q.num, "question_text": q.question_text,
        "input_type": q.input_type, "response": q.response,
    }


def _build_report(r: SubtaskReport):
    return {
        "id": r.id, "report_number": r.report_number, "report_name": r.report_name,
        "department": r.department, "status": r.status, "assigned_to": r.assigned_to,
        "due_date": r.due_date,
    }


def _build_milestone_report(r: MilestoneReport):
    return {
        "id": r.id, "report_number": r.report_number, "report_name": r.report_name,
        "department": r.department, "status": r.status, "assigned_to": r.assigned_to,
        "due_date": r.due_date,
    }


def _build_form_field(f: TaskFormField):
    return {
        "id": f.id, "num": f.num, "section_name": f.section_name,
        "question_text": f.question_text, "input_type": f.input_type,
        "response": f.response,
    }


def _build_task(t: CustomTask, db: Session = None):
    est, act = _task_hours(db, t) if db is not None else (0.0, 0.0)
    return {
        "id": t.id, "num": t.num, "name": t.name,
        "responsibility": t.responsibility,
        "status": t.status, "assignee": t.assignee,
        "planned_start": t.planned_start, "planned_end": t.planned_end,
        "actual_start": t.actual_start, "actual_end": t.actual_end,
        "start_time": t.start_time, "end_time": t.end_time,
        "estimated_hours": est, "actual_hours": act,
        "total_days": _total_days(t.planned_start, t.planned_end) or _total_days(t.actual_start, t.actual_end),
        # New: form_fields replaces the subtask+question hierarchy
        "form_fields": [_build_form_field(f) for f in sorted(t.form_fields, key=lambda x: x.num or 0)],
        # Keep subtasks in payload for backward compat during transition
        "subtasks": [_build_subtask(s, db) for s in sorted(t.subtasks, key=lambda x: x.num or 0)],
    }


def _build(ms: CustomMilestone, db: Session = None):
    est, act = _milestone_hours(db, ms) if db is not None else (0.0, 0.0)
    return {
        "id": ms.id, "num": ms.num, "name": ms.name,
        "description": ms.description, "responsible": ms.responsible,
        "is_active": ms.is_active,
        "status": ms.status, "assignee": ms.assignee,
        "planned_start": ms.planned_start, "planned_end": ms.planned_end,
        "actual_start": ms.actual_start, "actual_end": ms.actual_end,
        "start_time": ms.start_time, "end_time": ms.end_time,
        "schedule_variance_reason": ms.schedule_variance_reason,
        # Iteration fields — iteration=1 for the original, 2+ for revisits.
        "iteration": ms.iteration or 1,
        "revision_reason": ms.revision_reason,
        "revision_description": ms.revision_description,
        "estimated_hours": est, "actual_hours": act,
        "total_days": _total_days(ms.planned_start, ms.planned_end) or _total_days(ms.actual_start, ms.actual_end),
        "reports": [_build_milestone_report(r) for r in sorted(ms.reports, key=lambda x: x.id)],
        "tasks": [_build_task(t, db) for t in sorted(ms.tasks, key=lambda x: x.num or 0)],
    }


# ── IMPORTANT: /templates must come before /{milestone_id} ───────────────────
@router.get("/templates")
def get_milestone_templates(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Return all 10 standard milestones for the picker."""
    standard = db.query(Milestone).options(
        joinedload(Milestone.tasks).joinedload(Task.subtasks)
    ).order_by(Milestone.num).all()

    # Check which nums are already in CustomMilestone for this project
    existing_nums = {
        cm.num for cm in db.query(CustomMilestone).filter_by(project_id=project_id).all()
    }

    return [
        {
            "num": ms.num,
            "name": ms.name,
            "already_added": ms.num in existing_nums,
            "task_count": len(ms.tasks),
            "subtask_count": sum(len(t.subtasks) for t in ms.tasks),
        }
        for ms in standard
    ]


# ── Full standard tree for one milestone (drill-down picker) ─────────────────
@router.get("/templates/{ms_num}/detail")
def get_milestone_template_detail(
    project_id: int, ms_num: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    standard = db.query(Milestone).options(
        joinedload(Milestone.tasks).joinedload(Task.subtasks)
    ).filter_by(num=ms_num).first()
    if not standard:
        raise HTTPException(404, f"Standard milestone {ms_num} not found")

    existing_ct_nums, existing_cs_nums = {}, {}
    cm = db.query(CustomMilestone).filter_by(project_id=project_id, num=ms_num).first()
    if cm:
        for t in cm.tasks:
            existing_ct_nums[t.num] = True
            for s in t.subtasks:
                existing_cs_nums[(t.num, s.num)] = True

    return {
        "num": standard.num, "name": standard.name,
        "tasks": [
            {
                "num": t.num, "name": t.name, "responsibility": t.responsibility,
                "already_added": existing_ct_nums.get(t.num, False),
                "subtasks": [
                    {"num": s.num, "name": s.name, "input_type": s.input_type,
                     "already_added": existing_cs_nums.get((t.num, s.num), False)}
                    for s in sorted(t.subtasks, key=lambda x: x.num or 0)
                ]
            }
            for t in sorted(standard.tasks, key=lambda x: x.num or 0)
        ]
    }


# ── List selected milestones for this project ─────────────────────────────────
@router.get("")
def list_custom_milestones(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    milestones = db.query(CustomMilestone).options(
        joinedload(CustomMilestone.tasks)
            .joinedload(CustomTask.form_fields),
        joinedload(CustomMilestone.tasks)
            .joinedload(CustomTask.subtasks)
            .joinedload(CustomSubtask.activities),
        joinedload(CustomMilestone.tasks)
            .joinedload(CustomTask.subtasks)
            .joinedload(CustomSubtask.questions),
        joinedload(CustomMilestone.tasks)
            .joinedload(CustomTask.subtasks)
            .joinedload(CustomSubtask.reports),
        joinedload(CustomMilestone.reports),
    ).filter_by(project_id=project_id).order_by(
        CustomMilestone.num, CustomMilestone.iteration
    ).all()
    # Pre-fetch WorkHours + TaskAssignments in 3 queries, then build entirely
    # from in-memory lookups — eliminates the O(N×T×S) per-entity DB calls.
    ctx = _make_list_ctx(db, project_id)
    return [_build_ctx(ms, ctx) for ms in milestones]


# ── Add from standard template (selective Task/Subtask picking) ──────────────
@router.post("/from-template/{ms_num}")
def add_from_template(
    project_id: int, ms_num: int,
    payload: Optional[TemplateSelection] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    existing = db.query(CustomMilestone).filter_by(project_id=project_id, num=ms_num).first()
    if existing:
        raise HTTPException(400, f"Milestone {ms_num} already added")

    standard = db.query(Milestone).options(
        joinedload(Milestone.tasks).joinedload(Task.subtasks)
    ).filter_by(num=ms_num).first()
    if not standard:
        raise HTTPException(404, f"Standard milestone {ms_num} not found")

    # Resolve the closure of which tasks/subtasks to copy.
    # No selection at all => copy everything (back-compat "select all").
    task_nums = payload.task_nums if payload else None
    subtask_nums = payload.subtask_nums if payload else None

    selected_task_nums = set(task_nums) if task_nums else None
    if subtask_nums:
        # Auto-select the parent task of any directly-picked subtask.
        if selected_task_nums is None:
            selected_task_nums = set()
        for tnum in subtask_nums.keys():
            selected_task_nums.add(tnum)

    cm = CustomMilestone(
        project_id=project_id, num=standard.num, name=standard.name,
        description="From standard template", is_active=True
    )
    db.add(cm); db.flush()

    # Milestones 5-10 use the new Task-promotion structure: each Subtask
    # becomes an independent CustomTask with its Questions as FormFields.
    # Milestones 1-4 keep the original Tasks; Subtasks fold as Form sections.
    from main import _SUBTASK_TO_NEW_TASK

    for i, t in enumerate(sorted(standard.tasks, key=lambda x: x.num or 0)):
        t_num = t.num or (i + 1)
        if selected_task_nums is not None and t_num not in selected_task_nums:
            continue

        need_new_tasks = ms_num in (5, 6, 7, 8, 9, 10)

        if need_new_tasks:
            # Build promoted tasks from subtasks
            new_task_map: dict[str, CustomTask] = {}
            for j, s in enumerate(sorted(t.subtasks, key=lambda x: x.num or 0)):
                s_num = s.num or (j + 1)
                wanted_subs = subtask_nums.get(t_num) if subtask_nums else None
                if wanted_subs is not None and s_num not in wanted_subs:
                    continue
                new_name = _SUBTASK_TO_NEW_TASK.get((ms_num, t_num, s_num))
                if new_name is None:
                    new_name = s.name  # fallback: use subtask name as task name
                if new_name not in new_task_map:
                    nt = CustomTask(milestone_id=cm.id, project_id=project_id,
                                    num=None, name=new_name,
                                    responsibility=t.responsibility, status="Not Started")
                    db.add(nt); db.flush()
                    new_task_map[new_name] = nt
                tgt = new_task_map[new_name]
                questions = sorted(s.questions, key=lambda x: x.num or 0)
                if not questions:
                    db.add(TaskFormField(task_id=tgt.id, milestone_id=cm.id,
                                         project_id=project_id, num=1,
                                         section_name=None,
                                         question_text=s.name,
                                         input_type=s.input_type or "text"))
                else:
                    for idx, q in enumerate(questions, 1):
                        db.add(TaskFormField(task_id=tgt.id, milestone_id=cm.id,
                                              project_id=project_id, num=idx,
                                              section_name=None,
                                              question_text=q.question_text,
                                              input_type=q.input_type or "text"))
        else:
            # M1-M4: create the original Task, fold Subtasks as Form sections
            ct = CustomTask(milestone_id=cm.id, project_id=project_id,
                            num=t_num, name=t.name, responsibility=t.responsibility,
                            status="Not Started")
            db.add(ct); db.flush()

            wanted_subs = subtask_nums.get(t_num) if subtask_nums else None
            field_num = 0
            for j, s in enumerate(sorted(t.subtasks, key=lambda x: x.num or 0)):
                s_num = s.num or (j + 1)
                if wanted_subs is not None and s_num not in wanted_subs:
                    continue
                questions = sorted(s.questions, key=lambda x: x.num or 0)
                if not questions:
                    field_num += 1
                    db.add(TaskFormField(task_id=ct.id, milestone_id=cm.id,
                                         project_id=project_id, num=field_num,
                                         section_name=s.name,
                                         question_text=s.name,
                                         input_type=s.input_type or "text"))
                else:
                    for q in questions:
                        field_num += 1
                        db.add(TaskFormField(task_id=ct.id, milestone_id=cm.id,
                                              project_id=project_id, num=field_num,
                                              section_name=s.name,
                                              question_text=q.question_text,
                                              input_type=q.input_type or "text"))

    log_action(db, actor=current_user.name, action="add_milestone",
               description=f"Added M{ms_num:02d} to project",
               project_id=project_id, user_id=current_user.id)
    db.commit(); db.refresh(cm)
    return _build(cm, db)


# ── Add one standard task (+ optional subtask selection) into an existing
#    custom milestone — covers incrementally adding more Tasks later. ───────
@router.post("/{milestone_id}/tasks/from-template")
def add_task_from_template(
    project_id: int, milestone_id: int,
    task_num: int,
    subtask_nums: Optional[List[int]] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    cm = db.query(CustomMilestone).filter_by(id=milestone_id, project_id=project_id).first()
    if not cm:
        raise HTTPException(404, "Milestone not found")
    if any(t.num == task_num for t in cm.tasks):
        raise HTTPException(400, f"Task {task_num} already added")

    standard = db.query(Milestone).options(
        joinedload(Milestone.tasks).joinedload(Task.subtasks)
    ).filter_by(num=cm.num).first()
    std_task = next((t for t in standard.tasks if t.num == task_num), None) if standard else None
    if not std_task:
        raise HTTPException(404, f"Standard task {task_num} not found under milestone {cm.num}")

    from main import _SUBTASK_TO_NEW_TASK
    ms_num = cm.num
    t_num  = std_task.num or task_num
    need_new_tasks = ms_num in (5, 6, 7, 8, 9, 10)

    if need_new_tasks:
        # M5-M10: Subtasks become new Tasks with FormFields
        new_task_map: dict[str, CustomTask] = {}
        for j, s in enumerate(sorted(std_task.subtasks, key=lambda x: x.num or 0)):
            s_num = s.num or (j + 1)
            if subtask_nums is not None and s_num not in subtask_nums:
                continue
            new_name = _SUBTASK_TO_NEW_TASK.get((ms_num, t_num, s_num)) or s.name
            if new_name not in new_task_map:
                nt = CustomTask(milestone_id=cm.id, project_id=project_id, num=None,
                                name=new_name, responsibility=std_task.responsibility,
                                status="Not Started")
                db.add(nt); db.flush()
                new_task_map[new_name] = nt
            tgt = new_task_map[new_name]
            questions = sorted(s.questions, key=lambda x: x.num or 0)
            if not questions:
                db.add(TaskFormField(task_id=tgt.id, milestone_id=cm.id,
                                     project_id=project_id, num=1, section_name=None,
                                     question_text=s.name, input_type=s.input_type or "text"))
            else:
                for idx, q in enumerate(questions, 1):
                    db.add(TaskFormField(task_id=tgt.id, milestone_id=cm.id,
                                         project_id=project_id, num=idx, section_name=None,
                                         question_text=q.question_text,
                                         input_type=q.input_type or "text"))
        db.commit()
        return {"status": "ok", "tasks_created": list(new_task_map.keys())}

    else:
        # M1-M4: create Task, fold Subtasks as Form sections
        ct = CustomTask(milestone_id=cm.id, project_id=project_id, num=std_task.num,
                        name=std_task.name, responsibility=std_task.responsibility,
                        status="Not Started")
        db.add(ct); db.flush()

        field_num = 0
        for j, s in enumerate(sorted(std_task.subtasks, key=lambda x: x.num or 0)):
            s_num = s.num or (j + 1)
            if subtask_nums is not None and s_num not in subtask_nums:
                continue
            questions = sorted(s.questions, key=lambda x: x.num or 0)
            if not questions:
                field_num += 1
                db.add(TaskFormField(task_id=ct.id, milestone_id=cm.id,
                                     project_id=project_id, num=field_num,
                                     section_name=s.name, question_text=s.name,
                                     input_type=s.input_type or "text"))
            else:
                for q in questions:
                    field_num += 1
                    db.add(TaskFormField(task_id=ct.id, milestone_id=cm.id,
                                         project_id=project_id, num=field_num,
                                         section_name=s.name, question_text=q.question_text,
                                         input_type=q.input_type or "text"))
        db.commit(); db.refresh(ct)
        return _build_task(ct, db)


# ── Add one standard subtask into an existing custom task ────────────────────
@router.post("/{milestone_id}/tasks/{task_id}/subtasks/from-template")
def add_subtask_from_template(
    project_id: int, milestone_id: int, task_id: int,
    subtask_num: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    ct = db.query(CustomTask).filter_by(id=task_id, milestone_id=milestone_id).first()
    if not ct:
        raise HTTPException(404, "Task not found")
    if any(s.num == subtask_num for s in ct.subtasks):
        raise HTTPException(400, f"Subtask {subtask_num} already added")

    cm = db.query(CustomMilestone).filter_by(id=milestone_id, project_id=project_id).first()
    if not cm:
        raise HTTPException(404, "Milestone not found")
    standard_ms = db.query(Milestone).filter_by(num=cm.num).first()
    std_task = db.query(Task).options(joinedload(Task.subtasks)).filter_by(
        milestone_id=standard_ms.id if standard_ms else -1, num=ct.num
    ).first()
    if not std_task:
        raise HTTPException(404, "Standard task not found")
    std_sub = next((s for s in std_task.subtasks if s.num == subtask_num), None)
    if not std_sub:
        raise HTTPException(404, f"Standard subtask {subtask_num} not found")
    # Need std_sub.questions loaded — std_task was fetched without
    # joinedload(Subtask.questions), so this triggers a normal lazy load.
    std_questions = sorted(std_sub.questions, key=lambda x: x.num or 0)

    cs = CustomSubtask(task_id=ct.id, project_id=project_id, num=std_sub.num, name=std_sub.name,
                       input_type=std_sub.input_type or "text", status="Not Started")
    db.add(cs); db.flush()
    # See add_from_template — copy multi-question form data too.
    for q in std_questions:
        db.add(SubtaskQuestion(subtask_id=cs.id, project_id=project_id,
                                num=q.num, question_text=q.question_text,
                                input_type=q.input_type or "text"))
    db.commit(); db.refresh(cs)
    return {"id": cs.id, "num": cs.num, "name": cs.name, "input_type": cs.input_type, "status": cs.status,
            "questions": [_build_question(q) for q in sorted(cs.questions, key=lambda x: (x.num or 0, x.id))]}


# ── Reset all custom milestones ───────────────────────────────────────────────
@router.delete("/reset")
def reset_milestones(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    milestones = db.query(CustomMilestone).filter_by(project_id=project_id).all()
    count = len(milestones)
    for ms in milestones:
        db.delete(ms)
    db.commit()
    return {"status": "reset", "removed": count}


# ── Create custom milestone ───────────────────────────────────────────────────
@router.post("")
def create_custom_milestone(
    project_id: int, payload: MilestoneCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    existing = db.query(CustomMilestone).filter_by(project_id=project_id, num=payload.num).first()
    if existing:
        raise HTTPException(400, f"Milestone {payload.num} already exists")
    cm = CustomMilestone(project_id=project_id, num=payload.num, name=payload.name,
                         description=payload.description, responsible=payload.responsible, is_active=True,
                         status=payload.status or "Not Started", assignee=payload.assignee,
                         planned_start=_parse_dt(payload.planned_start), planned_end=_parse_dt(payload.planned_end),
                         actual_start=_parse_dt(payload.actual_start), actual_end=_parse_dt(payload.actual_end),
                         start_time=payload.start_time, end_time=payload.end_time)
    db.add(cm)
    log_action(db, actor=current_user.name, action="create_milestone",
               description=f"Created custom M{payload.num}: {payload.name}",
               project_id=project_id, user_id=current_user.id)
    db.commit(); db.refresh(cm)
    return _build(cm, db)


# ── Update milestone ──────────────────────────────────────────────────────────
@router.patch("/{milestone_id}")
def update_milestone(
    project_id: int, milestone_id: int, payload: MilestoneUpdate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    ms = db.query(CustomMilestone).filter_by(id=milestone_id, project_id=project_id).first()
    if not ms: raise HTTPException(404, "Milestone not found")
    old_status = ms.status
    old_assignee = ms.assignee
    _set_fields(ms, payload, ["name", "description", "responsible", "is_active", "status",
                              "assignee", "planned_start", "planned_end", "actual_start",
                              "actual_end", "start_time", "end_time",
                              "schedule_variance_reason"])
    if ms.status == "Completed" and old_status != "Completed" and not ms.actual_end:
        ms.actual_end = datetime.utcnow()
    db.flush()
    if ms.assignee and ms.assignee != old_assignee:
        _notify_assignee(db, project_id, "Milestone", ms.name, ms.assignee, ms.planned_end)
    log_action(db, actor=current_user.name, action="update_milestone",
               description=f"Milestone {ms.num} '{ms.name}' updated",
               project_id=project_id, entity_type="custom_milestone", entity_id=ms.id,
               old_value=old_status, new_value=ms.status, user_id=current_user.id)
    db.commit(); db.refresh(ms)
    return _build(ms, db)


# ── Delete milestone ──────────────────────────────────────────────────────────
@router.delete("/{milestone_id}")
def delete_milestone(
    project_id: int, milestone_id: int,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    ms = db.query(CustomMilestone).filter_by(id=milestone_id, project_id=project_id).first()
    if not ms: raise HTTPException(404, "Milestone not found")
    db.delete(ms); db.commit()
    return {"status": "deleted"}


# ── Revisit milestone ─────────────────────────────────────────────────────────
# Creates a fresh iteration of an existing milestone (same num) while keeping
# the original row intact. Clones task + subtask structure (cleared dates/status
# so the new iteration starts clean). Valid reasons: REVISION_REASONS list.
@router.post("/{milestone_id}/revisit")
def revisit_milestone(
    project_id: int, milestone_id: int, payload: RevisitPayload,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    if payload.reason not in REVISION_REASONS:
        raise HTTPException(400, f"Invalid reason. Must be one of: {', '.join(REVISION_REASONS)}")

    source = db.query(CustomMilestone).filter_by(id=milestone_id, project_id=project_id).first()
    if not source:
        raise HTTPException(404, "Milestone not found")

    # Find highest existing iteration for this (project_id, num) pair.
    from sqlalchemy import func as _func
    max_iter = db.query(_func.max(CustomMilestone.iteration)).filter_by(
        project_id=project_id, num=source.num
    ).scalar() or 1
    new_iteration = max_iter + 1

    # Create the new milestone row — same identity fields, fresh tracking fields.
    new_ms = CustomMilestone(
        project_id=project_id,
        num=source.num,
        name=source.name,
        description=source.description,
        responsible=source.responsible,
        is_active=True,
        status="Not Started",
        assignee=source.assignee,
        # Carry forward planned dates as a guide; clear actuals.
        planned_start=source.planned_start,
        planned_end=source.planned_end,
        actual_start=None,
        actual_end=None,
        start_time=None,
        end_time=None,
        schedule_variance_reason=None,
        iteration=new_iteration,
        revision_reason=payload.reason,
        revision_description=payload.description,
    )
    db.add(new_ms); db.flush()

    # Clone tasks + subtasks (structure only — cleared status/actuals/response).
    for t in sorted(source.tasks, key=lambda x: x.num or 0):
        new_t = CustomTask(
            milestone_id=new_ms.id,
            project_id=project_id,
            num=t.num,
            name=t.name,
            responsibility=t.responsibility,
            status="Not Started",
            assignee=t.assignee,
            planned_start=t.planned_start,
            planned_end=t.planned_end,
            actual_start=None,
            actual_end=None,
            start_time=None,
            end_time=None,
        )
        db.add(new_t); db.flush()

        for s in sorted(t.subtasks, key=lambda x: x.num or 0):
            new_s = CustomSubtask(
                task_id=new_t.id,
                project_id=project_id,
                num=s.num,
                name=s.name,
                input_type=s.input_type or "text",
                status="Not Started",
                assignee=s.assignee,
                planned_start=s.planned_start,
                planned_end=s.planned_end,
                actual_start=None,
                actual_end=None,
                start_time=None,
                end_time=None,
                estimated_hours=s.estimated_hours or 0.0,
                response=None,
            )
            db.add(new_s); db.flush()

            # Clone subtask questions (the form structure, not responses).
            for q in sorted(s.questions, key=lambda x: x.num or 0):
                db.add(SubtaskQuestion(
                    subtask_id=new_s.id, project_id=project_id,
                    num=q.num, question_text=q.question_text,
                    input_type=q.input_type or "text",
                ))

    log_action(db, actor=current_user.name, action="revisit_milestone",
               description=f"Revisited M{source.num} '{source.name}' (iteration {new_iteration}): {payload.reason}",
               project_id=project_id, entity_type="custom_milestone", entity_id=new_ms.id,
               user_id=current_user.id)
    db.commit(); db.refresh(new_ms)
    return _build(new_ms, db)


# ── Revision reasons list (for frontend dropdown) ────────────────────────────
@router.get("/revision-reasons")
def get_revision_reasons(
    project_id: int,
    current_user: User = Depends(get_current_user)
):
    return REVISION_REASONS


# ── Add task (manual) ─────────────────────────────────────────────────────────
@router.post("/{milestone_id}/tasks")
def add_task(
    project_id: int, milestone_id: int, payload: TaskCreate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    ms = db.query(CustomMilestone).filter_by(id=milestone_id, project_id=project_id).first()
    if not ms: raise HTTPException(404, "Milestone not found")
    max_num = max((t.num or 0 for t in ms.tasks), default=0)
    ct = CustomTask(milestone_id=milestone_id, project_id=project_id,
                    num=payload.num or (max_num+1), name=payload.name,
                    responsibility=payload.responsibility, status=payload.status or "Not Started",
                    assignee=payload.assignee,
                    planned_start=_parse_dt(payload.planned_start), planned_end=_parse_dt(payload.planned_end),
                    actual_start=_parse_dt(payload.actual_start), actual_end=_parse_dt(payload.actual_end),
                    start_time=payload.start_time, end_time=payload.end_time)
    db.add(ct); db.commit(); db.refresh(ct)
    if ct.assignee:
        _notify_task_assignment(db, project_id, ct, ms.name)
        db.commit()
    return _build_task(ct, db)


# ── Update task ────────────────────────────────────────────────────────────────
@router.patch("/{milestone_id}/tasks/{task_id}")
def update_task(
    project_id: int, milestone_id: int, task_id: int, payload: TaskUpdate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    t = db.query(CustomTask).filter_by(id=task_id, milestone_id=milestone_id).first()
    if not t: raise HTTPException(404, "Task not found")
    old_status = t.status
    old_assignee = t.assignee
    _set_fields(t, payload, ["name", "responsibility", "status", "assignee",
                              "planned_start", "planned_end", "actual_start", "actual_end",
                              "start_time", "end_time"])
    if t.status == "Completed" and old_status != "Completed" and not t.actual_end:
        t.actual_end = datetime.utcnow()
    db.flush()
    # Requirement 1(g): once a Task is assigned (newly assigned or reassigned), notify.
    if t.assignee and t.assignee != old_assignee:
        ms = db.query(CustomMilestone).filter_by(id=milestone_id).first()
        _notify_task_assignment(db, project_id, t, ms.name if ms else "—")
    # Bubble completion up Task -> Milestone
    if t.milestone_id:
        ms = db.query(CustomMilestone).filter_by(id=t.milestone_id).first()
        if ms:
            _bubble_milestone(db, ms)
    log_action(db, actor=current_user.name, action="update_task",
               description=f"Task '{t.name}' updated",
               project_id=project_id, entity_type="custom_task", entity_id=t.id,
               old_value=old_status, new_value=t.status, user_id=current_user.id)
    db.commit(); db.refresh(t)
    return _build_task(t, db)


# ── Delete task ───────────────────────────────────────────────────────────────
@router.delete("/{milestone_id}/tasks/{task_id}")
def delete_task(
    project_id: int, milestone_id: int, task_id: int,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    t = db.query(CustomTask).filter_by(id=task_id, milestone_id=milestone_id).first()
    if not t: raise HTTPException(404, "Task not found")
    db.delete(t); db.commit()
    return {"status": "deleted"}


# ── Task Form Fields (replaces Subtask+Question hierarchy) ───────────────────
@router.post("/{milestone_id}/tasks/{task_id}/form-fields")
def add_form_field(
    project_id: int, milestone_id: int, task_id: int, payload: FormFieldCreate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    task = db.query(CustomTask).filter_by(id=task_id, milestone_id=milestone_id).first()
    if not task: raise HTTPException(404, "Task not found")
    max_num = max((f.num or 0 for f in task.form_fields), default=0)
    ff = TaskFormField(
        task_id=task_id, milestone_id=milestone_id, project_id=project_id,
        num=payload.num or (max_num + 1),
        section_name=payload.section_name,
        question_text=payload.question_text,
        input_type=payload.input_type or "text",
    )
    db.add(ff); db.commit(); db.refresh(ff)
    return _build_form_field(ff)


@router.patch("/{milestone_id}/tasks/{task_id}/form-fields/{field_id}")
def update_form_field(
    project_id: int, milestone_id: int, task_id: int, field_id: int,
    payload: FormFieldUpdate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    ff = db.query(TaskFormField).filter_by(id=field_id, task_id=task_id).first()
    if not ff: raise HTTPException(404, "Form field not found")
    data = payload.model_dump(exclude_none=True)
    for k, v in data.items():
        setattr(ff, k, v)
    db.commit(); db.refresh(ff)
    return _build_form_field(ff)


@router.delete("/{milestone_id}/tasks/{task_id}/form-fields/{field_id}")
def delete_form_field(
    project_id: int, milestone_id: int, task_id: int, field_id: int,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    ff = db.query(TaskFormField).filter_by(id=field_id, task_id=task_id).first()
    if not ff: raise HTTPException(404, "Form field not found")
    db.delete(ff); db.commit()
    return {"status": "deleted"}


# ── Add subtask (manual) ─────────────────────────────────────────────────────
@router.post("/{milestone_id}/tasks/{task_id}/subtasks")
def add_subtask(
    project_id: int, milestone_id: int, task_id: int, payload: SubtaskCreate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    task = db.query(CustomTask).filter_by(id=task_id, milestone_id=milestone_id).first()
    if not task: raise HTTPException(404, "Task not found")
    max_num = max((s.num or 0 for s in task.subtasks), default=0)
    cs = CustomSubtask(task_id=task_id, project_id=project_id,
                       num=payload.num or (max_num+1), name=payload.name,
                       input_type=payload.input_type, status=payload.status or "Not Started",
                       assignee=payload.assignee,
                       planned_start=_parse_dt(payload.planned_start), planned_end=_parse_dt(payload.planned_end),
                       actual_start=_parse_dt(payload.actual_start), actual_end=_parse_dt(payload.actual_end),
                       start_time=payload.start_time, end_time=payload.end_time,
                       estimated_hours=payload.estimated_hours or 0.0)
    db.add(cs); db.commit(); db.refresh(cs)
    return _build_subtask(cs, db)


# ── Update subtask ────────────────────────────────────────────────────────────
@router.patch("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}")
def update_subtask(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int,
    payload: SubtaskUpdate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    s = db.query(CustomSubtask).filter_by(id=subtask_id, task_id=task_id).first()
    if not s: raise HTTPException(404, "Subtask not found")
    old_status = s.status
    old_assignee = s.assignee
    _set_fields(s, payload, ["name", "input_type", "response", "status", "assignee",
                              "planned_start", "planned_end", "actual_start", "actual_end",
                              "start_time", "end_time", "estimated_hours"])
    if s.status == "Completed" and old_status != "Completed" and not s.actual_end:
        s.actual_end = datetime.utcnow()
    db.flush()
    if s.assignee and s.assignee != old_assignee:
        _notify_assignee(db, project_id, "Subtask", s.name, s.assignee, s.planned_end)
    # Bubble completion: Subtask -> Task -> Milestone (req 1b)
    _bubble_subtask(db, s)
    log_action(db, actor=current_user.name, action="update_subtask",
               description=f"Subtask '{s.name}' updated",
               project_id=project_id, entity_type="custom_subtask", entity_id=s.id,
               old_value=old_status, new_value=s.status, user_id=current_user.id)
    db.commit(); db.refresh(s)
    return _build_subtask(s, db)


# ── Delete subtask ────────────────────────────────────────────────────────────
@router.delete("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}")
def delete_subtask(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    s = db.query(CustomSubtask).filter_by(id=subtask_id, task_id=task_id).first()
    if not s: raise HTTPException(404, "Subtask not found")
    db.delete(s); db.commit()
    return {"status": "deleted"}


# ── Activities (optional, nested under a Subtask) ────────────────────────────
@router.post("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}/activities")
def add_activity(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int,
    payload: ActivityCreate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    s = db.query(CustomSubtask).filter_by(id=subtask_id, task_id=task_id).first()
    if not s: raise HTTPException(404, "Subtask not found")
    a = Activity(subtask_id=subtask_id, project_id=project_id, name=payload.name,
                status=payload.status or "Not Started", assignee=payload.assignee,
                planned_start=_parse_dt(payload.planned_start), planned_end=_parse_dt(payload.planned_end),
                actual_start=_parse_dt(payload.actual_start), actual_end=_parse_dt(payload.actual_end),
                start_time=payload.start_time, end_time=payload.end_time,
                estimated_hours=payload.estimated_hours or 0.0)
    db.add(a); db.commit(); db.refresh(a)
    return _build_activity(a, db)


@router.patch("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}/activities/{activity_id}")
def update_activity(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int, activity_id: int,
    payload: ActivityUpdate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    a = db.query(Activity).filter_by(id=activity_id, subtask_id=subtask_id).first()
    if not a: raise HTTPException(404, "Activity not found")
    old_status = a.status
    old_assignee = a.assignee
    _set_fields(a, payload, ["name", "status", "assignee",
                              "planned_start", "planned_end", "actual_start", "actual_end",
                              "start_time", "end_time", "estimated_hours"])
    if a.status == "Completed" and old_status != "Completed" and not a.actual_end:
        a.actual_end = datetime.utcnow()
    db.flush()
    if a.assignee and a.assignee != old_assignee:
        _notify_assignee(db, project_id, "Activity", a.name, a.assignee, a.planned_end)
    # Bubble completion: Activity -> Subtask -> Task -> Milestone (req 1b)
    s = db.query(CustomSubtask).filter_by(id=subtask_id).first()
    if s:
        _bubble_subtask(db, s)
    log_action(db, actor=current_user.name, action="update_activity",
               description=f"Activity '{a.name}' updated",
               project_id=project_id, entity_type="activity", entity_id=a.id,
               old_value=old_status, new_value=a.status, user_id=current_user.id)
    db.commit(); db.refresh(a)
    return _build_activity(a, db)


@router.delete("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}/activities/{activity_id}")
def delete_activity(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int, activity_id: int,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    a = db.query(Activity).filter_by(id=activity_id, subtask_id=subtask_id).first()
    if not a: raise HTTPException(404, "Activity not found")
    db.delete(a); db.commit()
    return {"status": "deleted"}


# ── Subtask Questions (optional, multiple per Subtask) ───────────────────────
# A Subtask can hold several question/answer rows — same idea as the standard
# Milestone system's Question+Response, just inline on one row. The Subtask's
# own input_type/response stays available too (a subtask can answer directly
# with no questions added, or grow a question list when more than one
# question belongs under it).
@router.post("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}/questions")
def add_question(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int,
    payload: SubtaskQuestionCreate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    s = db.query(CustomSubtask).filter_by(id=subtask_id, task_id=task_id).first()
    if not s: raise HTTPException(404, "Subtask not found")
    next_num = payload.num or (max([q.num or 0 for q in s.questions], default=0) + 1)
    q = SubtaskQuestion(subtask_id=subtask_id, project_id=project_id,
                         num=next_num, question_text=payload.question_text,
                         input_type=payload.input_type or "text")
    db.add(q); db.commit(); db.refresh(q)
    log_action(db, actor=current_user.name, action="add_question",
               description=f"Question added to subtask '{s.name}'",
               project_id=project_id, entity_type="subtask_question", entity_id=q.id,
               user_id=current_user.id)
    return _build_question(q)


@router.patch("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}/questions/{question_id}")
def update_question(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int, question_id: int,
    payload: SubtaskQuestionUpdate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    q = db.query(SubtaskQuestion).filter_by(id=question_id, subtask_id=subtask_id).first()
    if not q: raise HTTPException(404, "Question not found")
    data = payload.model_dump(exclude_none=True)
    for k in ("question_text", "input_type", "response"):
        if k in data:
            setattr(q, k, data[k])
    db.commit(); db.refresh(q)
    return _build_question(q)


@router.delete("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}/questions/{question_id}")
def delete_question(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int, question_id: int,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    q = db.query(SubtaskQuestion).filter_by(id=question_id, subtask_id=subtask_id).first()
    if not q: raise HTTPException(404, "Question not found")
    db.delete(q); db.commit()
    return {"status": "deleted"}


# ── Subtask Reports (optional, multiple per Subtask) ─────────────────────────
# Lets several "reports" — each identified by Report Number / Report Name /
# Department, plus Status / Assigned To / Due Date — point at the same
# Milestone -> Task -> Subtask, instead of needing a separate milestone
# structure built per report. report_number must be unique within a Subtask.
@router.post("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}/reports")
def add_report(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int,
    payload: ReportCreate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    s = db.query(CustomSubtask).filter_by(id=subtask_id, task_id=task_id).first()
    if not s: raise HTTPException(404, "Subtask not found")
    if not payload.report_number.strip():
        raise HTTPException(400, "Report Number is required")
    if not payload.report_name.strip():
        raise HTTPException(400, "Report Name is required")
    dup = db.query(SubtaskReport).filter_by(subtask_id=subtask_id, report_number=payload.report_number.strip()).first()
    if dup:
        raise HTTPException(400, f"Report Number '{payload.report_number}' already exists for this subtask")
    r = SubtaskReport(subtask_id=subtask_id, project_id=project_id,
                       report_number=payload.report_number.strip(), report_name=payload.report_name.strip(),
                       department=payload.department, status=payload.status or "Not Started",
                       assigned_to=payload.assigned_to, due_date=_parse_date(payload.due_date))
    db.add(r)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(400, f"Report Number '{payload.report_number}' already exists for this subtask")
    db.refresh(r)
    log_action(db, actor=current_user.name, action="add_report",
               description=f"Report '{r.report_name}' added to subtask '{s.name}'",
               project_id=project_id, entity_type="subtask_report", entity_id=r.id,
               user_id=current_user.id)
    return _build_report(r)


@router.patch("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}/reports/{report_id}")
def update_report(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int, report_id: int,
    payload: ReportUpdate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    r = db.query(SubtaskReport).filter_by(id=report_id, subtask_id=subtask_id).first()
    if not r: raise HTTPException(404, "Report not found")
    data = payload.model_dump(exclude_none=True)
    if "report_number" in data:
        new_num = data["report_number"].strip()
        if not new_num:
            raise HTTPException(400, "Report Number is required")
        dup = db.query(SubtaskReport).filter(SubtaskReport.subtask_id == subtask_id,
                                              SubtaskReport.report_number == new_num,
                                              SubtaskReport.id != report_id).first()
        if dup:
            raise HTTPException(400, f"Report Number '{new_num}' already exists for this subtask")
        r.report_number = new_num
    if "report_name" in data:
        if not data["report_name"].strip():
            raise HTTPException(400, "Report Name is required")
        r.report_name = data["report_name"].strip()
    for k in ("department", "status", "assigned_to"):
        if k in data:
            setattr(r, k, data[k])
    if "due_date" in data:
        r.due_date = _parse_date(data["due_date"])
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(400, f"Report Number '{r.report_number}' already exists for this subtask")
    db.refresh(r)
    return _build_report(r)


@router.delete("/{milestone_id}/tasks/{task_id}/subtasks/{subtask_id}/reports/{report_id}")
def delete_report(
    project_id: int, milestone_id: int, task_id: int, subtask_id: int, report_id: int,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    r = db.query(SubtaskReport).filter_by(id=report_id, subtask_id=subtask_id).first()
    if not r: raise HTTPException(404, "Report not found")
    db.delete(r); db.commit()
    return {"status": "deleted"}


# ── Milestone Reports (optional, multiple per Milestone) ─────────────────────
# Reports are now attached at Milestone level so they cover the whole
# milestone rather than a specific subtask. The same Report Number / Report
# Name / Department + tracking-field shape is kept (ReportCreate/ReportUpdate
# schemas are reused). report_number must be unique within a Milestone.

@router.get("/{milestone_id}/reports")
def list_milestone_reports(
    project_id: int, milestone_id: int,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    ms = db.query(CustomMilestone).filter_by(id=milestone_id, project_id=project_id).first()
    if not ms: raise HTTPException(404, "Milestone not found")
    return [_build_milestone_report(r) for r in sorted(ms.reports, key=lambda x: x.id)]


@router.post("/{milestone_id}/reports")
def add_milestone_report(
    project_id: int, milestone_id: int,
    payload: ReportCreate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    ms = db.query(CustomMilestone).filter_by(id=milestone_id, project_id=project_id).first()
    if not ms: raise HTTPException(404, "Milestone not found")
    if not payload.report_number.strip():
        raise HTTPException(400, "Report Number is required")
    if not payload.report_name.strip():
        raise HTTPException(400, "Report Name is required")
    dup = db.query(MilestoneReport).filter_by(milestone_id=milestone_id, report_number=payload.report_number.strip()).first()
    if dup:
        raise HTTPException(400, f"Report Number '{payload.report_number}' already exists for this milestone")
    r = MilestoneReport(
        milestone_id=milestone_id, project_id=project_id,
        report_number=payload.report_number.strip(), report_name=payload.report_name.strip(),
        department=payload.department, status=payload.status or "Not Started",
        assigned_to=payload.assigned_to, due_date=_parse_date(payload.due_date)
    )
    db.add(r)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(400, f"Report Number '{payload.report_number}' already exists for this milestone")
    db.refresh(r)
    log_action(db, actor=current_user.name, action="add_milestone_report",
               description=f"Report '{r.report_number} — {r.report_name}' added to milestone {ms.name}",
               project_id=project_id, entity_type="milestone_report", entity_id=r.id,
               user_id=current_user.id)
    return _build_milestone_report(r)


@router.patch("/{milestone_id}/reports/{report_id}")
def update_milestone_report(
    project_id: int, milestone_id: int, report_id: int,
    payload: ReportUpdate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    r = db.query(MilestoneReport).filter_by(id=report_id, milestone_id=milestone_id).first()
    if not r: raise HTTPException(404, "Report not found")
    old_assigned_to = r.assigned_to
    data = payload.model_dump(exclude_none=True)
    if "report_number" in data:
        new_num = data["report_number"].strip()
        if not new_num:
            raise HTTPException(400, "Report Number is required")
        dup = db.query(MilestoneReport).filter(
            MilestoneReport.milestone_id == milestone_id,
            MilestoneReport.report_number == new_num,
            MilestoneReport.id != report_id
        ).first()
        if dup:
            raise HTTPException(400, f"Report Number '{new_num}' already exists for this milestone")
        r.report_number = new_num
    if "report_name" in data:
        if not data["report_name"].strip():
            raise HTTPException(400, "Report Name is required")
        r.report_name = data["report_name"].strip()
    if "department" in data:   r.department  = data["department"]
    if "status" in data:       r.status      = data["status"]
    if "assigned_to" in data:  r.assigned_to = data["assigned_to"]
    if "due_date" in data:     r.due_date    = _parse_date(data["due_date"])
    try:
        db.flush()
        if r.assigned_to and r.assigned_to != old_assigned_to:
            _notify_assignee(db, project_id, "Report", r.report_name, r.assigned_to, r.due_date)
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(400, f"Report Number '{r.report_number}' already exists for this milestone")
    db.refresh(r)
    return _build_milestone_report(r)


@router.delete("/{milestone_id}/reports/{report_id}")
def delete_milestone_report(
    project_id: int, milestone_id: int, report_id: int,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    r = db.query(MilestoneReport).filter_by(id=report_id, milestone_id=milestone_id).first()
    if not r: raise HTTPException(404, "Report not found")
    db.delete(r); db.commit()
    return {"status": "deleted"}


# ── Project-level Working Hours summary (req 1d / req #2) ────────────────────
# Project total = sum of its (selected) Milestones' totals.
@router.get("/hours-summary")
def hours_summary(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    milestones = db.query(CustomMilestone).options(
        joinedload(CustomMilestone.tasks).joinedload(CustomTask.subtasks).joinedload(CustomSubtask.activities)
    ).filter_by(project_id=project_id, is_active=True).all()

    by_milestone = []
    total_est = total_act = 0.0
    for ms in milestones:
        est, act = _milestone_hours(db, ms)
        total_est += est
        total_act += act
        by_milestone.append({
            "milestone_id": ms.id, "num": ms.num, "name": ms.name,
            "estimated_hours": est, "actual_hours": act,
        })

    # Per-person consumed hours within this project (req 1d: "if a Task is
    # assigned to a person, consumed hours auto-add to that person's total").
    # Uses Actual Hours (hours_spent minus buffer), not raw hours_spent, so
    # this lines up with every other Actual Hours figure on this page.
    by_person: Dict[str, float] = {}
    rows = db.query(WorkHours.user_id, WorkHours.hours_spent, WorkHours.buffer_hours) \
             .filter(WorkHours.project_id == project_id).all()
    for user_id, hours_spent, buffer_hours in rows:
        user = db.query(User).filter_by(id=user_id).first()
        name = user.name if user else f"User {user_id}"
        by_person[name] = round(by_person.get(name, 0.0) + max((hours_spent or 0.0) - (buffer_hours or 0.0), 0.0), 2)

    return {
        "project_id": project_id,
        "estimated_hours": total_est,
        "actual_hours": total_act,
        "by_milestone": by_milestone,
        "by_person": by_person,
    }


# ── Mailbox helpers ──────────────────────────────────────────────────────────

def _dfmt(dt) -> str:
    if not dt: return ""
    try: return str(dt)[:10]
    except: return ""

def _xl_header(ws, headers, color="4F46E5"):
    fill = _Fill(start_color=color, end_color=color, fill_type="solid")
    font = _Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill; c.font = font
        c.alignment = _Align(horizontal="center")
    ws.freeze_panes = "A2"

def _generate_milestone_excel(db: Session, ms: CustomMilestone) -> bytes:
    wb = _xl.Workbook()
    # Sheet 1 — Summary
    ws1 = wb.active; ws1.title = "Milestone"
    _xl_header(ws1, ["Field", "Value"])
    for r, (f, v) in enumerate([
        ("Milestone #", f"M{ms.num:02d}"), ("Name", ms.name or ""),
        ("Status", ms.status or ""), ("Assignee", ms.assignee or ""),
        ("Planned Start", _dfmt(ms.planned_start)), ("Planned End", _dfmt(ms.planned_end)),
        ("Actual Start", _dfmt(ms.actual_start)), ("Actual End", _dfmt(ms.actual_end)),
        ("Description", ms.description or ""), ("Responsible", ms.responsible or ""),
    ], 2):
        ws1.cell(row=r, column=1, value=f).font = _Font(bold=True)
        ws1.cell(row=r, column=2, value=v)
    ws1.column_dimensions["A"].width = 20; ws1.column_dimensions["B"].width = 45
    # Sheet 2 — Tasks
    ws2 = wb.create_sheet("Tasks")
    _xl_header(ws2, ["Task #", "Name", "Status", "Assignee", "Planned Start", "Planned End", "Responsibility"])
    tasks = db.query(CustomTask).filter_by(milestone_id=ms.id).order_by(CustomTask.num).all()
    for r, t in enumerate(tasks, 2):
        for c, v in enumerate([f"T{t.num:02d}" if t.num else "", t.name or "", t.status or "",
                                t.assignee or "", _dfmt(t.planned_start), _dfmt(t.planned_end), t.responsibility or ""], 1):
            ws2.cell(row=r, column=c, value=v)
    for i, w in enumerate([10,30,15,20,15,15,20], 1):
        ws2.column_dimensions[ws2.cell(1,i).column_letter].width = w
    # Sheet 3 — Subtasks
    ws3 = wb.create_sheet("Subtasks")
    _xl_header(ws3, ["Task", "Subtask #", "Name", "Status", "Assignee", "Input Type", "Est. Hours"], "7C3AED")
    r3 = 2
    for t in tasks:
        for s in db.query(CustomSubtask).filter_by(task_id=t.id).order_by(CustomSubtask.num).all():
            for c, v in enumerate([t.name or "", f"S{s.num:02d}" if s.num else "", s.name or "",
                                    s.status or "", s.assignee or "", s.input_type or "", s.estimated_hours or 0], 1):
                ws3.cell(row=r3, column=c, value=v)
            r3 += 1
    for i, w in enumerate([25,12,30,15,20,15,12], 1):
        ws3.column_dimensions[ws3.cell(1,i).column_letter].width = w
    # Sheet 4 — Activities
    ws4 = wb.create_sheet("Activities")
    _xl_header(ws4, ["Task", "Subtask", "Activity", "Status", "Assignee", "Planned Start", "Planned End", "Est. Hours"], "0D3E7A")
    r4 = 2
    for t in tasks:
        for s in db.query(CustomSubtask).filter_by(task_id=t.id).order_by(CustomSubtask.num).all():
            for a in db.query(Activity).filter_by(subtask_id=s.id).all():
                for c, v in enumerate([t.name or "", s.name or "", a.name or "", a.status or "",
                                        a.assignee or "", _dfmt(a.planned_start), _dfmt(a.planned_end), a.estimated_hours or 0], 1):
                    ws4.cell(row=r4, column=c, value=v)
                r4 += 1
    for i, w in enumerate([25,25,30,15,20,15,15,12], 1):
        ws4.column_dimensions[ws4.cell(1,i).column_letter].width = w
    buf = _io.BytesIO(); wb.save(buf); return buf.getvalue()


def _generate_task_excel(db: Session, task: CustomTask, ms: CustomMilestone) -> bytes:
    wb = _xl.Workbook()
    # Sheet 1 — Summary
    ws1 = wb.active; ws1.title = "Task"
    _xl_header(ws1, ["Field", "Value"])
    ms_label = f"M{ms.num:02d} — {ms.name}" if ms else ""
    for r, (f, v) in enumerate([
        ("Milestone", ms_label), ("Task #", f"T{task.num:02d}" if task.num else ""),
        ("Name", task.name or ""), ("Status", task.status or ""),
        ("Assignee", task.assignee or ""), ("Planned Start", _dfmt(task.planned_start)),
        ("Planned End", _dfmt(task.planned_end)), ("Responsibility", task.responsibility or ""),
    ], 2):
        ws1.cell(row=r, column=1, value=f).font = _Font(bold=True)
        ws1.cell(row=r, column=2, value=v)
    ws1.column_dimensions["A"].width = 20; ws1.column_dimensions["B"].width = 45
    # Sheet 2 — Subtasks
    ws2 = wb.create_sheet("Subtasks")
    _xl_header(ws2, ["Subtask #", "Name", "Status", "Assignee", "Input Type", "Est. Hours"], "7C3AED")
    subs = db.query(CustomSubtask).filter_by(task_id=task.id).order_by(CustomSubtask.num).all()
    for r, s in enumerate(subs, 2):
        for c, v in enumerate([f"S{s.num:02d}" if s.num else "", s.name or "", s.status or "",
                                s.assignee or "", s.input_type or "", s.estimated_hours or 0], 1):
            ws2.cell(row=r, column=c, value=v)
    for i, w in enumerate([12,30,15,20,15,12], 1):
        ws2.column_dimensions[ws2.cell(1,i).column_letter].width = w
    # Sheet 3 — Activities
    ws3 = wb.create_sheet("Activities")
    _xl_header(ws3, ["Subtask", "Activity", "Status", "Assignee", "Planned Start", "Planned End", "Est. Hours"], "0D3E7A")
    r3 = 2
    for s in subs:
        for a in db.query(Activity).filter_by(subtask_id=s.id).all():
            for c, v in enumerate([s.name or "", a.name or "", a.status or "",
                                    a.assignee or "", _dfmt(a.planned_start), _dfmt(a.planned_end), a.estimated_hours or 0], 1):
                ws3.cell(row=r3, column=c, value=v)
            r3 += 1
    for i, w in enumerate([25,30,15,20,15,15,12], 1):
        ws3.column_dimensions[ws3.cell(1,i).column_letter].width = w
    buf = _io.BytesIO(); wb.save(buf); return buf.getvalue()


class MailboxPayload(BaseModel):
    to: List[str]
    note: str = ""


@router.post("/{milestone_id}/mailbox")
def send_milestone_mailbox(
    project_id: int, milestone_id: int, payload: MailboxPayload,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    ms = db.query(CustomMilestone).filter_by(id=milestone_id, project_id=project_id).first()
    if not ms: raise HTTPException(404, "Milestone not found")
    if not payload.to: raise HTTPException(400, "At least one recipient required")
    project = db.query(Project).filter_by(id=project_id).first()
    project_name = project.name if project else "—"
    excel_b64 = _b64.b64encode(_generate_milestone_excel(db, ms)).decode("utf-8")
    file_name = f"Milestone_M{ms.num:02d}_{(ms.name or 'details').replace(' ','_')}.xlsx"
    note_html = f"<p><strong>Note from {current_user.name}:</strong> {payload.note}</p><hr/>" if payload.note else ""
    tasks = db.query(CustomTask).filter_by(milestone_id=ms.id).order_by(CustomTask.num).all()
    task_rows = "".join(
        f"<tr><td style='padding:4px 8px;border:1px solid #e2e8f0'>T{t.num:02d}</td>"
        f"<td style='padding:4px 8px;border:1px solid #e2e8f0'>{t.name or ''}</td>"
        f"<td style='padding:4px 8px;border:1px solid #e2e8f0'>{t.assignee or '—'}</td>"
        f"<td style='padding:4px 8px;border:1px solid #e2e8f0'>{t.status or ''}</td></tr>"
        for t in tasks
    ) or "<tr><td colspan='4' style='padding:8px;text-align:center;color:#94a3b8;font-size:12px'>No tasks</td></tr>"
    body = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
      <div style="background:linear-gradient(135deg,#091525,#0f2448);padding:24px 28px;border-radius:12px 12px 0 0;text-align:center;">
        <h1 style="color:#fff;font-size:20px;margin:0;">AXON WBS</h1>
        <p style="color:#4a6080;font-size:11px;margin:4px 0 0;letter-spacing:.08em;">MILESTONE DETAILS</p>
      </div>
      <div style="padding:24px;background:#f8fafc;border:1px solid #e2e8f0;border-top:none;border-radius:0 0 12px 12px;">
        {note_html}
        <table style="width:100%;border-collapse:collapse;margin-bottom:16px;">
          <tr><td style="padding:5px 8px;border:1px solid #e2e8f0;font-weight:bold;background:#f1f5f9;width:35%">Project</td><td style="padding:5px 8px;border:1px solid #e2e8f0">{project_name}</td></tr>
          <tr><td style="padding:5px 8px;border:1px solid #e2e8f0;font-weight:bold;background:#f1f5f9">Milestone</td><td style="padding:5px 8px;border:1px solid #e2e8f0">M{ms.num:02d} — {ms.name or ''}</td></tr>
          <tr><td style="padding:5px 8px;border:1px solid #e2e8f0;font-weight:bold;background:#f1f5f9">Status</td><td style="padding:5px 8px;border:1px solid #e2e8f0">{ms.status or ''}</td></tr>
          <tr><td style="padding:5px 8px;border:1px solid #e2e8f0;font-weight:bold;background:#f1f5f9">Assignee</td><td style="padding:5px 8px;border:1px solid #e2e8f0">{ms.assignee or '—'}</td></tr>
          <tr><td style="padding:5px 8px;border:1px solid #e2e8f0;font-weight:bold;background:#f1f5f9">Planned Start</td><td style="padding:5px 8px;border:1px solid #e2e8f0">{_dfmt(ms.planned_start)}</td></tr>
          <tr><td style="padding:5px 8px;border:1px solid #e2e8f0;font-weight:bold;background:#f1f5f9">Planned End</td><td style="padding:5px 8px;border:1px solid #e