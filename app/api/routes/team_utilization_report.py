"""Team Utilization Report — per-person hours, billable split, and cost.

Sheet 1 — Summary (one row per team member):
  Name | Role | Total Hours | Billable Hours | Non-Billable Hours
  | Utilization % | Manpower Cost (₹) | # Projects

Sheet 2 — By Project (one row per user × project):
  Name | Role | Project | Client | Total Hours | Billable Hours
  | Utilization % | Manpower Cost (₹)

Row colour coding (both sheets):
  Green  (E2EFDA) — Utilization ≥ 80 %
  Amber  (FFF2CC) — Utilization 50 – 79 %
  Red    (FCE4D6) — Utilization < 50 % or no hours logged
"""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.database import get_db
from app.models.models import User, WorkHours, Project
from app.core.deps import get_current_user

router = APIRouter(prefix="/team-utilization-report", tags=["Team Utilization Report"])


# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL  = _fill("1F3864")
COL_FILL  = _fill("BDD7EE")
EVEN_FILL = _fill("EBF3FB")
ODD_FILL  = _fill("FFFFFF")
TOT_FILL  = _fill("D9E8F5")
GREEN_FILL = _fill("E2EFDA")
AMBER_FILL = _fill("FFF2CC")
RED_FILL   = _fill("FCE4D6")


def _util_fill(util):
    """Return fill based on utilization percentage (or '' if no hours)."""
    if util == "" or util is None:
        return RED_FILL
    if util >= 80:
        return GREEN_FILL
    if util >= 50:
        return AMBER_FILL
    return RED_FILL


def _pct(num, den, dec=1):
    if not den:
        return ""
    return round(float(num) / float(den) * 100, dec)


def _fmt(v, dec=2):
    if v is None:
        return 0.0
    return round(float(v), dec)


def _write_title(ws, title, subtitle, ncols):
    """Write title + optional subtitle; returns the next row number."""
    last_col = chr(ord("A") + ncols - 1)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    if subtitle:
        ws.merge_cells(f"A2:{last_col}2")
        c2 = ws["A2"]
        c2.value = subtitle
        c2.font = Font(italic=True, color="555555", size=9, name="Calibri")
        c2.fill = _fill("D9E8F5")
        c2.alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 16
        return 4   # header row
    return 3


def _write_header(ws, headers, hrow):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(hrow, col, h)
        cell.font = Font(bold=True, color="1F3864", size=9, name="Calibri")
        cell.fill = COL_FILL
        cell.border = _bdr()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hrow].height = 28


def _write_row(ws, row_num, values, fill, right_cols=None):
    right_cols = right_cols or []
    for col, val in enumerate(values, 1):
        cell = ws.cell(row_num, col, val if val not in (None, "") else "—")
        cell.font = Font(size=9, name="Calibri")
        cell.fill = fill
        cell.border = _bdr()
        align = "right" if (isinstance(val, (int, float)) or col in right_cols) else "left"
        cell.alignment = Alignment(horizontal=align)


# ── Aggregation ───────────────────────────────────────────────────────────────
def _aggregate(wh_rows, db):
    """
    Returns:
      user_summary: {uid: {user, total, billable, manpower_cost, project_ids}}
      user_proj:    {(uid, pid): {user, total, billable, manpower_cost}}
    """
    user_summary = {}
    user_proj    = {}
    rate_cache   = {}   # uid -> (user_obj, cost_rate)
    proj_cache   = {}   # pid -> project_obj

    for wh in wh_rows:
        uid = wh.user_id
        pid = wh.project_id
        if uid is None:
            continue

        if uid not in rate_cache:
            u = db.query(User).filter_by(id=uid).first()
            rate_cache[uid] = (u, float(u.cost_rate or 0) if u else 0.0)

        u_obj, rate = rate_cache[uid]
        hrs  = float(wh.hours_spent or 0)
        bill = (wh.work_type == 'Billable') if wh.work_type else bool(wh.is_billable)

        # ── per-user summary ──────────────────────────────────────────────────
        if uid not in user_summary:
            user_summary[uid] = {
                "user": u_obj, "total": 0.0, "billable": 0.0,
                "manpower_cost": 0.0, "project_ids": set(),
            }
        user_summary[uid]["total"]        += hrs
        user_summary[uid]["manpower_cost"] += hrs * rate
        if bill:
            user_summary[uid]["billable"] += hrs
        if pid:
            user_summary[uid]["project_ids"].add(pid)

        # ── per-user-project detail ───────────────────────────────────────────
        if pid:
            if pid not in proj_cache:
                proj_cache[pid] = db.query(Project).filter_by(id=pid).first()
            key = (uid, pid)
            if key not in user_proj:
                user_proj[key] = {
                    "user": u_obj, "project": proj_cache[pid],
                    "total": 0.0, "billable": 0.0, "manpower_cost": 0.0,
                }
            user_proj[key]["total"]        += hrs
            user_proj[key]["manpower_cost"] += hrs * rate
            if bill:
                user_proj[key]["billable"] += hrs

    return user_summary, user_proj


# ── Export ────────────────────────────────────────────────────────────────────
@router.get("/export")
def team_utilization_export(
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    role:        Optional[str] = None,
    user_id:     Optional[int] = None,
    project_id:  Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import date as date_cls

    # Build query
    q = db.query(WorkHours).join(User, WorkHours.user_id == User.id)
    if start_date:
        q = q.filter(WorkHours.date >= date_cls.fromisoformat(start_date))
    if end_date:
        q = q.filter(WorkHours.date <= date_cls.fromisoformat(end_date))
    if role:
        q = q.filter(User.role == role)
    if user_id:
        q = q.filter(WorkHours.user_id == user_id)
    if project_id:
        q = q.filter(WorkHours.project_id == project_id)

    wh_rows = q.all()
    user_summary, user_proj = _aggregate(wh_rows, db)

    # ── Subtitle ──────────────────────────────────────────────────────────────
    parts = []
    if start_date or end_date:
        parts.append(f"Period: {start_date or '—'} → {end_date or '—'}")
    if role:
        parts.append(f"Role: {role}")
    subtitle = " | ".join(parts)

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    HEADERS_S = [
        "Name", "Role",
        "Total Hours", "Billable Hours", "Non-Billable Hours",
        "Utilization %", "Manpower Cost (₹)", "# Projects",
    ]
    hrow = _write_title(ws1, "Team Utilization Report — Summary", subtitle, len(HEADERS_S))
    _write_header(ws1, HEADERS_S, hrow)

    # Sort by name
    sorted_users = sorted(user_summary.items(), key=lambda x: (x[1]["user"].name if x[1]["user"] else ""))

    grand_total = grand_bill = grand_cost = 0.0
    row_num = hrow + 1
    for uid, d in sorted_users:
        u    = d["user"]
        tot  = d["total"]
        bill = d["billable"]
        nbill = tot - bill
        util  = _pct(bill, tot)
        cost  = d["manpower_cost"]
        nproj = len(d["project_ids"])

        grand_total += tot
        grand_bill  += bill
        grand_cost  += cost

        util_str = f"{util}%" if util != "" else "—"
        bg = _util_fill(util if util != "" else 0)

        _write_row(ws1, row_num, [
            u.name if u else "—",
            u.role if u else "—",
            _fmt(tot), _fmt(bill), _fmt(nbill),
            util_str,
            _fmt(cost), nproj,
        ], bg, right_cols=[3, 4, 5, 7, 8])
        row_num += 1

    # Totals row
    if sorted_users:
        grand_util = _pct(grand_bill, grand_total)
        _write_row(ws1, row_num, [
            "TOTAL", "",
            _fmt(grand_total), _fmt(grand_bill), _fmt(grand_total - grand_bill),
            f"{grand_util}%" if grand_util != "" else "—",
            _fmt(grand_cost), "",
        ], TOT_FILL, right_cols=[3, 4, 5, 7])
        ws1.cell(row_num, 1).font = Font(bold=True, size=9, name="Calibri")

    # Empty state
    if not sorted_users:
        ws1.merge_cells(f"A{row_num}:H{row_num}")
        ws1.cell(row_num, 1, "No work hours found for the selected filters")
        ws1.cell(row_num, 1).font = Font(italic=True, size=9, color="999999")

    # Colour legend note
    note_row = row_num + 2
    ws1.cell(note_row, 1, "Colour key:")
    ws1.cell(note_row, 1).font = Font(bold=True, size=8, name="Calibri")
    for col, (label, fg) in enumerate([
        ("≥ 80 % Utilization", "E2EFDA"),
        ("50 – 79 %", "FFF2CC"),
        ("< 50 %", "FCE4D6"),
    ], 2):
        c = ws1.cell(note_row, col, label)
        c.font = Font(size=8, name="Calibri")
        c.fill = _fill(fg)
        c.border = _bdr()
        c.alignment = Alignment(horizontal="center")

    # Column widths Sheet 1
    for i, w in enumerate([22, 22, 13, 14, 16, 13, 18, 10], 1):
        ws1.column_dimensions[chr(ord("A") + i - 1)].width = w

    # ── Sheet 2: By Project ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Project")

    HEADERS_P = [
        "Name", "Role", "Project", "Client",
        "Total Hours", "Billable Hours", "Non-Billable Hours",
        "Utilization %", "Manpower Cost (₹)",
    ]
    hrow2 = _write_title(ws2, "Team Utilization Report — By Project", subtitle, len(HEADERS_P))
    _write_header(ws2, HEADERS_P, hrow2)

    # Sort by name then project
    sorted_proj = sorted(
        user_proj.items(),
        key=lambda x: (
            x[1]["user"].name if x[1]["user"] else "",
            x[1]["project"].name if x[1]["project"] else "",
        ),
    )

    row_num2 = hrow2 + 1
    for (uid, pid), d in sorted_proj:
        u   = d["user"]
        p   = d["project"]
        tot  = d["total"]
        bill = d["billable"]
        nbill = tot - bill
        util  = _pct(bill, tot)
        cost  = d["manpower_cost"]

        util_str = f"{util}%" if util != "" else "—"
        bg = _util_fill(util if util != "" else 0)

        _write_row(ws2, row_num2, [
            u.name if u else "—",
            u.role if u else "—",
            p.name if p else "—",
            p.client if p else "—",
            _fmt(tot), _fmt(bill), _fmt(nbill),
            util_str, _fmt(cost),
        ], bg, right_cols=[5, 6, 7, 9])
        row_num2 += 1

    if not sorted_proj:
        ws2.merge_cells(f"A{row_num2}:I{row_num2}")
        ws2.cell(row_num2, 1, "No project-level work hours found")
        ws2.cell(row_num2, 1).font = Font(italic=True, size=9, color="999999")

    # Column widths Sheet 2
    for i, w in enumerate([22, 22, 28, 20, 13, 14, 16, 13, 18], 1):
        ws2.column_dimensions[chr(ord("A") + i - 1)].width = w

    # ── Stream ────────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=team-utilization-report.xlsx"},
    )


# ── Filter options ────────────────────────────────────────────────────────────
@router.get("/filter-options")
def filter_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    users = (
        db.query(User)
        .filter(User.is_active == True)
        .order_by(User.name)
        .all()
    )
    projects = db.query(Project).order_by(Project.name).all()

    roles = sorted({u.role for u in users if u.role})

    return {
        "users":    [{"id": u.id, "name": u.name, "role": u.role} for u in users],
        "roles":    roles,
        "projects": [{"id": p.id, "name": p.name} for p in projects],
    }
