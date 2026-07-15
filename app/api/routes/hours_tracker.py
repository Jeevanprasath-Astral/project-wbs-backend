"""Hours Tracker — Tracking Hours Booked dashboard.

Summary and Excel export endpoints covering 8 hour categories:
  Leave | Holiday | Permission | Billable | Non-Billable | No Work | Training | R&D

Two views in every response / export sheet:
  1. Individual-wise — one row per user
  2. Project-wise   — one row per project (work-type hours only; leave/holiday/permission
                      are not project-scoped and are omitted from this view)

Leave hours calculation: approved leave days that overlap [start, end] × 8h/day.
Holiday hours calculation: holidays in [start, end] × 8h.
Permission hours: SUM of approved permission.hours in [start, end].
"""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import date as date_cls, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.database import get_db
from app.models.models import WorkHours, LeaveRequest, Permission, Holiday, User, Project
from app.core.deps import get_current_user

router = APIRouter(prefix="/hours-tracker", tags=["Hours Tracker"])

HOURS_PER_DAY = 8.0

WORK_TYPES = ['Billable', 'Non-Billable', 'No Work', 'Training', 'R&D']


# ── Date helpers ──────────────────────────────────────────────────────────────
def _resolve(start_date, end_date):
    today = date_cls.today()
    s = date_cls.fromisoformat(start_date) if start_date else today.replace(day=1)
    e = date_cls.fromisoformat(end_date)   if end_date   else today
    return s, e


def _leave_days_in_range(leave, start, end):
    """Calendar days the approved leave overlaps [start, end]."""
    overlap_start = max(leave.date_from, start)
    overlap_end   = min(leave.date_to,   end)
    if overlap_end < overlap_start:
        return 0
    return (overlap_end - overlap_start).days + 1


# ── Core aggregation ──────────────────────────────────────────────────────────
def _aggregate(db, start, end, user_id=None, project_id=None):
    """
    Returns:
      individual: list of dicts keyed by user_id
      project:    list of dicts keyed by project_id
      totals:     summary dict
    """
    # ── Work hours ────────────────────────────────────────────────────────────
    wh_q = db.query(WorkHours).filter(WorkHours.date >= start, WorkHours.date <= end)
    if user_id:    wh_q = wh_q.filter(WorkHours.user_id    == user_id)
    if project_id: wh_q = wh_q.filter(WorkHours.project_id == project_id)
    wh_rows = wh_q.all()

    # ── Leave (user-scoped, not project-scoped) ───────────────────────────────
    lv_q = db.query(LeaveRequest).filter(
        LeaveRequest.status    == 'Approved',
        LeaveRequest.date_from <= end,
        LeaveRequest.date_to   >= start,
    )
    if user_id: lv_q = lv_q.filter(LeaveRequest.user_id == user_id)
    leave_rows = lv_q.all()

    # ── Holidays (global) ─────────────────────────────────────────────────────
    holiday_count = db.query(func.count(Holiday.id)).filter(
        Holiday.date >= start, Holiday.date <= end
    ).scalar() or 0
    holiday_hours = holiday_count * HOURS_PER_DAY

    # ── Permissions (user-scoped) ─────────────────────────────────────────────
    perm_q = db.query(Permission).filter(
        Permission.status == 'Approved',
        Permission.date   >= start,
        Permission.date   <= end,
    )
    if user_id: perm_q = perm_q.filter(Permission.user_id == user_id)
    perm_rows = perm_q.all()

    # ── Build individual map ──────────────────────────────────────────────────
    ind = {}  # uid -> {user, metrics}

    def _ensure_user(uid):
        if uid not in ind:
            u = db.query(User).filter_by(id=uid).first()
            ind[uid] = {
                'user': u,
                'billable': 0.0, 'non_billable': 0.0,
                'no_work':  0.0, 'training':     0.0, 'rnd': 0.0,
                'leave': 0.0, 'permission': 0.0,
                'holiday': holiday_hours,
            }

    for wh in wh_rows:
        if wh.user_id is None: continue
        _ensure_user(wh.user_id)
        h  = float(wh.hours_spent or 0)
        wt = wh.work_type or ('Billable' if (wh.is_billable if wh.is_billable is not None else True) else 'Non-Billable')
        if   wt == 'Billable':     ind[wh.user_id]['billable']     += h
        elif wt == 'Non-Billable': ind[wh.user_id]['non_billable'] += h
        elif wt == 'No Work':      ind[wh.user_id]['no_work']       += h
        elif wt == 'Training':     ind[wh.user_id]['training']      += h
        elif wt == 'R&D':          ind[wh.user_id]['rnd']           += h

    for lv in leave_rows:
        if lv.user_id is None: continue
        _ensure_user(lv.user_id)
        days = _leave_days_in_range(lv, start, end)
        ind[lv.user_id]['leave'] += days * HOURS_PER_DAY

    for pr in perm_rows:
        if pr.user_id is None: continue
        _ensure_user(pr.user_id)
        ind[pr.user_id]['permission'] += float(pr.hours or 0)

    # ── Build project map ─────────────────────────────────────────────────────
    proj = {}  # pid -> {project, metrics}

    def _ensure_proj(pid):
        if pid not in proj:
            p = db.query(Project).filter_by(id=pid).first()
            proj[pid] = {
                'project': p,
                'billable': 0.0, 'non_billable': 0.0,
                'no_work':  0.0, 'training':     0.0, 'rnd': 0.0,
            }

    for wh in wh_rows:
        if wh.project_id is None: continue
        _ensure_proj(wh.project_id)
        h  = float(wh.hours_spent or 0)
        wt = wh.work_type or ('Billable' if (wh.is_billable if wh.is_billable is not None else True) else 'Non-Billable')
        if   wt == 'Billable':     proj[wh.project_id]['billable']     += h
        elif wt == 'Non-Billable': proj[wh.project_id]['non_billable'] += h
        elif wt == 'No Work':      proj[wh.project_id]['no_work']       += h
        elif wt == 'Training':     proj[wh.project_id]['training']      += h
        elif wt == 'R&D':          proj[wh.project_id]['rnd']           += h

    # ── Totals ────────────────────────────────────────────────────────────────
    def _sum_key(key): return round(sum(v[key] for v in ind.values()), 2)

    totals = {
        'leave':       _sum_key('leave'),
        'holiday':     holiday_hours,
        'permission':  _sum_key('permission'),
        'billable':    _sum_key('billable'),
        'non_billable':_sum_key('non_billable'),
        'no_work':     _sum_key('no_work'),
        'training':    _sum_key('training'),
        'rnd':         _sum_key('rnd'),
    }

    # ── Serialize ─────────────────────────────────────────────────────────────
    def _fmt(v): return round(float(v), 2)

    individual_list = []
    for uid, d in sorted(ind.items(), key=lambda x: (x[1]['user'].name if x[1]['user'] else '')):
        u = d['user']
        total_work = _fmt(d['billable'] + d['non_billable'] + d['no_work'] + d['training'] + d['rnd'])
        individual_list.append({
            'user_id':     uid,
            'user_name':   u.name  if u else '—',
            'role':        u.role  if u else '—',
            'billable':    _fmt(d['billable']),
            'non_billable':_fmt(d['non_billable']),
            'no_work':     _fmt(d['no_work']),
            'training':    _fmt(d['training']),
            'rnd':         _fmt(d['rnd']),
            'leave':       _fmt(d['leave']),
            'holiday':     _fmt(d['holiday']),
            'permission':  _fmt(d['permission']),
            'total_work':  total_work,
        })

    project_list = []
    for pid, d in sorted(proj.items(), key=lambda x: (x[1]['project'].name if x[1]['project'] else '')):
        p = d['project']
        total_work = _fmt(d['billable'] + d['non_billable'] + d['no_work'] + d['training'] + d['rnd'])
        project_list.append({
            'project_id':  pid,
            'project_name':p.name   if p else '—',
            'client':      p.client if p else '—',
            'billable':    _fmt(d['billable']),
            'non_billable':_fmt(d['non_billable']),
            'no_work':     _fmt(d['no_work']),
            'training':    _fmt(d['training']),
            'rnd':         _fmt(d['rnd']),
            'total_work':  total_work,
        })

    return {
        'individual': individual_list,
        'project':    project_list,
        'totals':     totals,
        'holiday_count': holiday_count,
        'period':     {'start': str(start), 'end': str(end)},
    }


# ── Summary endpoint ──────────────────────────────────────────────────────────
@router.get("/summary")
def hours_tracker_summary(
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    user_id:     Optional[int] = None,
    project_id:  Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    start, end = _resolve(start_date, end_date)
    return _aggregate(db, start, end, user_id, project_id)


# ── Excel export ──────────────────────────────────────────────────────────────
def _fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def _bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL  = _fill("1F3864")
COL_FILL  = _fill("BDD7EE")
EVEN_FILL = _fill("EBF3FB")
ODD_FILL  = _fill("FFFFFF")
TOT_FILL  = _fill("D9E8F5")


def _write_title(ws, title, subtitle, ncols):
    last = chr(ord("A") + ncols - 1)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    if subtitle:
        ws.merge_cells(f"A2:{last}2")
        c2 = ws["A2"]
        c2.value = subtitle
        c2.font = Font(italic=True, color="555555", size=9, name="Calibri")
        c2.fill = _fill("D9E8F5")
        c2.alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 16
        return 4
    return 3


def _write_header(ws, headers, hrow):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(hrow, col, h)
        cell.font = Font(bold=True, color="1F3864", size=9, name="Calibri")
        cell.fill = COL_FILL
        cell.border = _bdr()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hrow].height = 28


def _write_row(ws, row_num, values, fill):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row_num, col, val if val not in (None, "") else "—")
        cell.font = Font(size=9, name="Calibri")
        cell.fill = fill
        cell.border = _bdr()
        cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")


@router.get("/export")
def hours_tracker_export(
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    user_id:     Optional[int] = None,
    project_id:  Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    start, end = _resolve(start_date, end_date)
    data = _aggregate(db, start, end, user_id, project_id)
    subtitle = f"Period: {start} → {end}"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Individual-wise ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Individual-wise"
    IND_HEADERS = [
        "Employee Name", "Role",
        "Billable (h)", "Non-Billable (h)", "No Work (h)", "Training (h)", "R&D (h)",
        "Leave (h)", "Holiday (h)", "Permission (h)", "Total Work (h)",
    ]
    hrow = _write_title(ws1, "Hours Tracker — Individual-wise", subtitle, len(IND_HEADERS))
    _write_header(ws1, IND_HEADERS, hrow)

    row_num = hrow + 1
    for i, d in enumerate(data['individual']):
        bg = EVEN_FILL if i % 2 == 0 else ODD_FILL
        _write_row(ws1, row_num, [
            d['user_name'], d['role'],
            d['billable'], d['non_billable'], d['no_work'], d['training'], d['rnd'],
            d['leave'], d['holiday'], d['permission'], d['total_work'],
        ], bg)
        row_num += 1

    # Totals row
    t = data['totals']
    if data['individual']:
        total_work_sum = round(sum(d['total_work'] for d in data['individual']), 2)
        _write_row(ws1, row_num, [
            "TOTAL", "",
            t['billable'], t['non_billable'], t['no_work'], t['training'], t['rnd'],
            t['leave'], t['holiday'], t['permission'], total_work_sum,
        ], TOT_FILL)
        ws1.cell(row_num, 1).font = Font(bold=True, size=9, name="Calibri")

    if not data['individual']:
        ws1.cell(row_num, 1, "No data for the selected filters")
        ws1.cell(row_num, 1).font = Font(italic=True, size=9, color="999999")

    col_widths = [22, 20, 13, 15, 12, 12, 10, 12, 13, 13, 14]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[chr(ord("A") + i - 1)].width = w

    # ── Sheet 2: Project-wise ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Project-wise")
    PROJ_HEADERS = [
        "Project Name", "Client",
        "Billable (h)", "No Work (h)", "Training (h)", "R&D (h)",
        "Total Work (h)",
    ]
    hrow2 = _write_title(ws2, "Hours Tracker — Project-wise", subtitle, len(PROJ_HEADERS))
    _write_header(ws2, PROJ_HEADERS, hrow2)

    row_num2 = hrow2 + 1
    for i, d in enumerate(data['project']):
        bg = EVEN_FILL if i % 2 == 0 else ODD_FILL
        _write_row(ws2, row_num2, [
            d['project_name'], d['client'],
            d['billable'], d['no_work'], d['training'], d['rnd'],
            d['total_work'],
        ], bg)
        row_num2 += 1

    if data['project']:
        proj_total_work = round(sum(d['total_work'] for d in data['project']), 2)
        _write_row(ws2, row_num2, [
            "TOTAL", "",
            t['billable'], t['no_work'], t['training'], t['rnd'],
            proj_total_work,
        ], TOT_FILL)
        ws2.cell(row_num2, 1).font = Font(bold=True, size=9, name="Calibri")

    if not data['project']:
        ws2.cell(row_num2, 1, "No project-level work hours found")
        ws2.cell(row_num2, 1).font = Font(italic=True, size=9, color="999999")

    col_widths2 = [28, 20, 13, 12, 12, 10, 14]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[chr(ord("A") + i - 1)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=hours-tracker-{start}-to-{end}.xlsx"},
    )


@router.get("/daily")
def hours_tracker_daily(
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    user_id:     Optional[int] = None,
    project_id:  Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Day-by-day aggregated hours for calendar heatmap visualization."""
    start, end = _resolve(start_date, end_date)

    # Work hours grouped by date
    wh_q = (db.query(WorkHours.date, func.sum(WorkHours.hours_spent))
            .filter(WorkHours.date >= start, WorkHours.date <= end))
    if user_id:    wh_q = wh_q.filter(WorkHours.user_id    == user_id)
    if project_id: wh_q = wh_q.filter(WorkHours.project_id == project_id)
    wh_by_date = {
        str(d): round(float(h or 0), 2)
        for d, h in wh_q.group_by(WorkHours.date).all()
    }

    # Holiday dates in range
    holiday_dates = {
        str(h.date)
        for h in db.query(Holiday).filter(
            Holiday.date >= start, Holiday.date <= end
        ).all()
    }

    # Build full date array
    dates = []
    cur = start
    while cur <= end:
        d_str = str(cur)
        dates.append({
            'date':       d_str,
            'hours':      wh_by_date.get(d_str, 0.0),
            'is_holiday': d_str in holiday_dates,
            'is_weekend': cur.weekday() >= 5,   # Sat=5, Sun=6
        })
        cur += timedelta(days=1)

    return {
        'dates':  dates,
        'period': {'start': str(start), 'end': str(end)},
    }
